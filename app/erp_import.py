from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import webview

REQUIRED_COLUMNS = ("제품코드", "제품명", "매출일자", "매출부수", "매출금액")
ALL_COLUMNS = (
    "브랜드", "계열", "시리즈", "제품코드", "제품명", "매출일자", "정가", "첫출고일",
    "출고부수", "출고금액", "반품부수", "반품금액", "매출부수", "매출금액", "입고부수", "기증부수", "재고",
)
INTEGER_COLUMNS = {
    "출고부수", "출고금액", "반품부수", "반품금액", "매출부수", "매출금액", "입고부수", "기증부수", "재고"
}
COMPARE_COLUMNS = (
    "브랜드", "계열", "시리즈", "제품명", "정가", "첫출고일",
    "출고부수", "출고금액", "반품부수", "반품금액", "매출부수", "매출금액", "입고부수", "기증부수", "재고",
)


def _date_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip().replace("/", "-").replace(".", "-")
    if not text:
        return None
    parts = [p for p in text.split("-") if p]
    if len(parts) >= 3:
        try:
            return date(int(parts[0]), int(parts[1]), int(parts[2])).isoformat()
        except ValueError:
            pass
    return text


def _number(value: Any, integer: bool = False) -> int | float | None:
    if value in (None, ""):
        return 0 if integer else None
    try:
        number = float(str(value).replace(",", ""))
        return int(round(number)) if integer else number
    except (TypeError, ValueError):
        return 0 if integer else None


def _find_header_row(ws) -> tuple[int, list[str]]:
    for row_no, values in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        headers = [str(x).strip() if x is not None else "" for x in values]
        if all(name in headers for name in REQUIRED_COLUMNS):
            return row_no, headers
    raise RuntimeError("필수 열을 찾을 수 없습니다: " + ", ".join(REQUIRED_COLUMNS))


def _normalize_for_compare(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return str(value).strip()


def choose_and_import_erp(backend: Any, expected_product_code: str | None = None) -> dict[str, Any]:
    """ERP 도서별 일별 매출 파일에서 제품코드를 자동 감지해 Supabase에 병합합니다.

    파일은 원칙적으로 한 도서만 포함해야 합니다. 기존 날짜는 값이 같으면 건너뛰고,
    값이 달라졌으면 수정하며, 처음 보는 날짜만 신규 추가합니다.
    expected_product_code는 하위 호환용이며 전달될 경우 일치 여부만 추가 검증합니다.
    """
    try:
        if not backend.db:
            raise RuntimeError("Supabase가 연결되지 않았습니다.")
        if not webview.windows:
            raise RuntimeError("파일 선택 창을 열 수 없습니다.")

        selected = webview.windows[0].create_file_dialog(
            webview.OPEN_DIALOG,
            allow_multiple=False,
            file_types=("Excel 파일 (*.xlsx)",),
        )
        if not selected:
            return {"ok": False, "cancelled": True, "message": "파일 선택을 취소했습니다."}

        path = Path(selected[0])
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = workbook[workbook.sheetnames[0]]
        header_row, headers = _find_header_row(ws)

        parsed_rows: list[dict[str, Any]] = []
        skipped = 0
        product_codes: set[str] = set()
        product_names: set[str] = set()

        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            source = dict(zip(headers, values))
            code = str(source.get("제품코드") or "").strip()
            sales_date = _date_text(source.get("매출일자"))
            if not code or not sales_date:
                skipped += 1
                continue

            product_codes.add(code)
            name = str(source.get("제품명") or "").strip()
            if name:
                product_names.add(name)

            row: dict[str, Any] = {}
            for col in ALL_COLUMNS:
                value = source.get(col)
                if col in {"매출일자", "첫출고일"}:
                    row[col] = _date_text(value)
                elif col in INTEGER_COLUMNS:
                    row[col] = _number(value, integer=True)
                elif col == "정가":
                    row[col] = _number(value, integer=False)
                elif value is None:
                    row[col] = None
                else:
                    row[col] = str(value).strip()
            row["제품코드"] = code
            row["매출일자"] = sales_date
            row["원본파일명"] = path.name
            parsed_rows.append(row)

        if not product_codes:
            raise RuntimeError("제품코드가 있는 ERP 일별 데이터가 없습니다. 소계/총계만 있는 파일인지 확인해 주세요.")
        if len(product_codes) != 1:
            found = ", ".join(sorted(product_codes))
            raise RuntimeError(f"한 파일에 여러 제품코드가 있습니다: {found}. ERP에서 도서 1권만 조회한 파일을 업로드해 주세요.")

        product_code = next(iter(product_codes))
        if expected_product_code and product_code != str(expected_product_code).strip():
            raise RuntimeError(f"업로드 파일의 제품코드는 {product_code}이며 선택된 도서 제품코드 {expected_product_code}와 다릅니다.")
        if not parsed_rows:
            raise RuntimeError("업로드할 ERP 일별 데이터가 없습니다.")

        # 제품인덱스에 존재하는 코드인지 확인해 잘못된 파일 연결을 방지합니다.
        product_match = (
            backend.db.client.table("제품인덱스")
            .select("제품코드,제품명")
            .eq("제품코드", product_code)
            .limit(1)
            .execute()
        ).data or []
        if not product_match:
            raise RuntimeError(f"제품코드 {product_code}를 제품인덱스에서 찾을 수 없습니다.")
        product_name = product_match[0].get("제품명") or (next(iter(product_names)) if product_names else product_code)

        min_date = min(r["매출일자"] for r in parsed_rows)
        max_date = max(r["매출일자"] for r in parsed_rows)
        existing_rows = (
            backend.db.client.table("ERP일별판매실적")
            .select("*")
            .eq("제품코드", product_code)
            .gte("매출일자", min_date)
            .lte("매출일자", max_date)
            .execute()
        ).data or []
        existing_map = {str(r.get("매출일자")): r for r in existing_rows if r.get("매출일자")}

        new_rows: list[dict[str, Any]] = []
        changed_rows: list[dict[str, Any]] = []
        unchanged = 0
        now = datetime.now().isoformat()

        for row in parsed_rows:
            old = existing_map.get(row["매출일자"])
            if not old:
                row["수정일시"] = now
                new_rows.append(row)
                continue

            changed = any(
                _normalize_for_compare(old.get(col)) != _normalize_for_compare(row.get(col))
                for col in COMPARE_COLUMNS
            )
            if changed:
                row["수정일시"] = now
                changed_rows.append(row)
            else:
                unchanged += 1

        rows_to_upsert = new_rows + changed_rows
        batch_size = 500
        for start in range(0, len(rows_to_upsert), batch_size):
            backend.db.client.table("ERP일별판매실적").upsert(
                rows_to_upsert[start:start + batch_size],
                on_conflict="제품코드,매출일자",
            ).execute()

        return {
            "ok": True,
            "file_name": path.name,
            "product_code": product_code,
            "product_name": product_name,
            "total": len(parsed_rows),
            "inserted": len(new_rows),
            "updated": len(changed_rows),
            "unchanged": unchanged,
            "skipped": skipped,
            "date_from": min_date,
            "date_to": max_date,
        }
    except Exception as exc:
        return {"ok": False, "message": str(exc)}
