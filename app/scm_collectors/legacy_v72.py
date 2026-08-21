# -*- coding: utf-8 -*-
r"""
Kyobo SCM downloader V30
- Runs from the folder where this .py file is saved.
- Reads Kyobo login info from scm_login.xlsx in the same folder.
- Automatically logs in, selects MiraeN(I-Seum) 0123910, opens Sales Info > Sales Search.
- Shows a small date range window, then downloads daily Excel files.

Required files in the same folder:
- kyobo_scm_download_v19_runtime.py
- run_kyobo_scm_download_v12.bat
- scm_login.xlsx
"""

import os
import re
import time
import shutil
import traceback
from pathlib import Path
from datetime import datetime, timedelta

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.common.exceptions import TimeoutException


VERSION = "V72-KYOBO-STALE-DATE-GRID-GUARD"

SCRIPT_DIR = Path(os.environ.get("KYOBO_BASE_DIR", Path(__file__).resolve().parent)).resolve()
BASE_DIR = SCRIPT_DIR.parent.parent
DOWNLOAD_DIR = SCRIPT_DIR / "download"

# V65 path-separated setup:
# - BAT/result/template folder:  ...\05. 영업 실적\01. 실판매
# - Python code folder:          ...\05. 영업 실적\90. 파이썬 코드
# - Login master data folder:    ...\05. 영업 실적\80. Master Data
MASTER_DATA_DIR = SCRIPT_DIR.parent / "80. Master Data"
LOGIN_XLSX = MASTER_DATA_DIR / "scm_login.xlsx"

LOGIN_URL = "https://scm.kyobobook.co.kr/scm/login.action"
SALES_URL = "https://scm.kyobobook.co.kr/scm/page.action?pageID=saleStockInfo"

PARTNER_KEYWORDS = ["아이세움", "0123910"]


def log(msg: str):
    print(msg, flush=True)


def ensure_dirs():
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)


def build_driver():
    ensure_dirs()
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": str(DOWNLOAD_DIR),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1,

        # Prevent Chrome's own password manager / leaked-password warning bubble.
        # This is not a Kyobo SCM popup, so Selenium cannot close it as a normal web element.
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "profile.password_manager_leak_detection": False,
        "password_manager_enabled": False,
        "password_manager_leak_detection": False,
        "credentials_enable_autosignin": False,
        "autofill.profile_enabled": False,
        "autofill.credit_card_enabled": False,
    }
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--start-maximized")
    options.add_argument("--disable-save-password-bubble")
    options.add_argument("--disable-features=PasswordLeakDetection,PasswordManagerOnboarding,PasswordCheck,AutofillServerCommunication")
    options.add_argument("--disable-password-generation")
    options.add_argument("--disable-popup-blocking")
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(60)
    return driver


def visible(el):
    try:
        return el.is_displayed() and el.size.get("width", 0) > 0 and el.size.get("height", 0) > 0
    except Exception:
        return False




def element_summary(driver, el):
    """Return a short, safe description of a Selenium element for logging only.

    This helper must never stop the automation.  Earlier integrated versions
    referenced element_summary from the Youngpoong search/download logging path
    but did not define it, so the run stopped after opening the sales page.
    """
    try:
        info = driver.execute_script(
            """
            const e = arguments[0];
            if (!e) return {};
            const r = e.getBoundingClientRect ? e.getBoundingClientRect() : {left:0, top:0, width:0, height:0};
            return {
              tag: e.tagName || '',
              id: e.id || '',
              cls: e.className || '',
              text: (e.innerText || e.value || e.title || '').trim().slice(0, 120),
              left: Math.round(r.left),
              top: Math.round(r.top),
              width: Math.round(r.width),
              height: Math.round(r.height)
            };
            """,
            el,
        )
        return str(info)
    except Exception as e:
        try:
            return f"tag={el.tag_name}, text={(el.text or '')[:80]}, error={e}"
        except Exception:
            return f"<element summary unavailable: {e}>"


def dump_debug(driver, tag: str):
    try:
        html_path = BASE_DIR / f"debug_kyobo_{tag}.html"
        png_path = BASE_DIR / f"debug_kyobo_{tag}.png"
        html_path.write_text(driver.page_source, encoding="utf-8", errors="ignore")
        driver.save_screenshot(str(png_path))
        log(f"Debug saved: {html_path}")
        log(f"Debug saved: {png_path}")
    except Exception as e:
        log(f"Could not save debug files: {e}")



def find_login_xlsx():
    r"""
    Find scm_login.xlsx.
    V65 path-separated priority:
    1) ...\05. 영업 실적\80. Master Data\scm_login.xlsx
    2) BAT/result folder
    3) Current working folder
    4) Same folder as the physical Python file
    5) Nearby fallback paths
    """
    physical_py_dir = Path(__file__).resolve().parent

    candidates = [
        MASTER_DATA_DIR / "scm_login.xlsx",
        SCRIPT_DIR / "scm_login.xlsx",
        Path.cwd() / "scm_login.xlsx",
        physical_py_dir / "scm_login.xlsx",
    ]

    candidates.extend(SCRIPT_DIR.glob("*/scm_login.xlsx"))
    candidates.extend(physical_py_dir.glob("*/scm_login.xlsx"))
    candidates.append(SCRIPT_DIR.parent / "scm_login.xlsx")
    candidates.append(physical_py_dir.parent / "scm_login.xlsx")

    checked = []
    seen = set()
    for p in candidates:
        try:
            p = p.resolve()
        except Exception:
            pass
        key = str(p)
        if key in seen:
            continue
        seen.add(key)
        checked.append(key)
        if p.exists():
            return p

    raise FileNotFoundError("Login file not found. Checked paths:\n" + "\n".join(checked))

def _normalize_header_name(value):
    text = str(value or "").strip().lower()
    text = text.replace(" ", "").replace("_", "").replace("-", "")
    if text in ("id", "아이디", "로그인id", "로그인아이디", "계정", "계정id", "userid", "user"):
        return "id"
    if text in ("password", "pw", "pwd", "비밀번호", "패스워드", "암호"):
        return "password"
    if text in ("scm", "사이트", "구분", "site", "siteurl", "url", "주소"):
        return text
    return text


def read_kyobo_login_info():
    """
    Reads Kyobo login info from scm_login.xlsx in the same folder as this script.

    Supported header examples:
    - scm / site / id / password
    - 구분 / 사이트 / 아이디 / 비밀번호
    """
    login_xlsx_path = find_login_xlsx()
    log(f"Using login file: {login_xlsx_path}")

    wb = load_workbook(login_xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    col_map = {}

    for idx, row in enumerate(rows):
        normalized = [_normalize_header_name(x) for x in row]
        if "id" in normalized and "password" in normalized:
            header_row_idx = idx
            for c_idx, name in enumerate(normalized):
                if name:
                    col_map[name] = c_idx
            break

    if header_row_idx is None:
        raise RuntimeError("Could not find header row containing id/password or 아이디/비밀번호 in scm_login.xlsx")

    id_col = col_map.get("id")
    pw_col = col_map.get("password")

    possible_site_cols = [
        col_map.get("site"),
        col_map.get("사이트"),
        col_map.get("siteurl"),
        col_map.get("url"),
        col_map.get("주소"),
    ]
    site_col = next((x for x in possible_site_cols if x is not None), None)

    for row in rows[header_row_idx + 1:]:
        row_text = " ".join(str(x) for x in row if x is not None)
        is_kyobo = ("교보" in row_text) or ("kyobo" in row_text.lower()) or ("scm.kyobobook" in row_text.lower())

        # If there is only one data row, allow it even if Kyobo text is missing.
        non_empty_data_rows = [r for r in rows[header_row_idx + 1:] if any(x is not None and str(x).strip() for x in r)]
        if len(non_empty_data_rows) == 1:
            is_kyobo = True

        if not is_kyobo:
            continue

        login_id = row[id_col] if id_col is not None and id_col < len(row) else None
        login_pw = row[pw_col] if pw_col is not None and pw_col < len(row) else None
        login_url = row[site_col] if site_col is not None and site_col < len(row) and row[site_col] else LOGIN_URL

        if login_id is None or login_pw is None:
            raise RuntimeError("Kyobo row found, but id/password is empty in scm_login.xlsx")

        return str(login_id).strip(), str(login_pw).strip(), str(login_url).strip()

    raise RuntimeError("Could not find Kyobo login row in scm_login.xlsx")

def set_text_input(el, value):
    el.click()
    time.sleep(0.1)
    try:
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.BACKSPACE)
        el.send_keys(value)
    except Exception:
        pass




def set_input_value(driver, el, value):
    """Safely set an input value and fire the events that web pages listen for.

    Youngpoong date fields need both the DOM value change and input/change/blur
    events.  This helper intentionally has the 3-argument signature used by
    yp_set_date_fields(driver, element, value).
    """
    value = str(value)
    try:
        driver.execute_script(
            """
            const el = arguments[0];
            const value = arguments[1];
            el.focus();
            const proto = Object.getPrototypeOf(el);
            const desc = Object.getOwnPropertyDescriptor(proto, 'value');
            if (desc && desc.set) { desc.set.call(el, value); }
            else { el.value = value; }
            for (const type of ['input','change','keyup','blur']) {
              try { el.dispatchEvent(new Event(type, {bubbles:true})); } catch(e) {}
            }
            """,
            el,
            value,
        )
        time.sleep(0.15)
    except Exception:
        set_text_input(el, value)
        try:
            driver.execute_script(
                "for (const type of ['input','change','keyup','blur']) { arguments[0].dispatchEvent(new Event(type,{bubbles:true})); }",
                el,
            )
        except Exception:
            pass

def auto_login_kyobo(driver):
    login_id, login_pw, login_url = read_kyobo_login_info()

    log(f"Opening Kyobo login page: {login_url}")
    driver.get(login_url)

    wait = WebDriverWait(driver, 25)

    # Find visible login fields. Kyobo login screen has one text input and one password input.
    wait.until(lambda d: len([x for x in d.find_elements(By.CSS_SELECTOR, "input") if visible(x)]) >= 2)

    inputs = [x for x in driver.find_elements(By.CSS_SELECTOR, "input") if visible(x)]
    password_inputs = [x for x in inputs if (x.get_attribute("type") or "").lower() == "password"]
    text_inputs = [
        x for x in inputs
        if (x.get_attribute("type") or "").lower() in ("text", "", "email", "tel", "number")
    ]

    if not password_inputs or not text_inputs:
        dump_debug(driver, "login_fields_not_found")
        raise RuntimeError("Could not find login id/password fields.")

    id_el = text_inputs[0]
    pw_el = password_inputs[0]

    log("Typing Kyobo login id/password from scm_login.xlsx.")
    set_text_input(id_el, login_id)
    set_text_input(pw_el, login_pw)

    # Kyobo SCM login button is not a normal <button>.
    # It is currently:
    # <a id="btn_login" class="w2anchor2 btn_login_blue" href="javascript:void(null);">로그인</a>
    # So V20 first targets #btn_login exactly, then falls back to class/text/Enter.
    time.sleep(0.5)
    clicked = False

    def click_login_element(el, label):
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", el)
            time.sleep(0.2)
        except Exception:
            pass

        # WebSquare/anchor buttons sometimes require browser-side mouse events.
        js_click = """
        const el = arguments[0];
        try { el.focus(); } catch(e) {}
        try { el.scrollIntoView({block:'center', inline:'center'}); } catch(e) {}
        const opts = {bubbles:true, cancelable:true, view:window};
        try { el.dispatchEvent(new MouseEvent('mouseover', opts)); } catch(e) {}
        try { el.dispatchEvent(new MouseEvent('mousedown', opts)); } catch(e) {}
        try { el.dispatchEvent(new MouseEvent('mouseup', opts)); } catch(e) {}
        try { el.dispatchEvent(new MouseEvent('click', opts)); } catch(e) {}
        try { el.click(); } catch(e) {}
        return true;
        """

        try:
            el.click()
            log(f"Clicked login button by {label} with Selenium click.")
            return True
        except Exception:
            pass

        try:
            driver.execute_script(js_click, el)
            log(f"Clicked login button by {label} with JavaScript mouse-event click.")
            return True
        except Exception:
            pass

        try:
            ActionChains(driver).move_to_element(el).click().perform()
            log(f"Clicked login button by {label} with ActionChains click.")
            return True
        except Exception:
            return False

    # 1) Exact Kyobo login button id. This is the most important fix.
    try:
        btn = WebDriverWait(driver, 10).until(
            lambda d: d.find_element(By.ID, "btn_login")
        )
        clicked = click_login_element(btn, "id=btn_login")
    except Exception as e:
        log(f"Could not click exact id=btn_login immediately: {e}")

    # 2) Direct CSS fallbacks for the same anchor/class.
    if not clicked:
        direct_selectors = [
            "#btn_login",
            "a#btn_login",
            "a.btn_login_blue",
            ".btn_login_blue",
            "[id='btn_login']",
            "[class*='btn_login']",
        ]
        for sel in direct_selectors:
            try:
                els = driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if click_login_element(el, sel):
                        clicked = True
                        break
                if clicked:
                    break
            except Exception:
                pass

    # 3) Browser-side querySelector fallback. This works even when Selenium's text/visibility is strange.
    if not clicked:
        try:
            clicked = bool(driver.execute_script(
                """
                const selectors = ['#btn_login','a#btn_login','a.btn_login_blue','.btn_login_blue','[id="btn_login"]','[class*="btn_login"]'];
                const opts = {bubbles:true, cancelable:true, view:window};
                for (const s of selectors) {
                    const el = document.querySelector(s);
                    if (el) {
                        try { el.focus(); } catch(e) {}
                        try { el.scrollIntoView({block:'center', inline:'center'}); } catch(e) {}
                        try { el.dispatchEvent(new MouseEvent('mouseover', opts)); } catch(e) {}
                        try { el.dispatchEvent(new MouseEvent('mousedown', opts)); } catch(e) {}
                        try { el.dispatchEvent(new MouseEvent('mouseup', opts)); } catch(e) {}
                        try { el.dispatchEvent(new MouseEvent('click', opts)); } catch(e) {}
                        try { el.click(); } catch(e) {}
                        return true;
                    }
                }
                return false;
                """
            ))
            if clicked:
                log("Clicked login button by JavaScript querySelector #btn_login fallback.")
        except Exception:
            pass

    # 4) Existing visible-text scan, expanded to id/class/value/href.
    if not clicked:
        candidates = []
        for sel in ["button", "a", "input[type='button']", "input[type='submit']", "div", "span"]:
            for el in driver.find_elements(By.CSS_SELECTOR, sel):
                try:
                    text = (el.text or el.get_attribute("value") or "").strip()
                    el_id = (el.get_attribute("id") or "").strip()
                    el_class = (el.get_attribute("class") or "").strip()
                    el_href = (el.get_attribute("href") or "").strip()
                    merged = f"{text} {el_id} {el_class} {el_href}".lower()
                    is_login = ("로그인" in text) or ("btn_login" in merged) or ("btn_login_blue" in merged)
                    if is_login:
                        rect = driver.execute_script(
                            "const r=arguments[0].getBoundingClientRect(); return {x:r.x,y:r.y,w:r.width,h:r.height};",
                            el,
                        )
                        area = (rect.get("w") or 1) * (rect.get("h") or 1)
                        candidates.append((area, el, text or el_id or el_class or el_href, rect))
                except Exception:
                    pass

        if candidates:
            candidates.sort(key=lambda x: x[0], reverse=True)
            login_btn = candidates[0][1]
            log(f"Clicking login button candidate: {candidates[0][2]!r}")
            clicked = click_login_element(login_btn, "expanded candidate scan")

    # 5) Final fallback: press Enter in the password field.
    if not clicked:
        try:
            pw_el.send_keys(Keys.ENTER)
            clicked = True
            log("Pressed Enter in password field as login fallback.")
        except Exception:
            pass

    if not clicked:
        dump_debug(driver, "login_button_not_found")
        raise RuntimeError("Could not find Kyobo login button. Tried id=btn_login, .btn_login_blue, text scan, and Enter fallback.")

    # Wait until logged in: top menu or logout appears.
    try:
        WebDriverWait(driver, 30).until(
            lambda d: ("로그아웃" in (d.execute_script("return document.body ? document.body.innerText : ''") or ""))
            or ("판매정보" in (d.execute_script("return document.body ? document.body.innerText : ''") or ""))
        )
        log("Kyobo login appears successful.")
    except TimeoutException:
        dump_debug(driver, "after_login_timeout")
        raise RuntimeError("Login did not complete within timeout.")


def close_password_change_popup(driver, max_wait=8):
    """
    Close only the real post-login password-change modal.

    V32 was too broad: because the page title contains "판매[주문] 확인", it could
    treat the normal sales page as a popup and click the first book/result area.
    That opened the 판매확인 detail popup.  V33 only closes a popup when the visible
    modal/window text contains password-change words.
    """
    log("Checking password-change popup only...")
    end = time.time() + max_wait
    password_keywords = ["비밀번호 변경", "비밀번호를 변경", "비밀번호가 정보 유출", "비밀번호"]

    while time.time() < end:
        try:
            alert = driver.switch_to.alert
            text = alert.text or ""
            if any(k in text for k in password_keywords):
                alert.accept()
                log(f"Closed browser password alert: {text[:80]}")
                time.sleep(0.8)
                return True
        except Exception:
            pass

        try:
            result = driver.execute_script(
                """
                return (function(){
                    function visible(el){
                        const r = el.getBoundingClientRect();
                        const st = window.getComputedStyle(el);
                        return r.width > 0 && r.height > 0 && st.display !== 'none' && st.visibility !== 'hidden' && r.bottom >= 0 && r.top <= innerHeight;
                    }
                    function text(el){ return (el.innerText || el.textContent || el.value || '').trim(); }
                    const roots = Array.from(document.querySelectorAll('[role="dialog"], .w2window, .w2popup, .w2modal, div, section'))
                        .filter(el => visible(el));
                    const modalRoots = roots.filter(el => {
                        const t = text(el);
                        if (!(t.includes('비밀번호 변경') || t.includes('비밀번호를 변경') || t.includes('비밀번호가 정보 유출') || t.includes('비밀번호'))) return false;
                        const r = el.getBoundingClientRect();
                        // Real password popup is a centered box, not the whole page.
                        return r.width >= 180 && r.width <= 900 && r.height >= 100 && r.height <= 700;
                    });
                    if (!modalRoots.length) return {ok:false, reason:'no password modal'};
                    modalRoots.sort((a,b) => {
                        const ra=a.getBoundingClientRect(), rb=b.getBoundingClientRect();
                        return (ra.width*ra.height) - (rb.width*rb.height);
                    });
                    const root = modalRoots[0];
                    const buttons = Array.from(root.querySelectorAll('button,a,input,div,span'))
                        .filter(el => visible(el) && (text(el) === '확인' || text(el).toLowerCase() === 'ok' || (text(el).includes('확인') && el.getBoundingClientRect().width < 200)));
                    if (!buttons.length) return {ok:false, reason:'password modal found but no confirm', popupText:text(root).slice(0,120)};
                    buttons.sort((a,b) => {
                        const ra=a.getBoundingClientRect(), rb=b.getBoundingClientRect();
                        return (rb.top - ra.top);
                    });
                    const btn = buttons[0];
                    try { btn.click(); } catch(e) {
                        const r = btn.getBoundingClientRect();
                        btn.dispatchEvent(new MouseEvent('click', {bubbles:true, cancelable:true, view:window, clientX:r.left+r.width/2, clientY:r.top+r.height/2}));
                    }
                    return {ok:true, text:text(btn), popupText:text(root).slice(0,120)};
                })();
                """
            )
            if result and result.get("ok"):
                log(f"Closed password-change popup: {result}")
                time.sleep(1.0)
                return True
        except Exception as e:
            log(f"Password popup check failed: {e}")

        time.sleep(0.4)

    log("No password-change popup detected.")
    return False


def get_current_partner_label(driver):
    """Return the actually selected partner label text from the WebSquare selectbox."""
    try:
        return driver.execute_script(
            """
            const el = document.getElementById('s_vndrList_label');
            return el ? (el.innerText || el.textContent || '').trim() : '';
            """
        ) or ""
    except Exception:
        return ""


def is_iseum_selected(driver):
    text = get_current_partner_label(driver)
    return ("아이세움" in text and "0123910" in text)


def try_websquare_component_select_iseum(driver):
    """Try selecting the vendor through WebSquare's own component API."""
    try:
        result = driver.execute_script(
            """
            return (function(){
                const TARGET_TEXT = '아이세움';
                const TARGET_CODE = '0123910';
                const ids = ['s_vndrList', 'sbx_vndrList', 'vndrList'];

                function getComp(id){
                    try { if (window.$p && $p.getComponentById) return $p.getComponentById(id); } catch(e) {}
                    try { if (window.WebSquare && WebSquare.util && WebSquare.util.getComponentById) return WebSquare.util.getComponentById(id); } catch(e) {}
                    try { if (window[ id ]) return window[id]; } catch(e) {}
                    return null;
                }
                function call(obj, names, args){
                    for (const n of names){
                        try { if (obj && typeof obj[n] === 'function') return obj[n].apply(obj, args || []); } catch(e) {}
                    }
                    return undefined;
                }
                function normalize(v){ return (v === null || v === undefined) ? '' : String(v).trim(); }
                function fireDomChange(){
                    const label = document.getElementById('s_vndrList_label');
                    const root = document.getElementById('s_vndrList') || (label ? label.closest('[id]') : null);
                    for (const el of [root, label].filter(Boolean)){
                        try { el.dispatchEvent(new Event('input', {bubbles:true})); } catch(e) {}
                        try { el.dispatchEvent(new Event('change', {bubbles:true})); } catch(e) {}
                    }
                }

                const log = [];
                for (const id of ids){
                    const comp = getComp(id);
                    if (!comp) { log.push(id + ': no component'); continue; }
                    log.push(id + ': component found');

                    // First: inspect items by index.
                    let count = call(comp, ['getItemCount', 'getItemLength', 'getLength'], []);
                    count = Number(count);
                    if (!Number.isFinite(count) || count <= 0 || count > 200) count = 20;

                    for (let i=0; i<count; i++){
                        const text = normalize(call(comp, ['getItemText', 'getItemLabel', 'getText'], [i]));
                        const value = normalize(call(comp, ['getItemValue', 'getValue'], [i]));
                        const item = normalize(call(comp, ['getItem', 'getOption'], [i]));
                        const merged = [text, value, item].join(' ');
                        if (merged.includes(TARGET_TEXT) && merged.includes(TARGET_CODE)){
                            log.push('target item index=' + i + ' text=' + text + ' value=' + value);
                            // Try several common WebSquare selectbox APIs.
                            // IMPORTANT:
                            // On Kyobo WebSquare, getItemValue/getValue(i) can incorrectly return
                            // the CURRENT vendor value (ex: 0123906) even when the target item's label
                            // is I-Seum 0123910. V23 called setValue(value) and accidentally changed
                            // the component back to Danhaengbon. V24 selects by INDEX first and never
                            // overwrites it with a wrong value unless the value itself is 0123910.
                            call(comp, ['setSelectedIndex'], [i]);
                            call(comp, ['setIndex'], [i]);
                            call(comp, ['select'], [i]);
                            call(comp, ['setSelectedItem'], [i]);
                            if (value && value.includes(TARGET_CODE)) {
                                call(comp, ['setValue', 'setSelectedValue'], [value]);
                            }
                            call(comp, ['trigger', 'fireEvent'], ['change']);
                            fireDomChange();
                            const labelAfterIndex = normalize((document.getElementById('s_vndrList_label') || {}).innerText);
                            if (labelAfterIndex.includes(TARGET_TEXT) && labelAfterIndex.includes(TARGET_CODE)) {
                                return {ok:true, method:'websquare-index-no-value-overwrite', id, index:i, text, value, labelAfterIndex, log};
                            }
                            log.push('index selected but label still=' + labelAfterIndex);
                        }
                    }

                    // Second: try the known vendor code directly.
                    call(comp, ['setValue', 'setSelectedValue'], [TARGET_CODE]);
                    call(comp, ['trigger', 'fireEvent'], ['change']);
                    fireDomChange();
                    const label = normalize((document.getElementById('s_vndrList_label') || {}).innerText);
                    if (label.includes(TARGET_TEXT) && label.includes(TARGET_CODE)){
                        return {ok:true, method:'websquare-setValue-code', id, value:TARGET_CODE, label, log};
                    }
                }
                return {ok:false, log};
            })();
            """
        )
        log(f"WebSquare partner select result: {result}")
        time.sleep(1.5)
        return bool(result and result.get("ok")) or is_iseum_selected(driver)
    except Exception as e:
        log(f"WebSquare component partner select failed: {e}")
        return False


def try_dom_click_select_iseum(driver):
    """Try selecting I-Seum by opening the WebSquare rendered dropdown and clicking the option."""
    try:
        # Click the exact selected-label area the user found: s_vndrList_label.
        label = driver.find_element(By.ID, "s_vndrList_label")
        driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", label)
        time.sleep(0.2)
        for target in [label, label.find_element(By.XPATH, "./ancestor::*[contains(@class,'w2selectbox') or @id='s_vndrList'][1]") if True else label]:
            try:
                target.click()
                break
            except Exception:
                try:
                    driver.execute_script("arguments[0].click();", target)
                    break
                except Exception:
                    pass
        time.sleep(1.0)

        # First try a browser-side click on the exact visible rendered option.
        # WebSquare selectbox options are often rendered as rows/divs outside the original component.
        js_result = driver.execute_script(
            """
            return (function(){
                const TARGET_TEXT = '아이세움';
                const TARGET_CODE = '0123910';
                function visible(el){
                    const r = el.getBoundingClientRect();
                    const st = getComputedStyle(el);
                    return r.width > 10 && r.height > 8 && st.visibility !== 'hidden' && st.display !== 'none' && r.bottom >= 0 && r.top <= innerHeight;
                }
                function score(el){
                    const txt = (el.innerText || el.textContent || '').trim();
                    const r = el.getBoundingClientRect();
                    let s = 0;
                    if (txt.includes(TARGET_TEXT)) s += 100;
                    if (txt.includes(TARGET_CODE)) s += 100;
                    if (r.width < 500) s += 10;
                    if (r.height < 80) s += 10;
                    if ((el.id || '').includes('s_vndrList')) s += 5;
                    return s;
                }
                const all = Array.from(document.querySelectorAll('div,td,tr,li,a,span'))
                    .filter(el => {
                        const txt = (el.innerText || el.textContent || '').trim();
                        return txt.includes(TARGET_TEXT) && txt.includes(TARGET_CODE) && visible(el);
                    })
                    .sort((a,b) => score(b)-score(a));
                for (const el of all) {
                    const r = el.getBoundingClientRect();
                    if (r.width > 900 || r.height > 150) continue;
                    try { el.scrollIntoView({block:'center', inline:'center'}); } catch(e) {}
                    const cx = Math.max(1, Math.floor(r.left + Math.min(r.width - 2, Math.max(2, r.width/2))));
                    const cy = Math.max(1, Math.floor(r.top + Math.min(r.height - 2, Math.max(2, r.height/2))));
                    const target = document.elementFromPoint(cx, cy) || el;
                    for (const type of ['mouseover','mousedown','mouseup','click']) {
                        try { target.dispatchEvent(new MouseEvent(type, {bubbles:true, cancelable:true, view:window, clientX:cx, clientY:cy})); } catch(e) {}
                        try { el.dispatchEvent(new MouseEvent(type, {bubbles:true, cancelable:true, view:window, clientX:cx, clientY:cy})); } catch(e) {}
                    }
                    try { target.click(); } catch(e) {}
                    try { el.click(); } catch(e) {}
                    return {ok:true, text:(el.innerText||el.textContent||'').trim(), tag:el.tagName, id:el.id, cls:el.className, rect:{left:r.left,top:r.top,width:r.width,height:r.height}};
                }
                return {ok:false, count:all.length};
            })();
            """
        )
        log(f"Browser-side visible I-Seum option click result: {js_result}")
        time.sleep(1.5)
        if is_iseum_selected(driver):
            log("Selected partner by browser-side visible option click.")
            return True

        # Visible WebSquare options often appear outside the selectbox root as a floating table/div.
        opt_candidates = driver.find_elements(
            By.XPATH,
            "//*[contains(normalize-space(.), '아이세움') and contains(normalize-space(.), '0123910')]"
        )
        for opt in opt_candidates:
            try:
                if not visible(opt):
                    continue
                # Avoid clicking the whole body/header by preferring small option-like elements.
                rect = driver.execute_script(
                    "const r=arguments[0].getBoundingClientRect(); return {w:r.width,h:r.height,top:r.top,left:r.left};", opt
                )
                if rect["w"] > 800 or rect["h"] > 120:
                    continue
                driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", opt)
                time.sleep(0.2)
                try:
                    ActionChains(driver).move_to_element(opt).click().perform()
                except Exception:
                    driver.execute_script("arguments[0].click();", opt)
                time.sleep(2.0)
                if is_iseum_selected(driver):
                    log("Selected partner by clicking visible WebSquare option.")
                    return True
            except Exception:
                continue
        return is_iseum_selected(driver)
    except Exception as e:
        log(f"DOM click partner select failed: {e}")
        return False


def select_iseum_partner(driver):
    """
    Selects MiraeN(I-Seum) 0123910 from the top WebSquare partner dropdown.
    The actually selected value is shown in #s_vndrList_label.
    """
    close_password_change_popup(driver, max_wait=6)
    log("Selecting partner: MiraeN(I-Seum) 0123910")

    current = get_current_partner_label(driver)
    log(f"Current partner label before selection: {current}")
    if is_iseum_selected(driver):
        log("I-Seum partner is already selected.")
        return

    # 1) WebSquare component API is the most reliable for this custom selectbox.
    if try_websquare_component_select_iseum(driver):
        current = get_current_partner_label(driver)
        log(f"Current partner label after WebSquare selection: {current}")
        if is_iseum_selected(driver):
            return

    # 2) DOM click fallback using the exact label id supplied by the user.
    if try_dom_click_select_iseum(driver):
        current = get_current_partner_label(driver)
        log(f"Current partner label after DOM-click selection: {current}")
        if is_iseum_selected(driver):
            return

    # 3) Last fallback: old native select approach, in case the page changes.
    selects = [s for s in driver.find_elements(By.TAG_NAME, "select") if visible(s)]
    for sel in selects:
        try:
            options = sel.find_elements(By.TAG_NAME, "option")
            for idx, opt in enumerate(options):
                text = (opt.text or opt.get_attribute("text") or "").strip()
                value = (opt.get_attribute("value") or "").strip()
                merged = f"{text} {value}"
                if all(k in merged for k in PARTNER_KEYWORDS):
                    driver.execute_script(
                        """
                        const sel = arguments[0];
                        const idx = arguments[1];
                        sel.selectedIndex = idx;
                        sel.dispatchEvent(new Event('input', {bubbles:true}));
                        sel.dispatchEvent(new Event('change', {bubbles:true}));
                        """,
                        sel,
                        idx,
                    )
                    time.sleep(2.0)
                    if is_iseum_selected(driver):
                        log(f"Selected partner by native select option: {text}")
                        return
        except Exception as e:
            log(f"Native select attempt failed: {e}")

    current = get_current_partner_label(driver)
    log(f"Current partner label after all attempts: {current}")
    dump_debug(driver, "select_partner_failed")
    raise RuntimeError("Could not automatically select I-Seum partner. Please check debug screenshot/html.")


def ensure_iseum_partner_on_current_page(driver, context=""):
    """
    Kyobo can reset the top vendor dropdown when moving to the sales page.
    Therefore, verify and, if needed, re-select I-Seum on the CURRENT page,
    immediately before setting dates/searching.
    """
    label = get_current_partner_label(driver)
    prefix = f"[{context}] " if context else ""
    log(f"{prefix}Partner label check: {label}")
    if is_iseum_selected(driver):
        return

    log(f"{prefix}Partner is not I-Seum. Re-selecting I-Seum on current page.")
    select_iseum_partner(driver)
    time.sleep(2.0)

    label = get_current_partner_label(driver)
    log(f"{prefix}Partner label after re-selection: {label}")
    if not is_iseum_selected(driver):
        dump_debug(driver, "partner_reset_or_not_selected")
        raise RuntimeError("I-Seum partner is not selected on the current page. Current label: " + label)


def open_sales_search_page(driver):
    """
    Opens Sales Info > Sales Search page.
    First try the direct page URL. If it fails, click the top menu.
    """
    log("Opening Sales Search page.")

    # Direct URL is more stable after login/partner selection.
    try:
        driver.get(SALES_URL)
        time.sleep(3.5)
        if driver.execute_script("return !!document.getElementById('btn_search');"):
            log("Sales Search page opened by direct URL.")
            return
    except Exception as e:
        log(f"Direct Sales URL failed: {e}")

    # Fallback: menu click.
    try:
        sales_info = None
        for el in driver.find_elements(By.XPATH, "//*[contains(normalize-space(.), '판매정보')]"):
            if visible(el):
                sales_info = el
                break

        if sales_info:
            ActionChains(driver).move_to_element(sales_info).pause(0.5).perform()
            time.sleep(0.8)
            try:
                sales_info.click()
            except Exception:
                pass
            time.sleep(0.8)

        sales_search_candidates = driver.find_elements(By.XPATH, "//*[contains(normalize-space(.), '판매조회')]")
        for el in sales_search_candidates:
            if visible(el):
                try:
                    el.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", el)
                time.sleep(3.5)
                if driver.execute_script("return !!document.getElementById('btn_search');"):
                    log("Sales Search page opened by menu click.")
                    return
    except Exception as e:
        log(f"Menu navigation failed: {e}")

    dump_debug(driver, "sales_page_not_found")
    raise RuntimeError("Could not open Sales Search page automatically.")


def find_date_inputs(driver):
    inputs = driver.find_elements(By.CSS_SELECTOR, "input")
    candidates = []
    for el in inputs:
        try:
            if not visible(el):
                continue
            value = (el.get_attribute("value") or "").strip()
            typ = (el.get_attribute("type") or "").lower()
            rect = driver.execute_script(
                "const r=arguments[0].getBoundingClientRect(); return {w:r.width,h:r.height,top:r.top,left:r.left};",
                el,
            )
            if typ in ("text", "", "date") and rect["w"] >= 60 and rect["h"] >= 15:
                if re.match(r"^\d{4}-\d{2}-\d{2}$", value) or value == "":
                    candidates.append(el)
        except Exception:
            continue
    if len(candidates) < 2:
        raise RuntimeError("Could not find two visible date input fields.")
    candidates.sort(key=lambda e: driver.execute_script(
        "const r=arguments[0].getBoundingClientRect(); return [Math.round(r.top), Math.round(r.left)];", e
    ))
    return candidates[0], candidates[1]


def set_one_date_field(driver, el, date_text):
    driver.execute_script(
        """
        const el = arguments[0];
        const val = arguments[1];
        el.removeAttribute('readonly');
        el.removeAttribute('disabled');
        el.focus();
        el.value = '';
        el.dispatchEvent(new Event('input', {bubbles:true}));
        el.value = val;
        el.dispatchEvent(new Event('input', {bubbles:true}));
        el.dispatchEvent(new Event('change', {bubbles:true}));
        el.dispatchEvent(new KeyboardEvent('keyup', {bubbles:true, key:'Enter'}));
        el.blur();
        """,
        el,
        date_text,
    )


def set_date_fields(driver, date_text):
    start_el, end_el = find_date_inputs(driver)
    set_one_date_field(driver, start_el, date_text)
    time.sleep(0.3)
    set_one_date_field(driver, end_el, date_text)
    time.sleep(0.7)

    values = driver.execute_script(
        """
        const inputs = Array.from(document.querySelectorAll('input'))
          .filter(x => x.offsetParent !== null)
          .map(x => x.value || '');
        return inputs;
        """
    )
    log(f"Visible input values after date set: {values[:8]}")



def visible_result_row_info(driver, partner=None):
    """Return visible result-row evidence from the current Kyobo page.

    V33 fix: the JS is passed as a raw Python string and uses join(' ') so it
    does not create an invalid JavaScript string token.  It also uses body-text
    evidence, because the Kyobo grid is WebSquare markup and may not expose rows
    as normal <tr> elements.
    """
    partner_name = partner.get("name", "") if isinstance(partner, dict) else ""
    partner_code = partner.get("code", "") if isinstance(partner, dict) else ""
    try:
        return driver.execute_script(
            r"""
            return (function(partnerName, partnerCode){
                function norm(v){ return (v === null || v === undefined) ? '' : String(v).replace(/\s+/g,' ').trim(); }
                const bodyText = document.body ? norm(document.body.innerText || '') : '';
                const isbnMatches = bodyText.match(/97[89][0-9]{10}/g) || [];
                const rowLike = Array.from(document.querySelectorAll('tr, .w2grid_body_row, .w2grid_row, [id*=row], [id*=data]')).filter(el => {
                    const r = el.getBoundingClientRect();
                    if (r.width <= 0 || r.height <= 0) return false;
                    const t = norm(el.innerText || el.textContent || '');
                    return /97[89][0-9]{10}/.test(t);
                });
                const rowTexts = rowLike.map(r => norm(r.innerText || r.textContent || ''));
                const sampleSource = rowTexts.length ? rowTexts.join(' ') : bodyText;
                const hasPartnerName = partnerName ? sampleSource.includes(partnerName) || bodyText.includes(partnerName) : false;
                const hasPartnerCode = partnerCode ? bodyText.includes(partnerCode) : false;
                const hasIseum = sampleSource.includes('아이세움') || bodyText.includes('아이세움');
                const hasBookfolio = sampleSource.includes('북폴리오') || bodyText.includes('북폴리오');
                return {
                    ok:true,
                    isbnCount:isbnMatches.length,
                    rowCount:rowLike.length,
                    hasPartnerName:hasPartnerName,
                    hasPartnerCode:hasPartnerCode,
                    hasIseum:hasIseum,
                    hasBookfolio:hasBookfolio,
                    sample: (rowTexts.length ? rowTexts : bodyText.split(/\s+(?=97[89][0-9]{10})/)).slice(0,5).map(t => String(t).slice(0,180))
                };
            })(arguments[0], arguments[1]);
            """,
            partner_name,
            partner_code,
        ) or {"ok": False, "isbnCount": 0, "rowCount": 0, "sample": []}
    except Exception as e:
        return {"ok": False, "isbnCount": 0, "rowCount": 0, "sample": [], "error": str(e)}


def page_has_visible_result_rows(driver):
    info = visible_result_row_info(driver)
    try:
        return int(info.get("isbnCount") or 0) > 0 or int(info.get("rowCount") or 0) > 0
    except Exception:
        return False


def page_has_current_partner_result_rows(driver, partner):
    info = visible_result_row_info(driver, partner)
    try:
        name = partner.get("name", "")
        code = partner.get("code", "")
        has_isbn_rows = int(info.get("isbnCount") or 0) > 0 or int(info.get("rowCount") or 0) > 0

        # V34 fix:
        # Kyobo's WebSquare grid does not always expose I-Seum rows to DOM text
        # immediately, even when the rows are visible on screen and Excel download
        # is available. V33 therefore blocked the child-sheet download when
        # isbnCount/rowCount was 0.
        #
        # Keep the stale-grid protection for Bookfolio, but allow I-Seum when
        # the current partner label/code is correct and no Bookfolio evidence is
        # present. This prevents I-Seum rows from being skipped while still
        # avoiding the earlier mistake of copying I-Seum rows into the adult sheet.
        if name == "아이세움":
            if bool(info.get("hasBookfolio")):
                return False
            if bool(info.get("hasIseum")) or (code and bool(info.get("hasPartnerCode"))):
                return True
            return has_isbn_rows and bool(info.get("hasPartnerName"))

        if name == "북폴리오":
            if bool(info.get("hasIseum")) and not bool(info.get("hasBookfolio")):
                return False
            if has_isbn_rows and (bool(info.get("hasBookfolio")) or (code and bool(info.get("hasPartnerCode")))):
                return True
            return False

        if has_isbn_rows and code and info.get("hasPartnerCode"):
            return True
        return has_isbn_rows and bool(info.get("hasPartnerName"))
    except Exception:
        return False


def current_grid_is_stale_for_partner(driver, partner):
    """True when ISBN rows exist, but they look like the previous partner's rows."""
    if not page_has_visible_result_rows(driver):
        return False
    return not page_has_current_partner_result_rows(driver, partner)


def no_data_text_present(driver):
    try:
        page_text = driver.execute_script("return document.body ? document.body.innerText : ''; ") or ""
    except Exception:
        page_text = ""
    return ("조회된 내역이 없습니다" in page_text) or ("조회된 내역이 없습니다." in page_text)


def click_popup_confirm_button(driver):
    """Click a visible Kyobo '확인' button. Returns True when a click was attempted."""
    try:
        result = driver.execute_script(
            """
            return (function(){
                function visible(el){
                    const r = el.getBoundingClientRect();
                    const st = getComputedStyle(el);
                    return r.width > 20 && r.height > 10 && st.display !== 'none' && st.visibility !== 'hidden';
                }
                function txt(el){ return (el.innerText || el.textContent || el.value || '').trim(); }
                const els = Array.from(document.querySelectorAll('a,button,input,div,span'))
                  .filter(el => visible(el) && txt(el) === '확인')
                  .map(el => { const r=el.getBoundingClientRect(); return {el, top:r.top, left:r.left, w:r.width, h:r.height, z:parseInt(getComputedStyle(el).zIndex)||0}; });
                if (!els.length) return {ok:false, reason:'no visible confirm'};
                els.sort((a,b) => (b.z-a.z) || (Math.abs(a.top-innerHeight/2)-Math.abs(b.top-innerHeight/2)));
                const target = els[0].el;
                target.click();
                const r = target.getBoundingClientRect();
                return {ok:true, text:txt(target), top:r.top, left:r.left};
            })();
            """
        )
        if result and result.get("ok"):
            log(f"Popup confirm clicked by JS: {result}")
            time.sleep(0.8)
            return True
    except Exception as e:
        log(f"Popup confirm JS click failed: {e}")

    candidates = driver.find_elements(
        By.XPATH,
        "//*[self::a or self::button or self::input or self::div or self::span][normalize-space(.)='확인' or @value='확인']"
    )
    for el in candidates:
        try:
            if not visible(el):
                continue
            rect = driver.execute_script(
                "const r=arguments[0].getBoundingClientRect(); return {w:r.width,h:r.height,top:r.top,left:r.left};",
                el,
            )
            if rect["w"] < 20 or rect["h"] < 10:
                continue
            driver.execute_script("arguments[0].click();", el)
            time.sleep(0.8)
            log("Popup confirm clicked by Selenium fallback.")
            return True
        except Exception:
            continue
    return False


def close_no_data_popup_but_keep_data(driver, partner=None, context=""):
    """Close Kyobo no-data popup only when rows match the current partner.

    Returns True if a false no-data popup was closed and it is safe to keep going.
    Returns False when no popup exists or the grid is stale/empty for the current partner.
    """
    if not no_data_text_present(driver):
        return False
    info = visible_result_row_info(driver, partner)
    log(f"No-data popup text detected during {context}. Current-partner row evidence: {info}")
    if partner is not None and not page_has_current_partner_result_rows(driver, partner):
        log("No-data popup belongs to current search, or grid is stale from previous partner. Closing popup and NOT using visible rows.")
        click_popup_confirm_button(driver)
        return False
    if not page_has_visible_result_rows(driver):
        return False
    clicked = click_popup_confirm_button(driver)
    if clicked:
        log("No-data popup was ignored because result grid has current partner rows. Continuing.")
    return clicked




def install_browser_activity_tracker(driver):
    """Install a lightweight XHR/fetch tracker so we can wait for real page activity instead of fixed sleeps."""
    try:
        return driver.execute_script(
            r"""
            return (function(){
                try {
                    if (!window.__scmAjaxTrackerInstalled) {
                        window.__scmAjaxTrackerInstalled = true;
                        window.__scmPendingAjax = 0;
                        window.__scmLastAjaxDone = Date.now();
                        const oldOpen = XMLHttpRequest.prototype.open;
                        const oldSend = XMLHttpRequest.prototype.send;
                        XMLHttpRequest.prototype.open = function(){ this.__scmTracked = true; return oldOpen.apply(this, arguments); };
                        XMLHttpRequest.prototype.send = function(){
                            try { window.__scmPendingAjax = Math.max(0, (window.__scmPendingAjax || 0) + 1); } catch(e) {}
                            try { this.addEventListener('loadend', function(){ window.__scmPendingAjax = Math.max(0, (window.__scmPendingAjax || 0) - 1); window.__scmLastAjaxDone = Date.now(); }); } catch(e) {}
                            return oldSend.apply(this, arguments);
                        };
                        if (window.fetch) {
                            const oldFetch = window.fetch;
                            window.fetch = function(){
                                try { window.__scmPendingAjax = Math.max(0, (window.__scmPendingAjax || 0) + 1); } catch(e) {}
                                return oldFetch.apply(this, arguments).finally(function(){
                                    try { window.__scmPendingAjax = Math.max(0, (window.__scmPendingAjax || 0) - 1); window.__scmLastAjaxDone = Date.now(); } catch(e) {}
                                });
                            };
                        }
                    }
                    window.__scmActionStart = Date.now();
                    return {ok:true, pending:window.__scmPendingAjax||0};
                } catch(e) { return {ok:false, error:String(e)}; }
            })();
            """
        )
    except Exception as e:
        return {"ok": False, "error": str(e)}


def browser_activity_info(driver):
    """Return whether the current page still looks busy/loading."""
    try:
        return driver.execute_script(
            r"""
            return (function(){
                function visible(el){
                    const r = el.getBoundingClientRect();
                    const st = getComputedStyle(el);
                    return r.width > 2 && r.height > 2 && st.display !== 'none' && st.visibility !== 'hidden' && r.bottom >= 0 && r.top <= innerHeight;
                }
                const busyEls = Array.from(document.querySelectorAll('*')).filter(el => {
                    const s = ((el.id||'') + ' ' + (el.className||'') + ' ' + (el.getAttribute('role')||'')).toLowerCase();
                    if (!(s.includes('loading') || s.includes('loader') || s.includes('progress') || s.includes('spinner') || s.includes('mask') || s.includes('blockui') || s.includes('modal'))) return false;
                    if (!visible(el)) return false;
                    const txt = (el.innerText || el.textContent || '').trim();
                    // Do not treat the normal page header as a modal/busy element.
                    if (txt.length > 2000) return false;
                    return true;
                }).slice(0, 5).map(el => ({id:el.id, cls:String(el.className).slice(0,80), text:(el.innerText||el.textContent||'').trim().slice(0,80)}));
                const pending = window.__scmPendingAjax || 0;
                const lastDone = window.__scmLastAjaxDone || 0;
                const actionStart = window.__scmActionStart || 0;
                const now = Date.now();
                return {ok:true, pending:pending, busyEls:busyEls, busy:pending > 0 || busyEls.length > 0, quietMs: now - Math.max(lastDone, actionStart)};
            })();
            """
        ) or {"ok": False, "busy": False, "pending": 0, "quietMs": 9999}
    except Exception as e:
        return {"ok": False, "busy": False, "pending": 0, "quietMs": 9999, "error": str(e)}


def wait_until_activity_quiet(driver, min_quiet_ms=600, max_wait=10, label="page activity"):
    """Wait until ajax/loading overlays appear to be finished, with a max timeout fallback."""
    end = time.time() + max_wait
    last = None
    while time.time() < end:
        info = browser_activity_info(driver)
        last = info
        if not info.get("busy") and int(info.get("quietMs") or 0) >= min_quiet_ms:
            return True, info
        time.sleep(0.25)
    return False, last or {}

def click_confirm_if_popup(driver, partner=None):
    if not no_data_text_present(driver):
        return False

    info = visible_result_row_info(driver, partner)
    if partner is not None and page_has_current_partner_result_rows(driver, partner):
        log(f"No-data popup detected, but grid has CURRENT partner rows. Row evidence: {info}. Closing popup and continuing.")
        click_popup_confirm_button(driver)
        return False

    if page_has_visible_result_rows(driver):
        log(f"No-data popup detected with stale rows for another partner. Row evidence: {info}. Closing popup and skipping this partner/date.")
        click_popup_confirm_button(driver)
        return True

    log("No-data popup/text detected and no result rows are visible. Trying to click confirm and skip this date...")
    if click_popup_confirm_button(driver):
        log("Confirm clicked. Skip this date.")
        return True

    log("No-data popup detected but confirm could not be clicked automatically. Please click OK manually, then press ENTER.")
    input()
    return True


def click_search_button(driver, ymd: str):
    install_browser_activity_tracker(driver)
    log("Clicking btn_search by exact id. Excel will NOT be clicked before this step finishes.")
    exists = driver.execute_script("return !!document.getElementById('btn_search');")
    log(f"btn_search exists: {exists}")
    if not exists:
        dump_debug(driver, ymd)
        raise RuntimeError("btn_search not found in current document.")

    info = driver.execute_script(
        """
        const el = document.getElementById('btn_search');
        const r = el.getBoundingClientRect();
        return {text: el.innerText, tag: el.tagName, id: el.id, cls: el.className, href: el.getAttribute('href'),
                top:r.top,left:r.left,width:r.width,height:r.height};
        """
    )
    log(f"btn_search info: {info}")

    btn = driver.find_element(By.ID, "btn_search")

    methods = [
        lambda: (driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", btn), time.sleep(0.3), btn.click()),
        lambda: ActionChains(driver).move_to_element(btn).pause(0.2).click().perform(),
        lambda: driver.execute_script("document.getElementById('btn_search').click();"),
        lambda: driver.execute_script(
            """
            const el = document.getElementById('btn_search');
            ['mouseover','mousedown','mouseup','click'].forEach(t => {
              el.dispatchEvent(new MouseEvent(t, {view:window, bubbles:true, cancelable:true}));
            });
            """
        ),
    ]

    last_error = None
    for idx, method in enumerate(methods, start=1):
        try:
            log(f"Search click method {idx}...")
            method()
            time.sleep(1.0)
            log("btn_search clicked. Now waiting for Kyobo search to complete before Excel.")
            return
        except Exception as e:
            last_error = e
            log(f"Search click method {idx} failed: {e}")

    dump_debug(driver, ymd)
    raise RuntimeError(f"Could not click btn_search. Last error: {last_error}")


def wait_after_search(driver, ymd: str, partner=None):
    """V72: conservative Kyobo wait to prevent previous-date grid reuse.

    Problem found in V71:
      When processing consecutive dates for the same partner, Kyobo keeps the
      previous result grid visible while the next search is still loading.
      Because the partner is the same, the old grid can look like a valid current
      partner result. That can make 20260626 copy the 20260625 result.

    Fix:
      Do not accept visible rows too early. Wait a conservative fixed floor after
      clicking search, then require the browser to be quiet before Excel export.
    """
    partner_name = partner.get("name", "") if isinstance(partner, dict) else ""

    # Conservative floor. Kyobo is slower on Y-drive/company PCs and WebSquare can
    # leave stale rows visible for several seconds.
    min_wait = 10.0
    max_wait = 28.0
    start_time = time.time()
    last_info = None

    log(f"Waiting for Kyobo refreshed search result. ymd={ymd}, partner={partner_name}, min={min_wait}s max={max_wait}s")

    while time.time() - start_time < max_wait:
        elapsed = time.time() - start_time
        time.sleep(0.5)

        if click_confirm_if_popup(driver, partner):
            return False

        activity = browser_activity_info(driver)
        if partner is not None:
            last_info = visible_result_row_info(driver, partner)

        # Before the conservative floor, do not trust visible rows at all.
        if elapsed < min_wait:
            if int(elapsed) in (3, 6, 9):
                log(f"Kyobo refresh guard: still waiting to avoid stale previous-date grid. elapsed={elapsed:.1f}s")
            continue

        # After min_wait, wait until page activity is quiet.
        if activity.get("busy") and elapsed < max_wait - 2:
            continue

        if partner is not None:
            info = visible_result_row_info(driver, partner)
            last_info = info
            if page_has_current_partner_result_rows(driver, partner):
                log(f"Kyobo refreshed result accepted for {ymd} / {partner_name}. Evidence: {info}")
                return True

            if page_has_visible_result_rows(driver):
                log(f"Rows visible but not confirmed for {ymd} / {partner_name}. Waiting. Evidence: {info}")
                continue
        else:
            if not activity.get("busy"):
                log(f"Kyobo page activity is quiet after search. Evidence: {activity}")
                return True

    if click_confirm_if_popup(driver, partner):
        return False

    if partner is not None:
        if page_has_current_partner_result_rows(driver, partner):
            info = visible_result_row_info(driver, partner)
            log(f"Kyobo max wait reached; accepting confirmed current partner rows for {ymd}. Evidence: {info}")
            return True
        if page_has_visible_result_rows(driver) and not page_has_current_partner_result_rows(driver, partner):
            info = visible_result_row_info(driver, partner)
            log(f"Kyobo stale grid detected for {ymd} / {partner_name}; download stopped. Evidence: {info}")
            return False

    log(f"Kyobo search max wait reached. Last evidence: {last_info}. Proceeding with conservative fallback.")
    return True


def get_existing_download_files():
    if not DOWNLOAD_DIR.exists():
        return set()
    return {p.name for p in DOWNLOAD_DIR.iterdir() if p.is_file()}


def wait_for_new_download(before_files, timeout=90):
    """Wait for a newly downloaded file.

    Chrome/Youngpoong sometimes creates a random *.tmp file and quickly renames
    or deletes it.  Older versions tried to stat that temporary file twice and
    failed when it disappeared between checks.  V41 ignores transient temp files
    and keeps waiting for the final .xls/.xlsx/.csv file.
    """
    end = time.time() + timeout
    transient_suffixes = (".crdownload", ".tmp", ".download")

    def safe_stat(path):
        try:
            return path.stat()
        except FileNotFoundError:
            return None
        except OSError:
            return None

    last_seen = None
    while time.time() < end:
        try:
            current = [p for p in DOWNLOAD_DIR.iterdir() if p.is_file()]
        except FileNotFoundError:
            DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
            current = []

        # Prefer final Excel-like files, not Chrome's temporary *.tmp file.
        candidates = []
        for p in current:
            if p.name in before_files:
                continue
            lower = p.name.lower()
            if lower.endswith(transient_suffixes):
                last_seen = p.name
                continue
            st = safe_stat(p)
            if st and st.st_size > 0:
                candidates.append((p, st.st_mtime, st.st_size))

        if candidates:
            candidates.sort(key=lambda item: item[1], reverse=True)
            newest, _, size1 = candidates[0]
            time.sleep(0.5)
            st2 = safe_stat(newest)
            if st2 and st2.st_size == size1 and st2.st_size > 0:
                return newest

        # If only a temp file is visible, wait because Chrome may rename it soon.
        time.sleep(0.5)

    log(f"Download wait timed out. Last transient file seen: {last_seen}")
    return None


def move_downloaded_file(downloaded_path, target_path):
    downloaded_path = Path(downloaded_path)
    target_path = Path(target_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    if not downloaded_path.exists():
        candidates = [
            p for p in DOWNLOAD_DIR.glob("*")
            if p.is_file() and not p.name.endswith(".crdownload") and p.stat().st_size > 0
        ]
        if not candidates:
            raise FileNotFoundError(f"Downloaded file was not found: {downloaded_path}")
        downloaded_path = max(candidates, key=lambda p: p.stat().st_mtime)

    if target_path.exists():
        target_path.unlink()

    shutil.move(str(downloaded_path), str(target_path))
    log(f"Saved: {target_path}")


def click_excel_button(driver, ymd: str):
    log("Clicking Excel download button only AFTER search wait.")
    before = get_existing_download_files()

    candidates = driver.find_elements(
        By.XPATH,
        "//*[self::a or self::button or self::input][contains(normalize-space(.), '엑셀다운') or contains(@value, '엑셀다운') or contains(normalize-space(.), 'Excel') or contains(normalize-space(.), 'excel')]"
    )
    visible_candidates = []
    for el in candidates:
        try:
            if not visible(el):
                continue
            rect = driver.execute_script(
                "const r=arguments[0].getBoundingClientRect(); return {w:r.width,h:r.height,top:r.top,left:r.left};",
                el,
            )
            if rect["w"] >= 40 and rect["h"] >= 15:
                visible_candidates.append(el)
        except Exception:
            continue

    if not visible_candidates:
        dump_debug(driver, ymd)
        raise RuntimeError("Could not find Excel download button.")

    btn = visible_candidates[0]
    info = driver.execute_script(
        "const r=arguments[0].getBoundingClientRect(); return {text:arguments[0].innerText, tag:arguments[0].tagName, cls:arguments[0].className, top:r.top,left:r.left,width:r.width,height:r.height};",
        btn,
    )
    log(f"Excel button info: {info}")

    driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", btn)
    time.sleep(0.4)
    try:
        btn.click()
    except Exception:
        driver.execute_script("arguments[0].click();", btn)

    downloaded = wait_for_new_download(before, timeout=90)
    if downloaded is None:
        dump_debug(driver, ymd)
        raise RuntimeError("Download file did not appear within timeout.")

    ext = downloaded.suffix.lower() if downloaded.suffix else ".xls"
    target = DOWNLOAD_DIR / f"{ymd}{ext}"
    move_downloaded_file(downloaded, target)


def process_one_date(driver, date_text: str):
    ymd = date_text.replace("-", "")
    log("=" * 70)
    log(f"Processing {ymd}")
    ensure_iseum_partner_on_current_page(driver, context=f"before {ymd}")
    set_date_fields(driver, date_text)
    log(f"Date set to {date_text} ~ {date_text}")

    click_search_button(driver, ymd)
    has_data = wait_after_search(driver, ymd, None)
    close_sales_detail_popup(driver, max_wait=1)
    if not has_data:
        log(f"Skipped {ymd}: no data.")
        return

    click_excel_button(driver, ymd)


def make_date_list(start_text: str, end_text: str):
    start = datetime.strptime(start_text, "%Y-%m-%d").date()
    end = datetime.strptime(end_text, "%Y-%m-%d").date()
    if start > end:
        raise ValueError("Start date is later than end date.")
    dates = []
    cur = start
    while cur <= end:
        dates.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)
    return dates


def select_date_range():
    # Default both start/end dates to yesterday based on the local PC date.
    # Example: if today is 2026-06-29, default_start/default_end = 2026-06-28.
    yesterday = (datetime.now().date() - timedelta(days=1)).strftime("%Y-%m-%d")
    default_start = yesterday
    default_end = yesterday

    try:
        import tkinter as tk
        from tkinter import messagebox

        result = {"dates": None}

        root = tk.Tk()
        root.title("Kyobo SCM Date Range")
        root.geometry("380x200")
        root.attributes("-topmost", True)
        root.lift()
        root.resizable(False, False)

        tk.Label(root, text="Kyobo SCM Download Date Range", font=("Arial", 12, "bold")).pack(pady=(14, 8))

        frame = tk.Frame(root)
        frame.pack(pady=4)

        tk.Label(frame, text="Start date  YYYY-MM-DD").grid(row=0, column=0, sticky="w", padx=8, pady=4)
        start_var = tk.StringVar(value=default_start)
        start_entry = tk.Entry(frame, textvariable=start_var, width=16)
        start_entry.grid(row=0, column=1, padx=8, pady=4)

        tk.Label(frame, text="End date    YYYY-MM-DD").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        end_var = tk.StringVar(value=default_end)
        end_entry = tk.Entry(frame, textvariable=end_var, width=16)
        end_entry.grid(row=1, column=1, padx=8, pady=4)

        def on_start():
            try:
                result["dates"] = make_date_list(start_var.get().strip(), end_var.get().strip())
                root.quit()
                root.destroy()
            except Exception as e:
                messagebox.showerror("Date error", str(e))

        def on_cancel():
            result["dates"] = None
            root.quit()
            root.destroy()

        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=14)
        tk.Button(btn_frame, text="Start download", width=16, command=on_start).pack(side="left", padx=6)
        tk.Button(btn_frame, text="Cancel", width=10, command=on_cancel).pack(side="left", padx=6)

        root.protocol("WM_DELETE_WINDOW", on_cancel)
        start_entry.focus_set()
        root.mainloop()
        try:
            if root.winfo_exists():
                root.destroy()
        except Exception:
            pass

        if not result["dates"]:
            raise RuntimeError("Date range selection was cancelled.")

        return result["dates"]

    except Exception as gui_error:
        log(f"Date range GUI could not be used: {gui_error}")
        log("Please enter date range in this terminal.")
        start_text = input(f"Start date YYYY-MM-DD [{default_start}]: ").strip() or default_start
        end_text = input(f"End date YYYY-MM-DD [{default_end}]: ").strip() or default_end
        return make_date_list(start_text, end_text)



# =====================================================================
# V28 additions
# - Create one date workbook from "SCM 복사 양식.xlsx" for every selected date.
# - Download I-Seum data to sheet "교보_아동".
# - Then download Bookfolio data to sheet "교보_성인".
# - Copy downloaded Excel content into the target sheet while preserving the downloaded form.
# =====================================================================

PARTNERS = [
    {"name": "아이세움", "code": "0123910", "sheet_hint": "교보_아동", "sheet_keywords": ["교보", "아동"], "label": "미래엔(아이세움) 0123910"},
    {"name": "북폴리오", "code": "0123905", "sheet_hint": "교보_성인", "sheet_keywords": ["교보", "성인"], "label": "미래엔(북폴리오) 0123905"},
]


def partner_sheet_hint(partner):
    return partner.get("sheet_hint") or "_".join(partner.get("sheet_keywords", []))


def safe_text_for_log(value):
    text = str(value)
    try:
        text.encode("cp949")
        return text
    except Exception:
        return text.encode("unicode_escape").decode("ascii", errors="ignore")



def close_extra_windows_only(driver, tag="common"):
    """Close only extra browser windows/tabs. Do not click page menu links."""
    try:
        handles = driver.window_handles
        if len(handles) <= 1:
            return

        main = handles[0]
        current = None
        try:
            current = driver.current_window_handle
        except Exception:
            current = main

        # Prefer current window if still available, otherwise first handle.
        if current in handles:
            main = current

        for h in list(handles):
            if h == main:
                continue
            try:
                driver.switch_to.window(h)
                log(f"Closing extra popup window/tab: {h}")
                driver.close()
                time.sleep(0.2)
            except Exception as e:
                log(f"Could not close popup window/tab {h}: {e}")

        driver.switch_to.window(main)
        time.sleep(0.3)
    except Exception as e:
        log(f"Extra window close failed ({tag}): {e}")


def close_safe_popups(driver, tag="common"):
    """
    Conservative popup closer.
    Important: do NOT click normal page links such as 판매 통계 or notice titles.
    Only closes:
      - extra browser windows/tabs
      - JS alerts
      - obvious X/닫기/close controls by class/id/alt/src
    """
    log(f"Checking safe popups: {tag}")

    close_extra_windows_only(driver, tag)

    # JS alerts
    for _ in range(3):
        try:
            alert = driver.switch_to.alert
            text = alert.text
            log(f"Closing JS alert: {text[:120]}")
            alert.accept()
            time.sleep(0.4)
        except Exception:
            break

    js = r"""
    function visible(el){
        const r = el.getBoundingClientRect();
        const st = getComputedStyle(el);
        return r.width > 0 && r.height > 0 &&
               st.display !== 'none' &&
               st.visibility !== 'hidden' &&
               r.bottom >= 0 && r.top <= window.innerHeight &&
               r.right >= 0 && r.left <= window.innerWidth;
    }
    function txt(el){
        return ((el.innerText || el.textContent || el.value || el.alt || el.title || el.getAttribute('aria-label') || '') + '').trim();
    }

    const selectors = [
        'button[class*=close]', 'button[id*=close]', 'button[aria-label*=close]',
        'button[class*=Close]', 'button[id*=Close]',
        'input[type=button][value="닫기"]',
        'input[type=button][value="Close"]',
        'input[type=button][value="close"]',
        'img[alt="닫기"]',
        'img[alt="close"]',
        'img[alt="Close"]',
        'img[src*=close]',
        'img[src*=btn_close]',
        '[class*=popup] [class*=close]',
        '[class*=modal] [class*=close]',
        '[id*=popup] [class*=close]',
        '[id*=modal] [class*=close]',
        '.close', '.btn_close', '.popup_close', '.modal-close'
    ];

    const seen = new Set();
    const candidates = [];

    for (const sel of selectors) {
        document.querySelectorAll(sel).forEach(el => {
            if (seen.has(el)) return;
            seen.add(el);
            if (!visible(el)) return;

            const t = txt(el);
            const html = (el.outerHTML || '').toLowerCase();
            const cls = ((el.className || '') + '').toLowerCase();
            const id = ((el.id || '') + '').toLowerCase();
            const r = el.getBoundingClientRect();

            let score = 0;
            if (t === '닫기' || t === 'Close' || t === 'close' || t === 'X' || t === '×') score += 80;
            if (cls.includes('close') || id.includes('close')) score += 70;
            if ((el.getAttribute('alt') || '') === '닫기') score += 80;
            if (html.includes('btn_close') || html.includes('popup_close')) score += 60;
            if (r.width <= 120 && r.height <= 80) score += 10;

            if (score > 0) candidates.push({el, info:{score, text:t, x:r.x, y:r.y, w:r.width, h:r.height}});
        });
    }

    candidates.sort((a,b) => b.info.score - a.info.score || a.info.y - b.info.y || a.info.x - b.info.x);

    const clicked = [];
    for (const c of candidates.slice(0, 3)) {
        try {
            const el = c.el;
            const r = el.getBoundingClientRect();
            const cx = r.left + Math.max(3, Math.min(r.width - 3, r.width / 2));
            const cy = r.top + Math.max(3, Math.min(r.height - 3, r.height / 2));
            ['mouseover','mousedown','mouseup','click'].forEach(type => {
                el.dispatchEvent(new MouseEvent(type, {bubbles:true, cancelable:true, view:window, clientX:cx, clientY:cy}));
            });
            try { el.click(); } catch(e) {}
            clicked.push(c.info);
        } catch(e) {}
    }

    return clicked;
    """

    try:
        clicked = driver.execute_script(js)
        if clicked:
            log(f"Safe popup close clicked={len(clicked)}")
            for info in clicked[:3]:
                log(f"  safe closed: text={str(info.get('text',''))[:40]!r}, score={info.get('score')}")
            time.sleep(0.5)
        else:
            log("No safe popup close targets found.")
    except Exception as e:
        log(f"Safe popup close script failed: {e}")

    # one more alert check
    try:
        alert = driver.switch_to.alert
        text = alert.text
        log(f"Closing JS alert after safe popup close: {text[:120]}")
        alert.accept()
        time.sleep(0.4)
    except Exception:
        pass


def find_template_xlsx():
    candidates = [
        SCRIPT_DIR / "SCM 복사 양식.xlsx",
        Path.cwd() / "SCM 복사 양식.xlsx",
    ]
    candidates.extend(SCRIPT_DIR.glob("*SCM*복사*양식*.xlsx"))
    candidates.extend(Path.cwd().glob("*SCM*복사*양식*.xlsx"))
    checked = []
    for p in candidates:
        try:
            p = p.resolve()
        except Exception:
            pass
        checked.append(str(p))
        if p.exists() and p.suffix.lower() == ".xlsx":
            return p
    raise FileNotFoundError("SCM 복사 양식.xlsx file not found. Checked paths:\n" + "\n".join(checked))


def date_to_ymd(date_text: str) -> str:
    return date_text.replace("-", "")


def target_workbook_path_for_date(date_text: str) -> Path:
    return SCRIPT_DIR / f"{date_to_ymd(date_text)}.xlsx"


def create_date_workbooks(selected_dates):
    template = find_template_xlsx()
    log(f"Template workbook: {template}")
    outputs = {}
    for d in selected_dates:
        ymd = date_to_ymd(d)
        target = target_workbook_path_for_date(d)
        if target.exists():
            target.unlink()
        shutil.copy2(str(template), str(target))
        outputs[d] = target
        log(f"Created date workbook from template: {target}")
    return outputs


def is_partner_selected(driver, partner):
    text = get_current_partner_label(driver)
    return (partner["name"] in text and partner["code"] in text)


def try_click_partner_visible_option(driver, partner):
    """Open the WebSquare vendor dropdown and click the exact visible option."""
    name = partner["name"]
    code = partner["code"]
    try:
        label = driver.find_element(By.ID, "s_vndrList_label")
        try:
            root = label.find_element(By.XPATH, "./ancestor::*[contains(@class,'w2selectbox') or @id='s_vndrList'][1]")
        except Exception:
            root = label

        driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", root)
        time.sleep(0.3)
        try:
            root.click()
        except Exception:
            driver.execute_script(
                """
                const el = arguments[0];
                const r = el.getBoundingClientRect();
                const cx = r.left + Math.max(3, Math.min(r.width - 3, r.width / 2));
                const cy = r.top + Math.max(3, Math.min(r.height - 3, r.height / 2));
                for (const type of ['mouseover','mousedown','mouseup','click']) {
                  el.dispatchEvent(new MouseEvent(type, {bubbles:true, cancelable:true, view:window, clientX:cx, clientY:cy}));
                }
                try { el.click(); } catch(e) {}
                """,
                root,
            )
        time.sleep(1.0)

        result = driver.execute_script(
            """
            return (function(name, code){
                function norm(v){ return (v === null || v === undefined) ? '' : String(v).trim(); }
                function visible(el){
                    const r = el.getBoundingClientRect();
                    const st = getComputedStyle(el);
                    return r.width > 20 && r.height > 8 && st.display !== 'none' && st.visibility !== 'hidden' && r.bottom >= 0 && r.top <= innerHeight;
                }
                function clickLikeHuman(el){
                    const r = el.getBoundingClientRect();
                    const cx = Math.floor(r.left + Math.min(r.width - 2, Math.max(2, r.width / 2)));
                    const cy = Math.floor(r.top + Math.min(r.height - 2, Math.max(2, r.height / 2)));
                    const target = document.elementFromPoint(cx, cy) || el;
                    for (const type of ['mouseover','mousemove','mousedown','mouseup','click']) {
                        try { target.dispatchEvent(new MouseEvent(type, {bubbles:true, cancelable:true, view:window, clientX:cx, clientY:cy})); } catch(e) {}
                        try { el.dispatchEvent(new MouseEvent(type, {bubbles:true, cancelable:true, view:window, clientX:cx, clientY:cy})); } catch(e) {}
                    }
                    try { target.click(); } catch(e) {}
                    try { el.click(); } catch(e) {}
                    return {x:cx, y:cy, tag:el.tagName, id:el.id, cls:el.className, text:norm(el.innerText || el.textContent)};
                }
                const candidates = Array.from(document.querySelectorAll('div,td,tr,li,a,span'))
                    .filter(el => {
                        const txt = norm(el.innerText || el.textContent);
                        if (!txt.includes(name) || !txt.includes(code)) return false;
                        if (!visible(el)) return false;
                        const r = el.getBoundingClientRect();
                        if (r.width > 900 || r.height > 160) return false;
                        return true;
                    })
                    .sort((a,b) => {
                        const ra = a.getBoundingClientRect();
                        const rb = b.getBoundingClientRect();
                        // Prefer option-like narrow rows, then lower-level/smaller elements.
                        const sa = (ra.width < 450 ? 50 : 0) + (ra.height < 45 ? 50 : 0) - a.querySelectorAll('*').length;
                        const sb = (rb.width < 450 ? 50 : 0) + (rb.height < 45 ? 50 : 0) - b.querySelectorAll('*').length;
                        return sb - sa;
                    });
                if (!candidates.length) return {ok:false, reason:'no visible candidate'};
                const clicked = clickLikeHuman(candidates[0]);
                return {ok:true, clicked};
            })(arguments[0], arguments[1]);
            """,
            name,
            code,
        )
        log(f"Visible partner option click result for {name}: {result}")
        time.sleep(2.0)
        return is_partner_selected(driver, partner)
    except Exception as e:
        log(f"Visible partner option click failed for {name}: {e}")
        return False


def try_websquare_select_partner(driver, partner):
    """Fallback: use WebSquare component API by index/code. Avoid wrong setValue overwrite."""
    name = partner["name"]
    code = partner["code"]
    try:
        result = driver.execute_script(
            """
            return (function(name, code){
                const ids = ['s_vndrList', 'sbx_vndrList', 'vndrList'];
                function getComp(id){
                    try { if (window.$p && $p.getComponentById) return $p.getComponentById(id); } catch(e) {}
                    try { if (window.WebSquare && WebSquare.util && WebSquare.util.getComponentById) return WebSquare.util.getComponentById(id); } catch(e) {}
                    try { if (window[id]) return window[id]; } catch(e) {}
                    return null;
                }
                function call(obj, names, args){
                    for (const n of names){
                        try { if (obj && typeof obj[n] === 'function') return obj[n].apply(obj, args || []); } catch(e) {}
                    }
                    return undefined;
                }
                function norm(v){ return (v === null || v === undefined) ? '' : String(v).trim(); }
                function fire(){
                    const label = document.getElementById('s_vndrList_label');
                    const root = document.getElementById('s_vndrList') || (label ? label.closest('[id]') : null);
                    for (const el of [root, label].filter(Boolean)){
                        try { el.dispatchEvent(new Event('input', {bubbles:true})); } catch(e) {}
                        try { el.dispatchEvent(new Event('change', {bubbles:true})); } catch(e) {}
                    }
                }
                const logs = [];
                for (const id of ids){
                    const comp = getComp(id);
                    if (!comp) { logs.push(id + ': no component'); continue; }
                    logs.push(id + ': component found');
                    let count = Number(call(comp, ['getItemCount', 'getItemLength', 'getLength'], []));
                    if (!Number.isFinite(count) || count <= 0 || count > 200) count = 30;
                    for (let i=0; i<count; i++){
                        const text = norm(call(comp, ['getItemText', 'getItemLabel', 'getText'], [i]));
                        const value = norm(call(comp, ['getItemValue'], [i]));
                        const item = norm(call(comp, ['getItem', 'getOption'], [i]));
                        const merged = [text, value, item].join(' ');
                        if (merged.includes(name) && merged.includes(code)){
                            logs.push('target index=' + i + ' text=' + text + ' value=' + value);
                            call(comp, ['setSelectedIndex'], [i]);
                            call(comp, ['setIndex'], [i]);
                            call(comp, ['select'], [i]);
                            call(comp, ['setSelectedItem'], [i]);
                            if (value && value.includes(code)) call(comp, ['setValue', 'setSelectedValue'], [value]);
                            call(comp, ['trigger', 'fireEvent'], ['change']);
                            fire();
                            const label = norm((document.getElementById('s_vndrList_label') || {}).innerText);
                            if (label.includes(name) && label.includes(code)) return {ok:true, method:'index', id, index:i, label, logs};
                        }
                    }
                    call(comp, ['setValue', 'setSelectedValue'], [code]);
                    call(comp, ['trigger', 'fireEvent'], ['change']);
                    fire();
                    const label = norm((document.getElementById('s_vndrList_label') || {}).innerText);
                    if (label.includes(name) && label.includes(code)) return {ok:true, method:'code', id, label, logs};
                }
                return {ok:false, logs};
            })(arguments[0], arguments[1]);
            """,
            name,
            code,
        )
        log(f"WebSquare select result for {name}: {result}")
        time.sleep(2.0)
        return is_partner_selected(driver, partner)
    except Exception as e:
        log(f"WebSquare select failed for {name}: {e}")
        return False


def select_partner(driver, partner):
    name = partner["name"]
    code = partner["code"]
    close_password_change_popup(driver, max_wait=3)
    log(f"Selecting partner: {name} {code}")
    label = get_current_partner_label(driver)
    log(f"Current partner label before selection: {label}")
    if is_partner_selected(driver, partner):
        log(f"Partner already selected: {name} {code}")
        return

    for attempt in range(1, 4):
        log(f"Partner selection attempt {attempt}: {name} {code}")
        if try_click_partner_visible_option(driver, partner):
            time.sleep(2.0)
            label = get_current_partner_label(driver)
            log(f"Partner label after visible click: {label}")
            if is_partner_selected(driver, partner):
                return
        if try_websquare_select_partner(driver, partner):
            time.sleep(2.0)
            label = get_current_partner_label(driver)
            log(f"Partner label after WebSquare API: {label}")
            if is_partner_selected(driver, partner):
                return
        time.sleep(1.5)

    label = get_current_partner_label(driver)
    dump_debug(driver, f"select_partner_failed_{name}_{code}")
    raise RuntimeError(f"Could not select partner {name} {code}. Current label: {label}")


def ensure_partner_on_current_page(driver, partner, context=""):
    prefix = f"[{context}] " if context else ""
    # If a previous quantity/detail popup is open, close it before touching the vendor dropdown.
    try:
        close_sales_detail_popup(driver, max_wait=1)
    except Exception:
        pass
    label = get_current_partner_label(driver)
    log(f"{prefix}Partner label check: {label}")
    if not is_partner_selected(driver, partner):
        log(f"{prefix}Partner is not {partner['name']}. Re-selecting on current page.")
        select_partner(driver, partner)

    # Verify again after a few seconds because Kyobo can briefly revert the dropdown.
    for sec in (1.5, 3.0):
        time.sleep(sec)
        label = get_current_partner_label(driver)
        log(f"{prefix}Partner stability check after {sec:.1f}s: {label}")
        if not is_partner_selected(driver, partner):
            log(f"{prefix}Partner reverted. Selecting again.")
            select_partner(driver, partner)
        else:
            return

    label = get_current_partner_label(driver)
    if not is_partner_selected(driver, partner):
        dump_debug(driver, f"partner_not_stable_{partner['name']}_{partner['code']}")
        raise RuntimeError(f"Partner is not stable: expected {partner['name']} {partner['code']}, current label: {label}")


def get_excel_app():
    try:
        import win32com.client as win32
    except Exception as e:
        raise RuntimeError("pywin32 is required to copy downloaded Excel sheets. The batch file will try to install it automatically. Original error: " + str(e))
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        excel.AskToUpdateLinks = False
    except Exception:
        pass
    try:
        excel.AutomationSecurity = 3
    except Exception:
        pass
    return excel


def normalize_korean_sheet_name(value):
    import unicodedata
    text = unicodedata.normalize("NFKC", str(value or ""))
    return text.replace(" ", "").replace("_", "").replace("-", "").strip()


def resolve_target_worksheet(dst_wb, partner):
    """Find the real sheet in the template by keywords instead of relying on a hard-coded Korean sheet name."""
    keywords = partner.get("sheet_keywords") or []
    hint = partner_sheet_hint(partner)
    available = []

    # 1) Exact hint match if possible.
    for ws in dst_wb.Worksheets:
        try:
            name = str(ws.Name).strip()
            available.append(name)
            if name == hint:
                return ws, name, available, "exact-hint"
        except Exception:
            pass

    # 2) Normalized hint match.
    norm_hint = normalize_korean_sheet_name(hint)
    for ws in dst_wb.Worksheets:
        try:
            name = str(ws.Name).strip()
            if normalize_korean_sheet_name(name) == norm_hint:
                return ws, name, available, "normalized-hint"
        except Exception:
            pass

    # 3) Keyword match: contains both 교보+아동 or 교보+성인.
    norm_keywords = [normalize_korean_sheet_name(k) for k in keywords if str(k).strip()]
    matches = []
    for ws in dst_wb.Worksheets:
        try:
            name = str(ws.Name).strip()
            norm_name = normalize_korean_sheet_name(name)
            if norm_keywords and all(k in norm_name for k in norm_keywords):
                matches.append((ws, name))
        except Exception:
            pass

    if len(matches) == 1:
        return matches[0][0], matches[0][1], available, "keyword-match"
    if len(matches) > 1:
        # Prefer a non-hidden visible sheet if Excel exposes Visible= -1.
        for ws, name in matches:
            try:
                if int(ws.Visible) == -1:
                    return ws, name, available, "keyword-visible-match"
            except Exception:
                pass
        return matches[0][0], matches[0][1], available, "keyword-first-match"

    # 4) Last resort: if the sheet exists with a weird encoding in log only, fail clearly.
    raise RuntimeError(
        "Target Kyobo sheet not found. "
        f"Need keywords={keywords}, hint={hint}. "
        f"Available sheets={available}"
    )


def copy_download_to_target_sheet(downloaded_path: Path, target_workbook: Path, partner):
    downloaded_path = Path(downloaded_path).resolve()
    target_workbook = Path(target_workbook).resolve()
    sheet_hint = partner_sheet_hint(partner)
    log(f"Copying downloaded file into {target_workbook.name} / {sheet_hint}: {downloaded_path.name}")
    if not downloaded_path.exists():
        raise FileNotFoundError(f"Downloaded file not found: {downloaded_path}")
    if not target_workbook.exists():
        raise FileNotFoundError(f"Target workbook not found: {target_workbook}")

    excel = get_excel_app()
    src_wb = None
    dst_wb = None
    try:
        src_wb = excel.Workbooks.Open(str(downloaded_path), UpdateLinks=0, ReadOnly=True)
        dst_wb = excel.Workbooks.Open(str(target_workbook), UpdateLinks=0, ReadOnly=False)
        src_ws = src_wb.Worksheets(1)
        dst_ws, actual_sheet_name, available, method = resolve_target_worksheet(dst_wb, partner)
        log(f"Resolved target sheet by {method}: {actual_sheet_name}")

        try:
            dst_ws.Activate()
        except Exception:
            pass

        dst_ws.Cells.Clear()
        src_ws.UsedRange.Copy()
        dst_ws.Range("A1").PasteSpecial(Paste=-4104)  # xlPasteAll
        try:
            dst_ws.Range("A1").PasteSpecial(Paste=8)  # xlPasteColumnWidths
        except Exception:
            pass
        excel.CutCopyMode = False

        # Preserve row heights and column widths for the used range as much as possible.
        try:
            rows_count = src_ws.UsedRange.Rows.Count
            cols_count = src_ws.UsedRange.Columns.Count
            for r in range(1, min(rows_count, 500) + 1):
                dst_ws.Rows(r).RowHeight = src_ws.Rows(r).RowHeight
            for c in range(1, min(cols_count, 120) + 1):
                dst_ws.Columns(c).ColumnWidth = src_ws.Columns(c).ColumnWidth
        except Exception:
            pass

        try:
            dst_ws.Activate()
            dst_wb.Worksheets(actual_sheet_name).Activate()
        except Exception:
            pass
        dst_wb.Save()
        log(f"Saved target workbook: {target_workbook} / sheet={actual_sheet_name}")
    finally:
        try:
            if src_wb is not None:
                src_wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if dst_wb is not None:
                dst_wb.Close(SaveChanges=True)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass


def close_sales_detail_popup(driver, max_wait=2):
    """Close Kyobo sales-detail popup (판매확인) if an accidental quantity-cell popup is open."""
    end = time.time() + max_wait
    closed = False
    while time.time() < end:
        try:
            page_text = driver.execute_script("return document.body ? document.body.innerText : ''; ") or ""
            if "판매확인" not in page_text and "판매수량" not in page_text:
                return closed
            log("Sales-detail popup detected. Closing it before continuing...")
            result = driver.execute_script(
                """
                return (function(){
                    function visible(el){
                        const r = el.getBoundingClientRect();
                        const st = getComputedStyle(el);
                        return r.width > 0 && r.height > 0 && st.display !== 'none' && st.visibility !== 'hidden' && r.bottom >= 0 && r.top <= innerHeight;
                    }
                    function txt(el){ return (el.innerText || el.textContent || el.value || '').trim(); }
                    const all = Array.from(document.querySelectorAll('a,button,input,div,span'));
                    const closeCandidates = all.filter(el => visible(el) && (txt(el) === '확인' || txt(el) === 'X' || txt(el) === '×' || (el.className && String(el.className).toLowerCase().includes('close'))));
                    // Prefer confirm button inside popup, then close X near top-right.
                    closeCandidates.sort((a,b) => {
                        const ta = txt(a), tb = txt(b);
                        const sa = (ta === '확인' ? 100 : 0) + (String(a.className).toLowerCase().includes('close') ? 50 : 0);
                        const sb = (tb === '확인' ? 100 : 0) + (String(b.className).toLowerCase().includes('close') ? 50 : 0);
                        return sb - sa;
                    });
                    if (!closeCandidates.length) return {ok:false, reason:'no close button'};
                    const el = closeCandidates[0];
                    try { el.click(); } catch(e) {
                        const r = el.getBoundingClientRect();
                        el.dispatchEvent(new MouseEvent('click', {bubbles:true, cancelable:true, view:window, clientX:r.left+r.width/2, clientY:r.top+r.height/2}));
                    }
                    return {ok:true, text:txt(el), id:el.id, cls:String(el.className)};
                })();
                """
            )
            log(f"Sales-detail popup close result: {result}")
            time.sleep(0.8)
            closed = True
        except Exception as e:
            log(f"Sales-detail popup close failed: {e}")
            return closed
    return closed


def find_excel_download_element_js(driver):
    """Find the real right-side Excel download button, not quantity cells in the result grid."""
    return driver.execute_script(
        """
        return (function(){
            function visible(el){
                const r = el.getBoundingClientRect();
                const st = getComputedStyle(el);
                return r.width > 20 && r.height > 10 && st.display !== 'none' && st.visibility !== 'hidden' && r.bottom >= 0 && r.top <= innerHeight;
            }
            function text(el){ return (el.innerText || el.textContent || el.value || el.getAttribute('title') || '').trim(); }
            const all = Array.from(document.querySelectorAll('a,button,input,div,span'));
            let candidates = all.filter(el => {
                const t = text(el);
                if (!visible(el)) return false;
                if (!(t.includes('엑셀다운') || t.toLowerCase().includes('excel'))) return false;
                const r = el.getBoundingClientRect();
                if (r.width > 220 || r.height > 80) return false;
                // The SCM excel button is normally on the right side above the grid.
                if (r.left < innerWidth * 0.55) return false;
                return true;
            });
            candidates.sort((a,b) => {
                const ra = a.getBoundingClientRect(), rb = b.getBoundingClientRect();
                const ta = text(a), tb = text(b);
                const sa = (ta === '엑셀다운' ? 100 : 0) + (String(a.id).toLowerCase().includes('excel') ? 50 : 0) + ra.left/10 - Math.abs(ra.top-350)/20;
                const sb = (tb === '엑셀다운' ? 100 : 0) + (String(b.id).toLowerCase().includes('excel') ? 50 : 0) + rb.left/10 - Math.abs(rb.top-350)/20;
                return sb - sa;
            });
            if (!candidates.length) return {ok:false, reason:'no candidate'};
            const el = candidates[0];
            const r = el.getBoundingClientRect();
            window.__kyoboExcelEl = el;
            return {ok:true, text:text(el), id:el.id, tag:el.tagName, cls:String(el.className), left:r.left, top:r.top, width:r.width, height:r.height};
        })();
        """
    )

def click_excel_button(driver, ymd: str, partner):
    log("Clicking Excel download button by exact Excel-download element. Quantity cells will NOT be clicked.")
    close_sales_detail_popup(driver, max_wait=2)
    before = get_existing_download_files()

    info = find_excel_download_element_js(driver)
    log(f"Excel download candidate by JS: {info}")
    if not info or not info.get("ok"):
        # Fallback to Selenium search, but keep strong filtering.
        candidates = driver.find_elements(
            By.XPATH,
            "//*[self::a or self::button or self::input or self::div or self::span][contains(normalize-space(.), '엑셀다운') or contains(@value, '엑셀다운') or contains(normalize-space(.), 'Excel') or contains(normalize-space(.), 'excel')]"
        )
        visible_candidates = []
        for el in candidates:
            try:
                if not visible(el):
                    continue
                rect = driver.execute_script(
                    "const r=arguments[0].getBoundingClientRect(); return {w:r.width,h:r.height,top:r.top,left:r.left,text:(arguments[0].innerText||arguments[0].value||'').trim(),id:arguments[0].id,cls:arguments[0].className};",
                    el,
                )
                if rect["w"] >= 40 and rect["h"] >= 15 and rect["w"] <= 220 and rect["left"] > 500:
                    visible_candidates.append(el)
            except Exception:
                continue
        if not visible_candidates:
            dump_debug(driver, f"excel_button_not_found_{ymd}_{partner['name']}")
            raise RuntimeError("Could not find Excel download button.")
        # Store selected fallback element in JS variable.
        driver.execute_script("window.__kyoboExcelEl = arguments[0];", visible_candidates[0])
        info = driver.execute_script(
            "const el=window.__kyoboExcelEl; const r=el.getBoundingClientRect(); return {ok:true,text:(el.innerText||el.value||'').trim(),tag:el.tagName,id:el.id,cls:el.className,left:r.left,top:r.top,width:r.width,height:r.height};"
        )
        log(f"Excel download fallback candidate: {info}")

    # Use DOM click first. This avoids Selenium coordinate-click accidentally hitting a sales quantity cell.
    click_result = driver.execute_script(
        """
        return (function(){
            const el = window.__kyoboExcelEl;
            if (!el) return {ok:false, reason:'window.__kyoboExcelEl missing'};
            try { el.scrollIntoView({block:'center', inline:'center'}); } catch(e) {}
            const r = el.getBoundingClientRect();
            try { el.click(); return {ok:true, method:'dom-click', text:(el.innerText||el.value||'').trim(), id:el.id, left:r.left, top:r.top}; }
            catch(e) {
                try {
                    ['mouseover','mousedown','mouseup','click'].forEach(t => el.dispatchEvent(new MouseEvent(t, {view:window, bubbles:true, cancelable:true, clientX:r.left+r.width/2, clientY:r.top+r.height/2})));
                    return {ok:true, method:'mouseevent', text:(el.innerText||el.value||'').trim(), id:el.id, left:r.left, top:r.top};
                } catch(e2) { return {ok:false, error:String(e2)}; }
            }
        })();
        """
    )
    log(f"Excel download click result: {click_result}")

    downloaded = None
    for attempt in range(1, 4):
        time.sleep(1.2)
        if close_sales_detail_popup(driver, max_wait=1):
            dump_debug(driver, f"wrong_popup_after_excel_click_{ymd}_{partner['name']}")
            raise RuntimeError("Excel click opened a sales-detail popup instead of downloading. Debug HTML/PNG saved.")

        if close_no_data_popup_but_keep_data(driver, partner, context=f"Excel click attempt {attempt} / {ymd} / {partner['name']}"):
            # The grid is visible, so this is a false no-data popup. Retry the Excel click.
            if attempt < 3:
                log("Retrying Excel download after closing false no-data popup.")
                click_result = driver.execute_script(
                    """
                    return (function(){
                        const el = window.__kyoboExcelEl;
                        if (!el) return {ok:false, reason:'window.__kyoboExcelEl missing'};
                        try { el.scrollIntoView({block:'center', inline:'center'}); } catch(e) {}
                        const r = el.getBoundingClientRect();
                        try { el.click(); return {ok:true, method:'dom-click-retry', text:(el.innerText||el.value||'').trim(), id:el.id, left:r.left, top:r.top}; }
                        catch(e) { return {ok:false, error:String(e)}; }
                    })();
                    """
                )
                log(f"Excel download retry click result: {click_result}")
                continue

        downloaded = wait_for_new_download(before, timeout=25 if attempt < 3 else 90)
        if downloaded is not None:
            break

        if attempt < 3:
            log(f"No downloaded file yet after Excel attempt {attempt}. Retrying Excel button click.")
            click_result = driver.execute_script(
                """
                return (function(){
                    const el = window.__kyoboExcelEl;
                    if (!el) return {ok:false, reason:'window.__kyoboExcelEl missing'};
                    try { el.scrollIntoView({block:'center', inline:'center'}); } catch(e) {}
                    const r = el.getBoundingClientRect();
                    try { el.click(); return {ok:true, method:'dom-click-timeout-retry', text:(el.innerText||el.value||'').trim(), id:el.id, left:r.left, top:r.top}; }
                    catch(e) { return {ok:false, error:String(e)}; }
                })();
                """
            )
            log(f"Excel download timeout-retry click result: {click_result}")

    if downloaded is None:
        dump_debug(driver, f"excel_download_timeout_{ymd}_{partner['name']}")
        raise RuntimeError("Download file did not appear within timeout.")

    ext = downloaded.suffix.lower() if downloaded.suffix else ".xls"
    safe_sheet = partner_sheet_hint(partner).replace("/", "_").replace("\\", "_")
    target = DOWNLOAD_DIR / f"{ymd}_{safe_sheet}{ext}"
    move_downloaded_file(downloaded, target)
    return target

def process_one_partner_date(driver, partner, date_text: str, target_workbook: Path):
    ymd = date_to_ymd(date_text)
    log("=" * 70)
    log(f"Processing {ymd} / {partner['name']} {partner['code']} -> {partner_sheet_hint(partner)}")

    ensure_partner_on_current_page(driver, partner, context=f"before date input {ymd} {partner['name']}")
    set_date_fields(driver, date_text)
    log(f"Date set to {date_text} ~ {date_text}")

    # Kyobo sometimes resets vendor after UI interaction. Verify one more time immediately before search.
    ensure_partner_on_current_page(driver, partner, context=f"before search {ymd} {partner['name']}")

    click_search_button(driver, ymd)
    has_data = wait_after_search(driver, ymd, partner)
    close_sales_detail_popup(driver, max_wait=1)
    if not has_data:
        log(f"Skipped download for {ymd} / {partner['name']}: no data. Target sheet remains as template.")
        return

    # Final guard: if the grid still belongs to the previous partner, do not download/copy.
    if not page_has_current_partner_result_rows(driver, partner):
        info = visible_result_row_info(driver, partner)
        log(f"Final guard stopped Excel download because grid is not current partner {partner['name']}. Evidence: {info}")
        return

    # V58:
    # Kyobo sometimes shows a transient "조회된 데이터가 없습니다" popup if Excel is clicked
    # immediately after the grid becomes visible. Give the export layer a short stability wait
    # and re-check that current-partner rows are still visible before clicking Excel.
    log(f"Kyobo export stability wait before Excel: {ymd} / {partner['name']} / 3.0s")
    time.sleep(3.0)
    close_no_data_popup_but_keep_data(driver, partner, context=f"pre-excel stability wait / {ymd} / {partner['name']}")
    close_sales_detail_popup(driver, max_wait=1)
    if not page_has_current_partner_result_rows(driver, partner):
        info = visible_result_row_info(driver, partner)
        log(f"Excel download stopped after stability wait because current partner rows disappeared. Evidence: {info}")
        return

    downloaded = click_excel_button(driver, ymd, partner)
    copy_download_to_target_sheet(downloaded, target_workbook, partner)


def process_partner_all_dates(driver, partner, selected_dates, output_map):
    log("")
    log("#" * 70)
    log(f"Starting partner batch: {partner['name']} {partner['code']} -> {partner_sheet_hint(partner)}")
    log("#" * 70)
    ensure_partner_on_current_page(driver, partner, context=f"start partner batch {partner['name']}")
    for d in selected_dates:
        process_one_partner_date(driver, partner, d, output_map[d])


def main():
    ensure_dirs()
    selected_dates = select_date_range()
    output_map = create_date_workbooks(selected_dates)

    log("Kyobo SCM downloader")
    log(f"VERSION: {VERSION}")
    log(f"Current directory: {Path.cwd()}")
    log(f"Script path: {Path(__file__).resolve()}")
    log(f"BASE_DIR: {BASE_DIR}")
    log(f"DOWNLOAD_DIR: {DOWNLOAD_DIR}")
    log(f"LOGIN_XLSX: {LOGIN_XLSX}")
    log(f"DATE_COUNT: {len(selected_dates)}")
    log(f"DATE_RANGE: {selected_dates[0]} ~ {selected_dates[-1]}")
    log("")

    driver = build_driver()

    try:
        auto_login_kyobo(driver)
        close_password_change_popup(driver, max_wait=8)

        # Open the Sales Search page first. Kyobo can reset the vendor dropdown on page move.
        open_sales_search_page(driver)
        close_password_change_popup(driver, max_wait=3)

        log("")
        log("Sales Search page is ready. Starting daily downloads into template date files.")
        log("")

        # User requested: I-Seum first, then Bookfolio.
        for partner in PARTNERS:
            process_partner_all_dates(driver, partner, selected_dates, output_map)

        log("=" * 70)
        log("Finished all dates and partners.")
        log("Created/updated files:")
        for d in selected_dates:
            log(f" - {output_map[d]}")

    except Exception as e:
        log("")
        log(f"ERROR: {e}")
        log(traceback.format_exc())
        log("Please send this terminal screen and debug files if created.")
    finally:
        input("Press ENTER to exit...")
        try:
            driver.quit()
        except Exception:
            pass




# =====================================================================
# V36 additions: run Youngpoong SCM after Kyobo V34 using the same dates.
# - Youngpoong login info is read from scm_login.xlsx in the same folder.
# - Downloaded Youngpoong Excel is copied into the date workbook sheet that contains "영풍".
# - Youngpoong sales status page is opened automatically by clicking li.sell_btn.
# =====================================================================

YPSCM_LOGIN_URL = "https://ypscm.ypbooks.co.kr/Account/Logout"
YPSCM_PARTNER = {"name": "영풍문고", "code": "", "sheet_hint": "영풍문고", "sheet_keywords": ["영풍"], "label": "영풍문고"}
YPSCM_SEARCH_WAIT_SECONDS = 18  # max seconds; V45 uses action-based polling
YPSCM_DOWNLOAD_WAIT_SECONDS = 90


def read_ypscm_login_info():
    login_xlsx_path = find_login_xlsx()
    log(f"Using login file for Youngpoong: {login_xlsx_path}")

    wb = load_workbook(login_xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    col_map = {}
    for idx, row in enumerate(rows):
        normalized = [_normalize_header_name(x) for x in row]
        if "id" in normalized and "password" in normalized:
            header_row_idx = idx
            for c_idx, name in enumerate(normalized):
                if name:
                    col_map[name] = c_idx
            break

    if header_row_idx is None:
        raise RuntimeError("Could not find header row containing id/password or 아이디/비밀번호 in scm_login.xlsx")

    id_col = col_map.get("id")
    pw_col = col_map.get("password")
    possible_site_cols = [col_map.get("site"), col_map.get("사이트"), col_map.get("siteurl"), col_map.get("url"), col_map.get("주소")]
    site_col = next((x for x in possible_site_cols if x is not None), None)

    data_rows = [r for r in rows[header_row_idx + 1:] if any(x is not None and str(x).strip() for x in r)]
    fallback_rows = []
    for row in data_rows:
        row_text = " ".join(str(x) for x in row if x is not None)
        row_lower = row_text.lower()
        is_ypscm = ("영풍" in row_text) or ("ypbooks" in row_lower) or ("ypscm" in row_lower) or ("youngpoong" in row_lower)
        is_kyobo = ("교보" in row_text) or ("kyobo" in row_lower) or ("scm.kyobobook" in row_lower)
        if not is_kyobo:
            fallback_rows.append(row)
        if not is_ypscm:
            continue
        login_id = row[id_col] if id_col is not None and id_col < len(row) else None
        login_pw = row[pw_col] if pw_col is not None and pw_col < len(row) else None
        login_url = row[site_col] if site_col is not None and site_col < len(row) and row[site_col] else YPSCM_LOGIN_URL
        if login_id is None or login_pw is None:
            raise RuntimeError("Youngpoong row found, but id/password is empty in scm_login.xlsx")
        return str(login_id).strip(), str(login_pw).strip(), str(login_url).strip()

    # If there are exactly two rows and one is Kyobo, use the other as Youngpoong.
    if len(fallback_rows) == 1:
        row = fallback_rows[0]
        login_id = row[id_col] if id_col is not None and id_col < len(row) else None
        login_pw = row[pw_col] if pw_col is not None and pw_col < len(row) else None
        login_url = row[site_col] if site_col is not None and site_col < len(row) and row[site_col] else YPSCM_LOGIN_URL
        if login_id and login_pw:
            log("Youngpoong row was inferred from the non-Kyobo row in scm_login.xlsx.")
            return str(login_id).strip(), str(login_pw).strip(), str(login_url).strip()

    raise RuntimeError("Could not find Youngpoong login row in scm_login.xlsx. Please include 영풍/ypscm/ypbooks in that row.")


def yp_visible_inputs(driver):
    return [e for e in driver.find_elements(By.TAG_NAME, "input") if visible(e)]


def yp_find_button_by_text(driver, target_text):
    xpath = (
        "//*[self::button or self::a or self::input or self::span]"
        f"[contains(normalize-space(.), '{target_text}') or @value='{target_text}' or contains(@title, '{target_text}')]"
    )
    els = []
    for e in driver.find_elements(By.XPATH, xpath):
        try:
            if not visible(e):
                continue
            txt = (e.text or e.get_attribute("value") or e.get_attribute("title") or "").strip()
            w, h = e.size.get("width", 0), e.size.get("height", 0)
            if w > 500 or h > 120:
                continue
            els.append(e)
        except Exception:
            pass
    return els


def yp_click_element(driver, el):
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", el)
    except Exception:
        pass
    time.sleep(0.2)
    for method in ("selenium", "action", "js"):
        try:
            if method == "selenium":
                el.click()
            elif method == "action":
                ActionChains(driver).move_to_element(el).pause(0.1).click().perform()
            else:
                driver.execute_script("arguments[0].click();", el)
            return True
        except Exception:
            pass
    return False


def auto_login_ypscm(driver):
    login_id, login_pw, login_url = read_ypscm_login_info()
    log(f"Opening Youngpoong SCM login page: {login_url}")
    driver.get(login_url)
    time.sleep(2)

    # Some sites redirect from Logout to Login. Wait for at least two fields, then fill.
    try:
        WebDriverWait(driver, 20).until(lambda d: len(yp_visible_inputs(d)) >= 2)
    except Exception:
        dump_debug(driver, "ypscm_login_fields_not_found")
        raise RuntimeError("Could not find Youngpoong login fields.")

    inputs = yp_visible_inputs(driver)
    password_inputs = [x for x in inputs if (x.get_attribute("type") or "").lower() == "password"]
    text_inputs = [x for x in inputs if (x.get_attribute("type") or "text").lower() in ("text", "", "email", "tel", "number")]
    if not password_inputs or not text_inputs:
        dump_debug(driver, "ypscm_login_fields_not_found")
        raise RuntimeError("Could not identify Youngpoong id/password fields.")

    log("Typing Youngpoong login id/password from scm_login.xlsx.")
    set_text_input(text_inputs[0], login_id)
    set_text_input(password_inputs[0], login_pw)
    time.sleep(0.5)

    buttons = []
    for text in ["로그인", "Login", "LOGIN"]:
        buttons.extend(yp_find_button_by_text(driver, text))
    if not buttons:
        # Fallback: submit button or Enter from password field.
        buttons = [e for e in driver.find_elements(By.CSS_SELECTOR, "button,input[type='submit'],a") if visible(e)]

    clicked = False
    for btn in buttons[:5]:
        try:
            log("Youngpoong login button candidate: " + element_summary(driver, btn))
        except Exception:
            pass
        if yp_click_element(driver, btn):
            clicked = True
            break
    if not clicked:
        try:
            password_inputs[0].send_keys(Keys.ENTER)
            clicked = True
        except Exception:
            pass
    if not clicked:
        dump_debug(driver, "ypscm_login_button_not_found")
        raise RuntimeError("Could not click Youngpoong login button.")

    time.sleep(3)
    log("Youngpoong login click completed.")


def yp_page_has_sales_controls(driver):
    try:
        inputs = yp_find_date_inputs(driver)
        excel_buttons = yp_find_button_by_text(driver, "Excel") + yp_find_button_by_text(driver, "EXCEL") + yp_find_button_by_text(driver, "엑셀")
        search_buttons = yp_find_button_by_text(driver, "검색") + yp_find_button_by_text(driver, "조회")
        return len(inputs) >= 2 and len(excel_buttons) >= 1 and len(search_buttons) >= 1
    except Exception:
        return False


def yp_close_notice_popups(driver):
    """Close Youngpoong notice popups shown after login.

    These are site notices with small close buttons labeled "닫기".
    We do not click broad page titles or large content areas.
    """
    log("Closing Youngpoong notice popups if present...")
    closed = 0
    for _ in range(8):
        candidate = None
        try:
            xpath = (
                "//*[self::button or self::a or self::span or self::input]"
                "[contains(normalize-space(.), '닫기') or @value='닫기' or contains(@title, '닫기')]"
            )
            for e in driver.find_elements(By.XPATH, xpath):
                try:
                    if not visible(e):
                        continue
                    w = e.size.get('width', 0)
                    h = e.size.get('height', 0)
                    txt = (e.text or e.get_attribute('value') or e.get_attribute('title') or '').strip()
                    # Notice close buttons are small. Avoid any large page area.
                    if w <= 160 and h <= 80 and '닫기' in txt:
                        candidate = e
                        break
                except Exception:
                    pass
        except Exception:
            candidate = None

        if candidate is None:
            break

        try:
            log("Closing Youngpoong notice popup by: " + element_summary(driver, candidate))
        except Exception:
            pass
        if yp_click_element(driver, candidate):
            closed += 1
            time.sleep(0.5)
        else:
            break

    log(f"Youngpoong notice popups closed: {closed}")
    return closed


def yp_click_sales_status_menu(driver):
    """Click the main Youngpoong '판매현황' tile/menu automatically.

    User confirmed the home tile is <li class="sell_btn"></li>.
    After clicking it, the sales page should show date fields, 검색, and Excel buttons.
    """
    yp_close_notice_popups(driver)
    log("Trying to open Youngpoong sales status page automatically...")

    click_scripts = [
        """
        const sels = ['li.sell_btn', '.sell_btn'];
        for (const sel of sels) {
          const el = document.querySelector(sel);
          if (el) {
            el.scrollIntoView({block:'center', inline:'center'});
            const r = el.getBoundingClientRect();
            const cx = r.left + r.width/2, cy = r.top + r.height/2;
            for (const type of ['mouseover','mousemove','mousedown','mouseup','click']) {
              try { el.dispatchEvent(new MouseEvent(type, {bubbles:true, cancelable:true, view:window, clientX:cx, clientY:cy})); } catch(e) {}
            }
            try { el.click(); } catch(e) {}
            return {ok:true, method:'css', selector:sel, text:(el.innerText||el.textContent||'').trim(), cls:el.className};
          }
        }
        return {ok:false, reason:'li.sell_btn not found'};
        """,
        """
        const nodes = Array.from(document.querySelectorAll('li, a, div, button, span'));
        const el = nodes.find(x => ((x.innerText||x.textContent||'').includes('판매현황')) && x.offsetWidth > 0 && x.offsetHeight > 0);
        if (el) {
          el.scrollIntoView({block:'center', inline:'center'});
          try { el.click(); } catch(e) {}
          const r = el.getBoundingClientRect();
          const cx = r.left + r.width/2, cy = r.top + r.height/2;
          for (const type of ['mouseover','mousemove','mousedown','mouseup','click']) {
            try { el.dispatchEvent(new MouseEvent(type, {bubbles:true, cancelable:true, view:window, clientX:cx, clientY:cy})); } catch(e) {}
          }
          return {ok:true, method:'text', text:(el.innerText||el.textContent||'').trim(), cls:el.className};
        }
        return {ok:false, reason:'판매현황 text not found'};
        """
    ]

    for js in click_scripts:
        try:
            result = driver.execute_script(js)
            log(f"Youngpoong sales menu click result: {result}")
        except Exception as e:
            log(f"Youngpoong sales menu script failed: {e}")
            result = {'ok': False}
        ok, act = wait_until_activity_quiet(driver, min_quiet_ms=500, max_wait=2.5, label="Youngpoong sales menu")
        for _ in range(12):
            if yp_page_has_sales_controls(driver):
                log("Youngpoong sales page opened automatically.")
                return True
            time.sleep(0.25)

    # Last fallback: known text buttons/links.
    for text in ['판매현황', '기간별 판매 내역 조회']:
        for btn in yp_find_button_by_text(driver, text):
            try:
                log("Youngpoong sales menu candidate: " + element_summary(driver, btn))
            except Exception:
                pass
            if yp_click_element(driver, btn):
                ok, act = wait_until_activity_quiet(driver, min_quiet_ms=500, max_wait=2.5, label="Youngpoong sales text menu")
                for _ in range(12):
                    if yp_page_has_sales_controls(driver):
                        log("Youngpoong sales page opened automatically by text candidate.")
                        return True
                    time.sleep(0.25)

    return False


def prepare_ypscm_sales_page(driver):
    log("Checking Youngpoong sales page controls...")
    if yp_page_has_sales_controls(driver):
        log("Youngpoong sales page appears ready.")
        return

    if yp_click_sales_status_menu(driver):
        return

    dump_debug(driver, "ypscm_sales_page_auto_open_failed")
    raise RuntimeError("Could not automatically open Youngpoong sales page. Expected home tile/menu li.sell_btn was not clickable or sales controls were not detected.")


def yp_find_date_inputs(driver):
    inputs = [e for e in driver.find_elements(By.TAG_NAME, "input") if visible(e)]
    candidates = []
    for e in inputs:
        try:
            value = (e.get_attribute("value") or "").strip()
            name = (e.get_attribute("name") or "").lower()
            idv = (e.get_attribute("id") or "").lower()
            cls = (e.get_attribute("class") or "").lower()
            typ = (e.get_attribute("type") or "").lower()
            placeholder = (e.get_attribute("placeholder") or "").lower()
            score = 0
            if re.match(r"^20\d{2}-\d{2}-\d{2}$", value): score += 100
            if "date" in name or "date" in idv or "date" in cls or "date" in typ or "date" in placeholder: score += 30
            if "from" in name or "to" in name or "start" in name or "end" in name or "시작" in name or "종료" in name: score += 20
            if e.size.get("width", 0) >= 70: score += 5
            candidates.append((score, e))
        except Exception:
            pass
    candidates.sort(key=lambda x: x[0], reverse=True)
    strong = [x[1] for x in candidates if x[0] >= 50]
    if len(strong) >= 2:
        strong_set = set(strong[:4])
        ordered = [e for e in inputs if e in strong_set]
        return ordered[:2]
    fallback = []
    for e in inputs:
        try:
            typ = (e.get_attribute("type") or "text").lower()
            if typ in ("text", "search", "date", "") and e.size.get("width", 0) >= 70:
                fallback.append(e)
        except Exception:
            pass
    return fallback[:2]


def yp_set_date_fields(driver, date_text):
    fields = yp_find_date_inputs(driver)
    if len(fields) < 2:
        dump_debug(driver, f"ypscm_date_fields_not_found_{date_to_ymd(date_text)}")
        raise RuntimeError("Could not find two Youngpoong date input fields.")
    set_input_value(driver, fields[0], date_text)
    time.sleep(0.2)
    set_input_value(driver, fields[1], date_text)
    time.sleep(0.5)
    log(f"Youngpoong date set to {date_text} ~ {date_text}")


def yp_handle_possible_popup(driver):
    try:
        alert = driver.switch_to.alert
        msg = alert.text
        log(f"Youngpoong alert found: {msg}")
        alert.accept()
        if "없" in msg or "no" in msg.lower():
            return "NO_DATA"
        return msg
    except Exception:
        pass
    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text
        if "조회된 내역이 없습니다" in body_text or "내역이 없습니다" in body_text or "데이터가 없습니다" in body_text:
            for text in ["확인", "OK", "Ok"]:
                for btn in yp_find_button_by_text(driver, text):
                    if yp_click_element(driver, btn):
                        log("Youngpoong no-data popup confirmed.")
                        time.sleep(1)
                        return "NO_DATA"
            return "NO_DATA"
    except Exception:
        pass
    return None


def yp_result_info(driver):
    try:
        return driver.execute_script(
            r"""
            return (function(){
                function norm(v){ return (v === null || v === undefined) ? '' : String(v).replace(/\s+/g,' ').trim(); }
                const bodyText = document.body ? norm(document.body.innerText || '') : '';
                const isbnMatches = bodyText.match(/97[89][0-9]{10}/g) || [];
                const rows = Array.from(document.querySelectorAll('tr, .k-master-row, .k-grid-content tr, .k-grid tr')).filter(el => {
                    const r = el.getBoundingClientRect();
                    if (r.width <= 0 || r.height <= 0) return false;
                    const t = norm(el.innerText || el.textContent || '');
                    return /97[89][0-9]{10}/.test(t);
                });
                return {ok:true, isbnCount:isbnMatches.length, rowCount:rows.length, sample:(rows.length ? rows.map(x => norm(x.innerText || x.textContent || '')) : bodyText.split(/\s+(?=97[89][0-9]{10})/)).slice(0,5).map(t => String(t).slice(0,180))};
            })();
            """
        ) or {"ok": False, "isbnCount": 0, "rowCount": 0, "sample": []}
    except Exception as e:
        return {"ok": False, "isbnCount": 0, "rowCount": 0, "sample": [], "error": str(e)}


def yp_has_result_rows(driver):
    info = yp_result_info(driver)
    try:
        return int(info.get("isbnCount") or 0) > 0 or int(info.get("rowCount") or 0) > 0
    except Exception:
        return False


def yp_wait_after_search(driver, ymd):
    max_wait = YPSCM_SEARCH_WAIT_SECONDS
    min_wait = 1.5
    start_time = time.time()
    last_info = None
    log(f"Waiting for Youngpoong result rows/page activity. min={min_wait}s max={max_wait}s")
    while time.time() - start_time < max_wait:
        time.sleep(0.35)
        popup = yp_handle_possible_popup(driver)
        if popup == "NO_DATA":
            return False
        info = yp_result_info(driver)
        last_info = info
        if (time.time() - start_time) >= min_wait and yp_has_result_rows(driver):
            log(f"Youngpoong result rows detected. Evidence: {info}")
            return True
        activity = browser_activity_info(driver)
        if (time.time() - start_time) >= 5 and not activity.get("busy") and int(activity.get("quietMs") or 0) >= 800:
            # If the page is quiet and no rows appeared, treat as no data rather than clicking Excel blindly.
            if not yp_has_result_rows(driver):
                log(f"Youngpoong page is quiet but no result rows were detected. Evidence: {info}")
                return False
    log(f"Youngpoong search wait reached max. Last evidence: {last_info}")
    return yp_has_result_rows(driver)


def yp_click_search(driver, ymd):
    install_browser_activity_tracker(driver)
    log("Clicking Youngpoong Search button...")
    buttons = yp_find_button_by_text(driver, "검색") + yp_find_button_by_text(driver, "조회")
    if not buttons:
        dump_debug(driver, f"ypscm_search_button_not_found_{ymd}")
        raise RuntimeError("Could not find Youngpoong Search button.")
    buttons.sort(key=lambda e: (abs(e.size.get("width", 0) - 50), e.location.get("y", 9999), e.location.get("x", 9999)))
    btn = buttons[0]
    log("Youngpoong Search candidate: " + element_summary(driver, btn))
    if not yp_click_element(driver, btn):
        dump_debug(driver, f"ypscm_search_click_failed_{ymd}")
        raise RuntimeError("Could not click Youngpoong Search button.")
    log("Youngpoong Search clicked. Waiting by result-row/activity signal, not fixed sleep.")


def yp_click_excel(driver, ymd):
    log("Clicking Youngpoong Excel button...")
    before = get_existing_download_files()
    buttons = yp_find_button_by_text(driver, "Excel") + yp_find_button_by_text(driver, "EXCEL") + yp_find_button_by_text(driver, "엑셀")
    if not buttons:
        dump_debug(driver, f"ypscm_excel_button_not_found_{ymd}")
        raise RuntimeError("Could not find Youngpoong Excel button.")
    buttons.sort(key=lambda e: (e.location.get("y", 9999), e.location.get("x", 9999)))
    btn = buttons[0]
    log("Youngpoong Excel candidate: " + element_summary(driver, btn))
    if not yp_click_element(driver, btn):
        dump_debug(driver, f"ypscm_excel_click_failed_{ymd}")
        raise RuntimeError("Could not click Youngpoong Excel button.")
    downloaded = wait_for_new_download(before, timeout=YPSCM_DOWNLOAD_WAIT_SECONDS)
    if not downloaded:
        dump_debug(driver, f"ypscm_download_not_found_{ymd}")
        raise RuntimeError(f"Youngpoong download file not found for {ymd}")
    ext = downloaded.suffix.lower() if downloaded.suffix else ".xls"
    target = DOWNLOAD_DIR / f"{ymd}_영풍문고{ext}"
    move_downloaded_file(downloaded, target)
    return target


def process_ypscm_date(driver, date_text, target_workbook):
    ymd = date_to_ymd(date_text)
    log("=" * 70)
    log(f"Processing {ymd} / 영풍문고 -> 영풍문고")
    close_safe_popups(driver, f"ypscm_before_date_{ymd}")
    yp_set_date_fields(driver, date_text)
    yp_click_search(driver, ymd)
    if not yp_wait_after_search(driver, ymd):
        log(f"Skipped Youngpoong download for {ymd}: no data or result rows not detected. Target sheet remains as template.")
        return
    downloaded = yp_click_excel(driver, ymd)
    copy_download_to_target_sheet(downloaded, target_workbook, YPSCM_PARTNER)


def run_kyobo_batch(selected_dates, output_map):
    log("")
    log("=" * 70)
    log("STEP 1/2: Kyobo SCM V34 batch starts.")
    log("=" * 70)
    driver = build_driver()
    try:
        auto_login_kyobo(driver)
        close_extra_windows_only(driver, "kyobo_after_login")
        close_safe_popups(driver, "kyobo_after_login")
        close_password_change_popup(driver, max_wait=8)
        open_sales_search_page(driver)
        close_safe_popups(driver, "kyobo_sales_search_page")
        close_password_change_popup(driver, max_wait=3)
        log("")
        log("Kyobo Sales Search page is ready. Starting daily downloads into template date files.")
        log("")
        for partner in PARTNERS:
            process_partner_all_dates(driver, partner, selected_dates, output_map)
        log("Kyobo SCM V34 batch finished.")
    finally:
        try:
            driver.quit()
        except Exception:
            pass


def run_ypscm_batch(selected_dates, output_map):
    log("")
    log("=" * 70)
    log("STEP 2/2: Youngpoong SCM batch starts. Auto-clicking Sales Status menu.")
    log("=" * 70)
    driver = build_driver()
    try:
        auto_login_ypscm(driver)
        close_extra_windows_only(driver, "ypscm_after_login")
        close_safe_popups(driver, "ypscm_after_login")
        prepare_ypscm_sales_page(driver)
        close_safe_popups(driver, "ypscm_sales_page")
        for d in selected_dates:
            try:
                process_ypscm_date(driver, d, output_map[d])
            except Exception as e:
                log(f"ERROR on Youngpoong {date_to_ymd(d)}: {e}")
                log(traceback.format_exc())
                dump_debug(driver, f"ypscm_error_{date_to_ymd(d)}")
                # Continue to next date rather than losing already-created Kyobo files.
        log("Youngpoong SCM batch finished.")
    finally:
        try:
            driver.quit()
        except Exception:
            pass





# =====================================================================
# V54 additions: run YES24 SCM first, then Aladin + Kyobo + Youngpoong.
# YES24 has SMS authentication, so the login/auth/sales-screen step is semi-manual.
# After the user completes SMS authentication and opens the sales screen, daily downloads are automated.
# =====================================================================

YES24_START_URL = "https://scm.yes24.com/"
YES24_SALES_ANALYSIS_URL = "https://scm.yes24.com/AnalysisManagement/ListSaleAnalysisGoods"


def yes24_save_debug(driver, tag):
    try:
        debug_dir = SCRIPT_DIR / "yes24_debug"
        debug_dir.mkdir(parents=True, exist_ok=True)
        html_path = debug_dir / f"debug_yes24_{tag}.html"
        png_path = debug_dir / f"debug_yes24_{tag}.png"
        html_path.write_text(driver.page_source, encoding="utf-8", errors="ignore")
        driver.save_screenshot(str(png_path))
        log(f"Saved YES24 debug html: {html_path}")
        log(f"Saved YES24 debug png: {png_path}")
    except Exception as e:
        log(f"Could not save YES24 debug files: {e}")


def yes24_find_login_row(kind=None):
    """
    kind:
      child -> row text should include 예스24 and 아동 / 아이세움
      adult -> row text should include 예스24 and 성인 / 북폴리오
      None  -> first YES24 row
    """
    login_xlsx_path = find_login_xlsx()
    log(f"Using login file for YES24 {kind or 'default'}: {login_xlsx_path}")

    wb = load_workbook(login_xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    col_map = {}

    for idx, row in enumerate(rows):
        normalized = [_normalize_header_name(x) for x in row]
        if "id" in normalized and "password" in normalized:
            header_row_idx = idx
            for c_idx, name in enumerate(normalized):
                if name:
                    col_map[name] = c_idx
            break

    if header_row_idx is None:
        return None

    id_col = col_map.get("id")
    pw_col = col_map.get("password")
    site_col = None
    for k in ("site", "사이트", "siteurl", "url", "주소"):
        if k in col_map:
            site_col = col_map[k]
            break

    if kind == "child":
        kind_words = ["아동", "아이세움", "iseum", "child", "kids"]
    elif kind == "adult":
        kind_words = ["성인", "북폴리오", "bookfolio", "adult", "adlt"]
    else:
        kind_words = []

    candidates = []

    for row in rows[header_row_idx + 1:]:
        if not any(x is not None and str(x).strip() for x in row):
            continue

        row_text = " ".join(str(x) for x in row if x is not None)
        row_text_lower = row_text.lower()
        is_yes24 = ("예스24" in row_text) or ("yes24" in row_text_lower) or ("scm.yes24" in row_text_lower)

        if not is_yes24:
            continue

        if kind_words:
            is_kind = any((w in row_text) or (w.lower() in row_text_lower) for w in kind_words)
            if not is_kind:
                continue

        login_id = row[id_col] if id_col is not None and id_col < len(row) else None
        login_pw = row[pw_col] if pw_col is not None and pw_col < len(row) else None
        login_url = row[site_col] if site_col is not None and site_col < len(row) and row[site_col] else YES24_START_URL

        if login_id is None or login_pw is None:
            log(f"YES24 {kind or 'default'} row found, but id/password is empty. Manual login will be required.")
            return None

        candidates.append({
            "id": str(login_id).strip(),
            "password": str(login_pw).strip(),
            "url": str(login_url).strip() if login_url else YES24_START_URL,
            "row_text": row_text,
        })

    if candidates:
        return candidates[0]

    log(f"YES24 {kind or 'default'} login row was not found in scm_login.xlsx. Manual login will be required.")
    return None

def yes24_set_input(driver, el, value, label):
    driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", el)
    time.sleep(0.2)
    try:
        el.click()
        time.sleep(0.1)
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.BACKSPACE)
        el.send_keys(value)
        time.sleep(0.2)
    except Exception as e:
        log(f"YES24 {label} keyboard input failed: {e}")

    actual = (el.get_attribute("value") or "").strip()
    if actual == value:
        return True

    try:
        driver.execute_script(
            """
            const el = arguments[0];
            const value = arguments[1];
            el.focus();
            const proto = window.HTMLInputElement.prototype;
            const descriptor = Object.getOwnPropertyDescriptor(proto, 'value');
            if (descriptor && descriptor.set) {
                descriptor.set.call(el, value);
            } else {
                el.value = value;
            }
            el.dispatchEvent(new Event('input', {bubbles:true}));
            el.dispatchEvent(new Event('change', {bubbles:true}));
            el.blur();
            """,
            el,
            value,
        )
        time.sleep(0.3)
    except Exception as e:
        log(f"YES24 {label} JS input failed: {e}")

    return (el.get_attribute("value") or "").strip() == value


def yes24_find_login_fields(driver):
    visible_inputs = [el for el in driver.find_elements(By.CSS_SELECTOR, "input") if visible(el)]
    text_inputs = []
    pw_inputs = []

    for el in visible_inputs:
        try:
            typ = (el.get_attribute("type") or "").lower()
            name = (el.get_attribute("name") or "").lower()
            el_id = (el.get_attribute("id") or "").lower()
            placeholder = (el.get_attribute("placeholder") or "").lower()
            rect = driver.execute_script(
                "const r=arguments[0].getBoundingClientRect(); return {x:r.x,y:r.y,w:r.width,h:r.height};",
                el,
            )

            if typ == "password":
                pw_inputs.append((rect["y"], rect["x"], el))

            if typ in ("text", "", "email"):
                score = 0
                if "id" in name or "id" in el_id or "아이디" in placeholder:
                    score += 30
                if rect["w"] > 150 and rect["h"] > 20:
                    score += 5
                text_inputs.append((-score, rect["y"], rect["x"], el))
        except Exception:
            pass

    text_inputs.sort()
    pw_inputs.sort()

    if not text_inputs or not pw_inputs:
        return None, None

    return text_inputs[0][3], pw_inputs[0][2]


def yes24_click_login_button(driver):
    candidates = []
    selectors = ["button", "a", "input[type='button']", "input[type='submit']", "[role='button']"]
    for sel in selectors:
        for el in driver.find_elements(By.CSS_SELECTOR, sel):
            try:
                if not visible(el):
                    continue
                text = (el.text or el.get_attribute("value") or el.get_attribute("title") or "").strip()
                html = (el.get_attribute("outerHTML") or "").lower()
                rect = driver.execute_script(
                    "const r=arguments[0].getBoundingClientRect(); return {x:r.x,y:r.y,w:r.width,h:r.height};",
                    el,
                )
                score = 0
                if "로그인" in text or "login" in text.lower():
                    score += 50
                if "login" in html:
                    score += 20
                if rect["w"] > 100 and rect["h"] > 30:
                    score += 5
                if score > 0:
                    candidates.append((score, rect["y"], rect["x"], el, text))
            except Exception:
                pass

    if not candidates:
        raise RuntimeError("Could not find YES24 login button.")

    candidates.sort(key=lambda x: (-x[0], x[1], x[2]))
    btn = candidates[0][3]
    log(f"Clicking YES24 login button: {candidates[0][4]!r}")

    try:
        btn.click()
    except Exception:
        driver.execute_script("arguments[0].click();", btn)


def yes24_try_auto_login(driver, kind=None):
    row = yes24_find_login_row(kind)
    if not row:
        return False

    login_url = row.get("url") or YES24_START_URL
    if "yes24" not in login_url.lower():
        login_url = YES24_START_URL

    log(f"Opening YES24 login page: {login_url}")
    driver.get(login_url)
    time.sleep(2.5)

    id_el, pw_el = yes24_find_login_fields(driver)
    if id_el is None or pw_el is None:
        log("YES24 login fields were not found. Please login manually.")
        return False

    safe_id = row["id"][:3] + "***" if len(row["id"]) >= 3 else "***"
    log(f"Typing YES24 id/password from scm_login.xlsx. id={safe_id}")
    yes24_set_input(driver, id_el, row["id"], "id")
    yes24_set_input(driver, pw_el, row["password"], "password")
    yes24_click_login_button(driver)
    time.sleep(2.5)

    return True


def yes24_visible(el):
    try:
        return el.is_displayed() and el.size.get("width", 0) > 0 and el.size.get("height", 0) > 0
    except Exception:
        return False


def yes24_close_modal_if_any(driver):
    try:
        buttons = driver.find_elements(By.CSS_SELECTOR, "button, a, [role='button']")
        for b in buttons:
            if not yes24_visible(b):
                continue
            text = (b.text or b.get_attribute("value") or "").strip()
            if text in ("확인", "OK", "Ok"):
                try:
                    b.click()
                    log("Closed YES24 modal.")
                    time.sleep(0.8)
                    return True
                except Exception:
                    try:
                        driver.execute_script("arguments[0].click();", b)
                        log("Closed YES24 modal by JS.")
                        time.sleep(0.8)
                        return True
                    except Exception:
                        pass
    except Exception:
        pass
    return False


def yes24_set_input_by_keyboard(driver, el, value, label):
    driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", el)
    time.sleep(0.2)

    try:
        el.click()
        time.sleep(0.1)
        el.send_keys(Keys.CONTROL, "a")
        time.sleep(0.1)
        el.send_keys(Keys.BACKSPACE)
        time.sleep(0.1)
        el.send_keys(value)
        time.sleep(0.2)
        el.send_keys(Keys.TAB)
        time.sleep(0.4)
    except Exception as e:
        log(f"YES24 {label}: keyboard input failed: {e}")

    actual = (el.get_attribute("value") or "").strip()
    log(f"YES24 {label}: after keyboard input value={actual!r}")

    if actual == value:
        return True

    try:
        driver.execute_script(
            """
            const el = arguments[0];
            const value = arguments[1];
            el.focus();
            const proto = window.HTMLInputElement.prototype;
            const descriptor = Object.getOwnPropertyDescriptor(proto, 'value');
            if (descriptor && descriptor.set) {
                descriptor.set.call(el, value);
            } else {
                el.value = value;
            }
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
            el.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: 'Enter' }));
            el.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: 'Enter' }));
            el.blur();
            """,
            el,
            value,
        )
        time.sleep(0.5)
    except Exception as e:
        log(f"YES24 {label}: JS input failed: {e}")

    actual = (el.get_attribute("value") or "").strip()
    log(f"YES24 {label}: after JS input value={actual!r}")
    return actual == value


def yes24_find_date_inputs(driver):
    all_inputs = [el for el in driver.find_elements(By.CSS_SELECTOR, "input") if yes24_visible(el)]
    candidates = []

    for el in all_inputs:
        try:
            val = (el.get_attribute("value") or "").strip()
            typ = (el.get_attribute("type") or "").lower()
            placeholder = (el.get_attribute("placeholder") or "").strip()
            name = (el.get_attribute("name") or "").lower()
            el_id = (el.get_attribute("id") or "").lower()
            cls = (el.get_attribute("class") or "").lower()
            rect = driver.execute_script(
                "const r=arguments[0].getBoundingClientRect(); return {x:r.x,y:r.y,w:r.width,h:r.height};",
                el,
            )

            if rect["w"] < 80 or rect["h"] < 20:
                continue

            score = 0
            if val.count("-") == 2:
                score += 30
            if placeholder.upper() == "YYYY-MM-DD":
                score += 30
            if typ in ("date", "text"):
                score += 5
            if "date" in name or "date" in el_id or "date" in cls:
                score += 5
            if 150 <= rect["y"] <= 600:
                score += 10

            if score >= 10:
                candidates.append((score, rect["y"], rect["x"], el, val, placeholder, typ, el_id, name, cls))
        except Exception:
            pass

    candidates.sort(key=lambda x: (-x[0], x[1], x[2]))

    log("YES24 date input candidates:")
    for c in candidates[:10]:
        log(f"  score={c[0]}, x={c[2]:.1f}, y={c[1]:.1f}, value={c[4]!r}, placeholder={c[5]!r}, id={c[7]!r}, name={c[8]!r}")

    if len(candidates) < 2:
        raise RuntimeError("Could not find two YES24 date input fields.")

    top_y = candidates[0][1]
    same_row = [c for c in candidates if abs(c[1] - top_y) < 60]

    if len(same_row) >= 2:
        selected = sorted(same_row[:4], key=lambda x: x[2])[:2]
    else:
        selected = sorted(candidates[:4], key=lambda x: x[2])[:2]

    return selected[0][3], selected[1][3]




def yes24_is_sms_auth_page(driver):
    """
    Detect whether the current visible page is still the YES24 SMS auth page.

    V63 fix:
    After SMS authentication succeeds, YES24 can move to the SCM main page while
    stale/hidden text still contains SMS-auth labels. This version requires
    either current URL contains SMSAuth or visible SMS-specific controls remain.
    """
    try:
        url = (driver.current_url or "").lower()
    except Exception:
        url = ""

    if "smsauth" in url:
        return True

    try:
        body_text = driver.execute_script("return document.body ? document.body.innerText : ''") or ""
    except Exception:
        body_text = ""

    if "관리자 SMS 인증" not in body_text and "휴대폰번호" not in body_text:
        return False

    try:
        visible_inputs = [el for el in driver.find_elements(By.CSS_SELECTOR, "input") if yes24_visible(el)]
        long_inputs = []
        for el in visible_inputs:
            try:
                typ = (el.get_attribute("type") or "").lower()
                if typ not in ("text", "tel", "number", "", "password"):
                    continue
                rect = driver.execute_script(
                    "const r=arguments[0].getBoundingClientRect(); return {w:r.width,h:r.height};",
                    el,
                )
                if rect["w"] >= 250 and rect["h"] >= 25:
                    long_inputs.append(el)
            except Exception:
                pass

        buttons = [el for el in driver.find_elements(By.CSS_SELECTOR, "button, a, input[type='button'], input[type='submit']") if yes24_visible(el)]
        button_texts = []
        for b in buttons:
            try:
                button_texts.append((b.text or b.get_attribute("value") or "").strip())
            except Exception:
                pass

        has_request_button = any("인증번호 요청" in t or "인증번호요청" in t for t in button_texts)
        has_auth_button = any(t == "인증" for t in button_texts)

        if len(long_inputs) >= 2 and (has_request_button or has_auth_button):
            return True
    except Exception:
        pass

    return False

def yes24_popup_input(title, prompt, default_value=""):
    """
    Show a topmost Windows input popup and return the entered text.
    This does not save the value anywhere.
    """
    try:
        import tkinter as tk
        from tkinter import simpledialog, messagebox

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        root.lift()
        root.focus_force()

        value = simpledialog.askstring(title, prompt, initialvalue=default_value, parent=root)
        root.destroy()

        if value is None:
            raise RuntimeError(f"User cancelled popup input: {title}")
        return str(value).strip()

    except Exception as e:
        log(f"tkinter input popup failed or cancelled: {e}")
        raise


def yes24_popup_info(title, message):
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        root.lift()
        messagebox.showinfo(title, message, parent=root)
        root.destroy()
    except Exception as e:
        log(f"YES24 info popup failed: {e}")


def yes24_find_sms_inputs(driver):
    """
    YES24 SMS screen has two large input fields:
      1) 휴대폰번호
      2) 인증번호

    V61 fix:
    After clicking '인증번호 요청', the previous scorer sometimes picked the phone
    field again as the auth-code field. We now sort the two visible large text
    inputs by screen position and force first=phone, second=code.
    """
    inputs = [el for el in driver.find_elements(By.CSS_SELECTOR, "input") if yes24_visible(el)]
    candidates = []

    for idx, el in enumerate(inputs):
        try:
            typ = (el.get_attribute("type") or "").lower()
            if typ not in ("text", "tel", "number", "", "password"):
                continue

            rect = driver.execute_script(
                "const r=arguments[0].getBoundingClientRect(); return {x:r.x,y:r.y,w:r.width,h:r.height};",
                el,
            )

            # The SMS fields are the long horizontal inputs.
            if rect["w"] < 250 or rect["h"] < 25:
                continue

            disabled = el.get_attribute("disabled")
            readonly = el.get_attribute("readonly")
            if disabled:
                continue

            nearby = driver.execute_script(
                """
                const el = arguments[0];
                let texts = [];
                let p = el;
                for (let i=0; i<5 && p; i++, p=p.parentElement) {
                    texts.push((p.innerText || p.textContent || '').trim());
                }
                return texts.join(' ');
                """,
                el,
            ) or ""

            value = (el.get_attribute("value") or "").strip()
            candidates.append((rect["y"], rect["x"], el, value, nearby[:120], rect))
        except Exception:
            pass

    candidates.sort(key=lambda x: (x[0], x[1]))

    log("YES24 SMS input candidates by position:")
    for i, item in enumerate(candidates[:6], start=1):
        log(f"  {i}) y={item[0]:.1f}, x={item[1]:.1f}, value={item[3]!r}, nearby={item[4]!r}, rect={item[5]}")

    if len(candidates) < 2:
        raise RuntimeError("Could not find two YES24 SMS phone/auth-code input fields.")

    phone_el = candidates[0][2]
    code_el = candidates[1][2]

    return phone_el, code_el


def yes24_click_auth_confirm_button_direct(driver):
    """
    V71:
    Click the YES24 SMS auth-confirm button robustly.
    It must click the exact '인증' button, never '인증번호 요청'.
    Uses JS elementFromPoint + native click/event dispatch to avoid Selenium
    occasionally clicking the wrong/old element.
    """
    info = driver.execute_script(
        """
        function visible(el){
            const r = el.getBoundingClientRect();
            const st = getComputedStyle(el);
            return r.width > 0 && r.height > 0 &&
                   st.display !== 'none' &&
                   st.visibility !== 'hidden' &&
                   r.bottom >= 0 && r.top <= window.innerHeight &&
                   r.right >= 0 && r.left <= window.innerWidth;
        }

        const selectors = ['button', 'a', 'input[type=button]', 'input[type=submit]', '[role=button]', '.btn'];
        const seen = new Set();
        const candidates = [];

        for (const sel of selectors) {
            document.querySelectorAll(sel).forEach(el => {
                if (seen.has(el)) return;
                seen.add(el);
                if (!visible(el)) return;

                const text = ((el.innerText || el.textContent || el.value || el.title || el.getAttribute('aria-label') || '') + '').trim();
                const html = (el.outerHTML || '').toLowerCase();
                const r = el.getBoundingClientRect();

                // Absolutely exclude the request-code button.
                if (text.includes('요청') || text.includes('인증번호')) return;

                let score = 0;
                if (text === '인증') score += 300;
                else return;

                // The auth button is normally to the right of the second long input.
                if (r.x > window.innerWidth * 0.45) score += 30;
                if (r.y > window.innerHeight * 0.35) score += 20;
                if (r.width >= 60 && r.height >= 25) score += 10;
                if (html.includes('btn')) score += 5;

                candidates.push({el, info:{score, text, x:r.x, y:r.y, w:r.width, h:r.height, html:html.slice(0,180)}});
            });
        }

        candidates.sort((a,b) => b.info.score - a.info.score || b.info.y - a.info.y || b.info.x - a.info.x);

        if (!candidates.length) return {ok:false, reason:'no exact auth button'};

        const c = candidates[0];
        const el = c.el;
        const r = el.getBoundingClientRect();
        const cx = r.left + r.width / 2;
        const cy = r.top + r.height / 2;

        try { el.scrollIntoView({block:'center', inline:'center'}); } catch(e) {}

        const topEl = document.elementFromPoint(cx, cy);
        const target = (topEl && (topEl === el || el.contains(topEl) || topEl.closest('button,a,input,[role=button]') === el))
            ? (topEl.closest('button,a,input,[role=button]') || el)
            : el;

        try { target.focus && target.focus(); } catch(e) {}

        try {
            target.click();
        } catch(e) {
            ['mouseover','mousedown','mouseup','click'].forEach(type => {
                target.dispatchEvent(new MouseEvent(type, {bubbles:true, cancelable:true, view:window, clientX:cx, clientY:cy}));
            });
        }

        return {ok:true, clicked:c.info, topText:((topEl && (topEl.innerText || topEl.textContent || topEl.value)) || '').trim()};
        """
    )
    log(f"YES24 direct auth-confirm click result: {info}")
    return bool(info and info.get("ok"))

def yes24_click_sms_button(driver, wanted_texts, label):
    """
    V71:
    For the final SMS auth-confirm button, use strict direct JS clicking.
    For request-code button, keep the normal candidate selection.
    """
    if label == "인증":
        if yes24_click_auth_confirm_button_direct(driver):
            log("YES24 final SMS auth button clicked by V71 direct method.")
            time.sleep(1.5)
            return True
        raise RuntimeError("Could not click exact YES24 final SMS auth button.")

    candidates = []
    selectors = ["button", "a", "input[type='button']", "input[type='submit']", "[role='button']"]
    is_auth_confirm = (label == "인증")

    for sel in selectors:
        for el in driver.find_elements(By.CSS_SELECTOR, sel):
            try:
                if not yes24_visible(el):
                    continue

                text = (
                    el.text
                    or el.get_attribute("value")
                    or el.get_attribute("title")
                    or el.get_attribute("aria-label")
                    or ""
                ).strip()
                html = (el.get_attribute("outerHTML") or "").lower()
                rect = driver.execute_script(
                    "const r=arguments[0].getBoundingClientRect(); return {x:r.x,y:r.y,w:r.width,h:r.height};",
                    el,
                )

                score = 0

                if is_auth_confirm:
                    # Exact auth-confirm button only.
                    if text == "인증":
                        score += 200
                    elif "인증" in text and "요청" not in text and "번호" not in text:
                        score += 80
                    else:
                        continue

                    # Auth button is usually on the same row as the auth-code input and below the request button.
                    if rect["y"] > 350:
                        score += 20

                else:
                    # Request-code button.
                    for wt in wanted_texts:
                        if wt in text:
                            score += 100
                        if wt.lower() in html:
                            score += 20
                    if "요청" in text:
                        score += 80
                    if score <= 0:
                        continue

                if rect["w"] >= 50 and rect["h"] >= 20:
                    score += 5

                candidates.append((score, rect["y"], rect["x"], el, text, html[:160], rect))
            except Exception:
                pass

    candidates.sort(key=lambda x: (-x[0], x[1], x[2]))

    log(f"YES24 SMS button candidates for {label}:")
    for c in candidates[:5]:
        log(f"  score={c[0]}, text={c[4]!r}, rect={c[6]}, html={c[5]!r}")

    if not candidates:
        raise RuntimeError(f"Could not find YES24 SMS button: {label}")

    btn = candidates[0][3]
    log(f"Clicking YES24 SMS button {label}: text={candidates[0][4]!r}, score={candidates[0][0]}")

    if not yes24_click_element(driver, btn, f"YES24 SMS {label}"):
        raise RuntimeError(f"Could not click YES24 SMS button: {label}")

    time.sleep(1.5)


def yes24_close_auth_success_popup(driver, max_wait=8):
    """
    V65 strict version.

    After clicking YES24 SMS '인증', YES24 may show a success confirmation popup.
    Close only:
      - JavaScript alert, or
      - an on-page button whose visible text is exactly '확인' / 'OK'

    Never click '인증번호 요청' or any button containing '요청' / '인증번호'.
    """
    log("Waiting for YES24 SMS auth success popup. Strict confirm-only close mode.")

    end_time = time.time() + max_wait
    closed_any = False

    while time.time() < end_time:
        # 1) Browser JS alert
        try:
            alert = driver.switch_to.alert
            text = alert.text
            log(f"Closing YES24 JS alert after SMS auth: {text[:120]}")
            alert.accept()
            time.sleep(1.0)
            closed_any = True
            continue
        except Exception:
            pass

        # 2) Exact confirm button only. Do NOT click 인증번호 요청.
        try:
            clicked = driver.execute_script(
                """
                function visible(el){
                    const r = el.getBoundingClientRect();
                    const st = getComputedStyle(el);
                    return r.width > 0 && r.height > 0 &&
                           st.display !== 'none' &&
                           st.visibility !== 'hidden' &&
                           r.bottom >= 0 && r.top <= window.innerHeight &&
                           r.right >= 0 && r.left <= window.innerWidth;
                }

                const selectors = [
                    'button', 'input[type=button]', 'input[type=submit]',
                    '.modal button', '.ui-dialog button', '.swal2-confirm', '[role=button]'
                ];

                const candidates = [];
                const seen = new Set();

                for (const sel of selectors) {
                    document.querySelectorAll(sel).forEach(el => {
                        if (seen.has(el)) return;
                        seen.add(el);
                        if (!visible(el)) return;

                        const text = ((el.innerText || el.textContent || el.value || el.title || el.getAttribute('aria-label') || '') + '').trim();
                        const html = (el.outerHTML || '').toLowerCase();
                        const r = el.getBoundingClientRect();

                        // Strict exclusion
                        if (text.includes('요청') || text.includes('인증번호')) return;

                        let score = 0;
                        if (text === '확인') score += 200;
                        else if (text === 'OK' || text === 'Ok' || text === 'ok') score += 180;
                        else return;

                        // Prefer centered dialog-ish controls.
                        if (r.width >= 35 && r.width <= 220 && r.height >= 20 && r.height <= 90) score += 10;
                        if (r.x > window.innerWidth * 0.20 && r.x < window.innerWidth * 0.90 &&
                            r.y > window.innerHeight * 0.15 && r.y < window.innerHeight * 0.95) score += 20;
                        if (html.includes('modal') || html.includes('dialog') || html.includes('alert') || html.includes('swal')) score += 20;

                        candidates.push({el, info:{score, text, x:r.x, y:r.y, w:r.width, h:r.height}});
                    });
                }

                candidates.sort((a,b) => b.info.score - a.info.score || a.info.y - b.info.y || a.info.x - b.info.x);
                if (!candidates.length) return null;

                const c = candidates[0];
                try { c.el.click(); }
                catch(e) {
                    const r = c.el.getBoundingClientRect();
                    ['mouseover','mousedown','mouseup','click'].forEach(type => {
                        c.el.dispatchEvent(new MouseEvent(type, {bubbles:true, cancelable:true, view:window, clientX:r.left+r.width/2, clientY:r.top+r.height/2}));
                    });
                }
                return c.info;
                """
            )
            if clicked:
                log(f"Closed YES24 auth confirm popup/button: {clicked}")
                time.sleep(1.0)
                closed_any = True
                continue
        except Exception as e:
            log(f"YES24 strict auth confirm close attempt failed: {e}")

        try:
            current_url = (driver.current_url or "").lower()
            if "smsauth" not in current_url:
                if closed_any:
                    log("YES24 URL moved away from SMSAuth after closing confirmation.")
                return closed_any
        except Exception:
            pass

        time.sleep(0.5)

    if not closed_any:
        log("No strict YES24 auth confirmation popup was detected within wait time.")
    return closed_any


def yes24_handle_sms_auth_with_popups(driver, display_label):
    """
    Handle YES24 SMS authentication using Windows popups:
      1) Ask user for phone number in popup.
      2) Fill phone field on YES24 page and click request button.
      3) Ask user for received SMS code in popup.
      4) Fill auth-code field and click 인증 button.
    """
    if not yes24_is_sms_auth_page(driver):
        log(f"YES24 SMS auth page not detected for {display_label}. Skipping SMS popup auth helper.")
        return

    log(f"YES24 SMS auth page detected for {display_label}. Opening phone/code popup flow.")

    phone_el, code_el = yes24_find_sms_inputs(driver)

    phone = yes24_popup_input(
        f"YES24 SMS 인증 - {display_label}",
        "휴대폰번호를 입력해 주세요.\\n하이픈(-)은 제외하고 숫자만 입력하는 것이 좋습니다."
    )
    if not phone:
        raise RuntimeError("YES24 phone number was empty.")

    yes24_set_input(driver, phone_el, phone, "SMS phone")
    yes24_click_sms_button(driver, ["인증번호 요청", "인증번호요청", "요청"], "인증번호 요청")

    # YES24 may show a browser modal/popup after requesting the SMS code.
    # Close it before asking for and entering the auth code.
    yes24_close_modal_if_any(driver)
    time.sleep(0.8)

    code = yes24_popup_input(
        f"YES24 인증번호 입력 - {display_label}",
        "문자로 받은 인증번호를 입력해 주세요."
    )
    if not code:
        raise RuntimeError("YES24 SMS code was empty.")

    # Re-find fields because the page may update after requesting the code.
    # V61: this now forces the second long input field as the auth-code field.
    try:
        _, code_el = yes24_find_sms_inputs(driver)
    except Exception as e:
        log(f"Could not re-find YES24 SMS code field. Using previous code field. Error: {e}")

    yes24_set_input(driver, code_el, code, "SMS auth code")

    # V71: click exact final auth button and retry if the SMS page does not react.
    for auth_try in range(1, 4):
        log(f"YES24 final SMS auth click attempt {auth_try}/3")
        yes24_click_sms_button(driver, ["인증"], "인증")
        time.sleep(1.5)

        # If a success confirmation popup appears, close it.
        yes24_close_auth_success_popup(driver, max_wait=3)

        if not yes24_is_sms_auth_page(driver):
            log("YES24 SMS page disappeared after final auth click.")
            return

        log("YES24 SMS page still visible after final auth click. Retrying if attempts remain.")

    # After closing the success popup, wait until the visible SMS-auth page disappears.
    # Do not show a warning popup just because stale/hidden text remains in the DOM.
    for wait_idx in range(1, 13):
        time.sleep(1.0)
        try:
            current_url = driver.current_url
        except Exception:
            current_url = ""
        still_sms = yes24_is_sms_auth_page(driver)
        log(f"YES24 SMS auth wait {wait_idx}/12: still_sms={still_sms}, url={current_url}")
        if not still_sms:
            log("YES24 SMS authentication appears complete.")
            return

    if yes24_is_sms_auth_page(driver):
        yes24_popup_info(
            "YES24 SMS 인증 확인 필요",
            "아직 YES24 SMS 인증 화면이 실제로 남아 있습니다.\n인증번호가 맞는지 확인한 뒤, 화면에서 직접 인증을 완료해 주세요."
        )
        input("YES24 SMS 인증을 화면에서 직접 완료한 뒤 Enter를 누르세요...")


def yes24_open_sales_analysis_page(driver):
    """
    Open the YES24 sales analysis goods page after manual SMS authentication.
    Target page confirmed by user:
      https://scm.yes24.com/AnalysisManagement/ListSaleAnalysisGoods
    """
    log(f"Opening YES24 sales analysis page: {YES24_SALES_ANALYSIS_URL}")
    driver.get(YES24_SALES_ANALYSIS_URL)
    time.sleep(4.0)
    close_safe_popups(driver, "yes24_sales_analysis_page")

    log(f"YES24 current URL after sales page open: {driver.current_url}")

    # Wait until date inputs appear. If still not authenticated, YES24 may redirect back to SMSAuth/Login.
    last_error = None
    for idx in range(1, 21):
        try:
            start_el, end_el = yes24_find_date_inputs(driver)
            log(f"YES24 sales analysis page ready. Date inputs detected on attempt {idx}.")
            return True
        except Exception as e:
            last_error = e
            current_url = ""
            try:
                current_url = driver.current_url
            except Exception:
                pass
            log(f"YES24 sales page wait attempt {idx}/20: date fields not ready yet. current_url={current_url}")
            time.sleep(1.0)

    raise RuntimeError(
        "YES24 sales analysis page did not show date fields. "
        "SMS authentication may not be complete, or the page may not be the sales analysis screen. "
        f"Last error: {last_error}"
    )


def yes24_set_date_fields(driver, date_text):
    yes24_close_modal_if_any(driver)
    start_el, end_el = yes24_find_date_inputs(driver)

    log(f"Setting YES24 start/end date to {date_text}")
    yes24_set_input_by_keyboard(driver, start_el, date_text, "Start date")
    yes24_set_input_by_keyboard(driver, end_el, date_text, "End date")

    time.sleep(0.8)
    start_el2, end_el2 = yes24_find_date_inputs(driver)
    start_val = (start_el2.get_attribute("value") or "").strip()
    end_val = (end_el2.get_attribute("value") or "").strip()

    log(f"Verify YES24 date fields: start={start_val!r}, end={end_val!r}")

    if start_val != date_text or end_val != date_text:
        log("YES24 date verification failed. Retrying date input once.")
        yes24_set_input_by_keyboard(driver, start_el2, date_text, "Start date retry")
        yes24_set_input_by_keyboard(driver, end_el2, date_text, "End date retry")
        time.sleep(0.8)

        start_el3, end_el3 = yes24_find_date_inputs(driver)
        start_val = (start_el3.get_attribute("value") or "").strip()
        end_val = (end_el3.get_attribute("value") or "").strip()
        log(f"Verify YES24 after retry: start={start_val!r}, end={end_val!r}")

    if start_val != date_text or end_val != date_text:
        raise RuntimeError(f"YES24 date fields were not set correctly. start={start_val}, end={end_val}")


def yes24_click_element(driver, el, name):
    driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", el)
    time.sleep(0.3)

    methods = [
        lambda: el.click(),
        lambda: ActionChains(driver).move_to_element(el).pause(0.2).click().perform(),
        lambda: driver.execute_script("arguments[0].click();", el),
        lambda: driver.execute_script(
            """
            const el = arguments[0];
            el.dispatchEvent(new MouseEvent('mouseover', {bubbles:true}));
            el.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
            el.dispatchEvent(new MouseEvent('mouseup', {bubbles:true}));
            el.dispatchEvent(new MouseEvent('click', {bubbles:true}));
            """,
            el,
        ),
    ]

    for idx, method in enumerate(methods, start=1):
        try:
            method()
            log(f"YES24 {name}: click method {idx} ok")
            return True
        except Exception as e:
            log(f"YES24 {name}: click method {idx} failed: {e}")

    return False


def yes24_find_search_button(driver):
    candidates = []
    for sel in ["button", "a", "input[type='button']", "input[type='submit']", "[role='button']", ".btn"]:
        for el in driver.find_elements(By.CSS_SELECTOR, sel):
            if not yes24_visible(el):
                continue
            try:
                text = (el.text or el.get_attribute("value") or el.get_attribute("title") or "").strip()
                cls = el.get_attribute("class") or ""
                rect = driver.execute_script(
                    "const r=arguments[0].getBoundingClientRect(); return {x:r.x,y:r.y,w:r.width,h:r.height,right:r.right,bottom:r.bottom};",
                    el,
                )
                if "조회" not in text:
                    continue

                score = 10
                if rect["x"] > 1000:
                    score += 10
                if 250 <= rect["y"] <= 750:
                    score += 5
                if "btn" in cls.lower():
                    score += 2
                candidates.append((score, rect["y"], -rect["x"], el, text, cls, rect))
            except Exception:
                pass

    candidates.sort(key=lambda x: (-x[0], x[1], x[2]))
    if not candidates:
        raise RuntimeError("Could not find YES24 search button.")

    best = candidates[0]
    log(f"YES24 search button candidate: text={best[4]!r}, class={best[5]!r}, rect={best[6]}")
    return best[3]


def yes24_click_search(driver):
    yes24_close_modal_if_any(driver)
    btn = yes24_find_search_button(driver)
    if not yes24_click_element(driver, btn, "Search button"):
        raise RuntimeError("Could not click YES24 search button.")
    time.sleep(1.2)
    if yes24_close_modal_if_any(driver):
        raise RuntimeError("YES24 warning modal appeared after search. Date fields may not be set correctly.")


def yes24_find_excel_button(driver):
    candidates = []
    selectors = [
        "button",
        "a",
        "[role='button']",
        ".btn",
        ".btn-info",
        ".btn-primary",
        "i",
        "svg",
        "span",
    ]

    for sel in selectors:
        for el in driver.find_elements(By.CSS_SELECTOR, sel):
            if not yes24_visible(el):
                continue
            try:
                text = (el.text or el.get_attribute("title") or el.get_attribute("aria-label") or "").strip()
                cls = (el.get_attribute("class") or "").lower()
                html = (el.get_attribute("outerHTML") or "").lower()
                rect = driver.execute_script(
                    "const r=arguments[0].getBoundingClientRect(); return {x:r.x,y:r.y,w:r.width,h:r.height,right:r.right,bottom:r.bottom};",
                    el,
                )

                score = 0
                if "excel" in text.lower() or "엑셀" in text:
                    score += 20
                if "download" in text.lower() or "다운" in text:
                    score += 20
                if "excel" in cls or "download" in cls:
                    score += 15
                if "excel" in html or "download" in html:
                    score += 15
                if "fa-download" in html or "mdi-download" in html or "icon-download" in html:
                    score += 20
                if rect["x"] > 900 and 250 <= rect["y"] <= 850:
                    score += 12
                if 10 <= rect["w"] <= 90 and 10 <= rect["h"] <= 90:
                    score += 6
                if "question" in html or "help" in html:
                    score -= 8

                if score >= 12:
                    candidates.append((score, rect["y"], -rect["x"], el, text, cls, rect, html[:160]))
            except Exception:
                pass

    candidates.sort(key=lambda x: (-x[0], x[1], x[2]))

    log("YES24 Excel/download candidates:")
    for c in candidates[:8]:
        log(f"  score={c[0]}, text={c[4]!r}, class={c[5]!r}, rect={c[6]}, html={c[7]!r}")

    if not candidates:
        raise RuntimeError("Could not find YES24 Excel/download button.")

    best_el = candidates[0][3]
    try:
        parent = driver.execute_script(
            """
            let el = arguments[0];
            for (let i=0; i<4 && el; i++) {
                if (el.tagName && ['BUTTON','A'].includes(el.tagName.toUpperCase())) return el;
                if (el.getAttribute && el.getAttribute('role') === 'button') return el;
                el = el.parentElement;
            }
            return arguments[0];
            """,
            best_el,
        )
        if parent:
            best_el = parent
    except Exception:
        pass

    return best_el


def yes24_list_download_files():
    return [p for p in DOWNLOAD_DIR.glob("*") if p.is_file() and not p.name.endswith(".crdownload")]


def yes24_wait_for_new_download(before_files, timeout=120):
    """
    V70 fix:
    YES24 sometimes creates a temporary .tmp file first, then replaces it with the real Excel file.
    Older code could grab that .tmp file too early, and the file vanished before shutil.move().
    This waits only for a stable .xls/.xlsx/.csv file and ignores .tmp/.crdownload files.
    """
    before_set = {str(p.resolve()) for p in before_files}
    end = time.time() + timeout
    last_seen = None

    valid_exts = {".xls", ".xlsx", ".csv"}

    while time.time() < end:
        try:
            partials = list(DOWNLOAD_DIR.glob("*.crdownload")) + list(DOWNLOAD_DIR.glob("*.tmp"))
            current = yes24_list_download_files()
            new_files = [
                p for p in current
                if str(p.resolve()) not in before_set
                and p.exists()
                and p.suffix.lower() in valid_exts
                and not p.name.startswith("~$")
            ]

            if new_files:
                newest = max(new_files, key=lambda p: p.stat().st_mtime)
                last_seen = newest

                # Wait until temp/partial files disappear and size is stable.
                size1 = newest.stat().st_size if newest.exists() else -1
                time.sleep(1.2)
                size2 = newest.stat().st_size if newest.exists() else -2

                partials_after = list(DOWNLOAD_DIR.glob("*.crdownload")) + list(DOWNLOAD_DIR.glob("*.tmp"))

                if newest.exists() and size1 == size2 and size2 > 0 and not partials_after:
                    return newest

        except FileNotFoundError:
            # File may be renamed by Chrome while we are checking it. Retry.
            pass
        except Exception as e:
            log(f"YES24 download wait retry after transient error: {e}")

        time.sleep(1)

    if last_seen and last_seen.exists() and last_seen.suffix.lower() in valid_exts:
        return last_seen

    raise RuntimeError("YES24 Excel download file did not appear as a stable .xls/.xlsx file within timeout.")



def yes24_click_excel_and_save(driver, date_text, file_label):
    before = yes24_list_download_files()
    excel_btn = yes24_find_excel_button(driver)

    log("Clicking YES24 Excel/download icon now.")
    if not yes24_click_element(driver, excel_btn, "Excel/download button"):
        raise RuntimeError("Could not click YES24 Excel/download button.")

    downloaded = yes24_wait_for_new_download(before, timeout=120)

    if downloaded.suffix.lower() not in [".xls", ".xlsx", ".csv"]:
        raise RuntimeError(f"YES24 download is not an Excel file: {downloaded}")

    ext = downloaded.suffix or ".xls"
    ymd = date_text.replace("-", "")
    safe_label = re.sub(r"[^0-9A-Za-z가-힣_\-]+", "_", file_label)
    target = DOWNLOAD_DIR / f"{ymd}_예스24_{safe_label}{ext}"

    # Network drives/security software can briefly lock files. Retry a few times.
    last_error = None
    for attempt in range(1, 8):
        try:
            if not downloaded.exists():
                raise FileNotFoundError(f"Downloaded file disappeared before move: {downloaded}")

            if target.exists():
                target.unlink()

            time.sleep(0.5)
            shutil.move(str(downloaded), str(target))
            log(f"Saved YES24 file: {target}")
            return target

        except Exception as e:
            last_error = e
            log(f"YES24 file move retry {attempt}/7: {e}")
            time.sleep(1.0)

    raise RuntimeError(f"YES24 downloaded file could not be moved to target: {last_error}")



def yes24_partner_config(kind):
    if kind == "child":
        return {
            "sheet_hint": "예스24_아동",
            "sheet_keywords": ["예스24", "아동"],
            "file_label": "아동",
            "display_label": "예스24_아동",
        }
    return {
        "sheet_hint": "예스24_성인",
        "sheet_keywords": ["예스24", "성인"],
        "file_label": "성인",
        "display_label": "예스24_성인",
    }


def yes24_process_one_date(driver, date_text, target_workbook, yes24_partner):
    ymd = date_to_ymd(date_text)
    log("=" * 70)
    log(f"Processing YES24 {ymd}")

    close_safe_popups(driver, f"yes24_before_date_{ymd}")
    yes24_set_date_fields(driver, date_text)

    log("Clicking YES24 search button.")
    yes24_click_search(driver)

    log("Waiting after YES24 search: 7 seconds")
    time.sleep(7)

    result_state = driver.execute_script(
        """
        const text=(document.body&&document.body.innerText||'').replace(/\\s+/g,'');
        return {
          empty: text.includes('총0건') || text.includes('조회된데이터가없습니다'),
          sample: text.slice(-1200)
        };
        """
    )
    if result_state and result_state.get("empty"):
        raise RuntimeError(
            f"YES24 {ymd} 조회 결과가 0건입니다. 원천 데이터가 제공된 뒤 다시 수집해주세요."
        )

    log("YES24 search wait done. Trying Excel download now.")
    downloaded = yes24_click_excel_and_save(driver, date_text, yes24_partner["file_label"])
    copy_download_to_target_sheet(downloaded, target_workbook, yes24_partner)


def run_yes24_one_account(kind, selected_dates, output_map):
    yes24_partner = yes24_partner_config(kind)
    display_label = yes24_partner["display_label"]

    log("")
    log("=" * 70)
    log(f"YES24 {display_label} batch starts. SMS authentication is manual.")
    log("=" * 70)

    driver = build_driver()

    try:
        try:
            yes24_try_auto_login(driver, kind)
        except Exception as e:
            log(f"YES24 {display_label} auto login attempt failed. Manual login is required: {e}")
            driver.get(YES24_START_URL)
            time.sleep(2.0)

        log("")
        log(f"YES24 SMS authentication step for {display_label}:")
        log("1) If SMS authentication page is shown, a Windows popup will ask for phone number.")
        log("2) The program will click '인증번호 요청'.")
        log("3) A second Windows popup will ask for the received SMS code.")
        log("4) The program will click '인증' and then open the sales-analysis page automatically.")
        yes24_handle_sms_auth_with_popups(driver, display_label)

        yes24_open_sales_analysis_page(driver)

        for d in selected_dates:
            try:
                yes24_process_one_date(driver, d, output_map[d], yes24_partner)
            except Exception as e:
                log(f"ERROR on YES24 {display_label} {date_to_ymd(d)}: {e}")
                log(traceback.format_exc())
                yes24_save_debug(driver, f"yes24_{display_label}_{date_to_ymd(d)}")
                raise

        log(f"YES24 {display_label} batch finished.")

    finally:
        try:
            driver.quit()
        except Exception:
            pass


def run_yes24_batch(selected_dates, output_map):
    log("")
    log("=" * 70)
    log("STEP 1/4: YES24 SCM batch starts. Child first, adult next.")
    log("=" * 70)

    run_yes24_one_account("child", selected_dates, output_map)
    run_yes24_one_account("adult", selected_dates, output_map)

    log("YES24 child/adult SCM batch finished.")


# =====================================================================
# V46 additions: run Aladin SCM child/adult before Kyobo + Youngpoong.
# - Aladin child account row: scm_login.xlsx row containing 알라딘 and 아동 or 아이세움
# - Aladin adult account row: scm_login.xlsx row containing 알라딘 and 성인 or 북폴리오
# - Aladin supplier login ID field is exactly name="SupplierId".
# - Aladin does not download Excel from the site. It reads the result table and creates XLSX.
# =====================================================================

ALADIN_START_URL = "https://www.aladin.co.kr/supplier/wStatSalesBook.aspx"


def aladin_find_login_row(kind):
    login_xlsx_path = find_login_xlsx()
    log(f"Using login file for Aladin {kind}: {login_xlsx_path}")

    wb = load_workbook(login_xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    col_map = {}

    for idx, row in enumerate(rows):
        normalized = [_normalize_header_name(x) for x in row]
        if "id" in normalized and "password" in normalized:
            header_row_idx = idx
            for c_idx, name in enumerate(normalized):
                if name:
                    col_map[name] = c_idx
            break

    if header_row_idx is None:
        raise RuntimeError("Could not find id/password header row in scm_login.xlsx for Aladin.")

    id_col = col_map.get("id")
    pw_col = col_map.get("password")

    site_col = None
    for k in ("site", "사이트", "siteurl", "url", "주소"):
        if k in col_map:
            site_col = col_map[k]
            break

    if kind == "child":
        keywords = ["아동", "아이세움", "iseum", "child", "kids"]
    else:
        keywords = ["성인", "북폴리오", "bookfolio", "adult", "adlt"]

    candidates = []

    for row in rows[header_row_idx + 1:]:
        if not any(x is not None and str(x).strip() for x in row):
            continue

        row_text = " ".join(str(x) for x in row if x is not None)
        row_text_lower = row_text.lower()

        is_aladin = ("알라딘" in row_text) or ("aladin" in row_text_lower)
        is_kind = any((k in row_text) or (k.lower() in row_text_lower) for k in keywords)

        if is_aladin and is_kind:
            login_id = row[id_col] if id_col is not None and id_col < len(row) else None
            login_pw = row[pw_col] if pw_col is not None and pw_col < len(row) else None
            login_url = row[site_col] if site_col is not None and site_col < len(row) and row[site_col] else ALADIN_START_URL

            if login_id is None or login_pw is None:
                raise RuntimeError(f"Aladin {kind} row found, but id/password is empty in scm_login.xlsx")

            candidates.append({
                "id": str(login_id).strip(),
                "password": str(login_pw).strip(),
                "url": str(login_url).strip(),
                "row_text": row_text,
            })

    if candidates:
        return candidates[0]

    raise RuntimeError(
        f"Could not find Aladin {kind} login row in scm_login.xlsx. "
        f"Please add row text like 알라딘_아동/알라딘_아이세움 or 알라딘_성인/알라딘_북폴리오."
    )


def aladin_set_text_input(el, value):
    el.click()
    time.sleep(0.1)
    try:
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.BACKSPACE)
        el.send_keys(value)
    except Exception:
        pass


def aladin_click_login_button(driver):
    """
    Aladin supplier login button confirmed by user:
      <input type="image" src="//image.aladin.co.kr/img/provider/bu_login.gif" width="43" height="43" border="0" tabindex="3">
    """
    log("Clicking Aladin login button by exact image button.")

    login_btn = None

    exact_selectors = [
        "input[type='image'][src*='bu_login.gif']",
        "input[type='image'][tabindex='3']",
        "input[src*='provider/bu_login.gif']",
    ]

    for sel in exact_selectors:
        try:
            els = [x for x in driver.find_elements(By.CSS_SELECTOR, sel) if visible(x)]
            if els:
                login_btn = els[0]
                log(f"Found Aladin login image button by CSS: {sel}")
                break
        except Exception as e:
            log(f"Aladin exact login selector failed {sel}: {e}")

    if login_btn is None:
        candidates = []
        selectors = [
            "input[type='image']",
            "input[type='submit'][value*='로그인']",
            "input[type='button'][value*='로그인']",
            "button",
            "a",
            "[role='button']",
        ]

        for sel in selectors:
            for el in driver.find_elements(By.CSS_SELECTOR, sel):
                try:
                    if not visible(el):
                        continue

                    text = (
                        el.text
                        or el.get_attribute("value")
                        or el.get_attribute("alt")
                        or el.get_attribute("title")
                        or ""
                    ).strip()
                    html = (el.get_attribute("outerHTML") or "").lower()
                    rect = driver.execute_script(
                        "const r=arguments[0].getBoundingClientRect(); return {w:r.width,h:r.height,x:r.x,y:r.y};",
                        el,
                    )

                    score = 0
                    if "bu_login.gif" in html:
                        score += 100
                    if "로그인" in text or "login" in text.lower():
                        score += 50
                    if "login" in html or "로그인" in html:
                        score += 30
                    if rect["x"] < 700 and 250 <= rect["y"] <= 750:
                        score += 20
                    if 30 <= rect["w"] <= 90 and 30 <= rect["h"] <= 90:
                        score += 10

                    if score > 0:
                        candidates.append((score, rect["y"], rect["x"], el, text, html[:160]))
                except Exception:
                    pass

        if not candidates:
            raise RuntimeError("Could not find Aladin supplier login image button.")

        candidates.sort(key=lambda x: (-x[0], x[1], x[2]))
        login_btn = candidates[0][3]
        log(f"Aladin login button candidate: text={candidates[0][4]!r}, score={candidates[0][0]}, html={candidates[0][5]!r}")

    driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", login_btn)
    time.sleep(0.3)

    click_methods = [
        lambda: login_btn.click(),
        lambda: ActionChains(driver).move_to_element(login_btn).pause(0.2).click().perform(),
        lambda: driver.execute_script("arguments[0].click();", login_btn),
        lambda: driver.execute_script(
            """
            const el = arguments[0];
            ['mouseover','mousedown','mouseup','click'].forEach(t => {
                el.dispatchEvent(new MouseEvent(t, {view:window, bubbles:true, cancelable:true}));
            });
            """,
            login_btn,
        ),
    ]

    last_error = None
    for idx, method in enumerate(click_methods, start=1):
        try:
            log(f"Aladin login click method {idx}...")
            method()
            time.sleep(1.5)
            return
        except Exception as e:
            last_error = e
            log(f"Aladin login click method {idx} failed: {e}")

    raise RuntimeError(f"Could not click Aladin login image button. Last error: {last_error}")


def aladin_open_sales_statistics_page(driver):
    """
    Open Aladin sales statistics page after login.
    User confirmed menu link:
      <a href="/supplier/wStatSalesBook.aspx">판매 통계</a>
    """
    log("Opening Aladin sales statistics page.")

    # First try clicking the exact menu link if it is visible.
    try:
        links = driver.find_elements(By.CSS_SELECTOR, "a[href='/supplier/wStatSalesBook.aspx'], a[href*='wStatSalesBook.aspx']")
        for a in links:
            if visible(a):
                log("Clicking Aladin sales statistics link: /supplier/wStatSalesBook.aspx")
                try:
                    a.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", a)
                time.sleep(2.5)
                close_safe_popups(driver, "aladin_sales_stats_after_menu")
                if "cmdGetStat" in driver.page_source:
                    log("Aladin sales statistics page opened by menu link.")
                    return
    except Exception as e:
        log(f"Aladin sales statistics link click failed: {e}")

    # Fallback: direct URL.
    log("Opening Aladin sales statistics page by direct URL.")
    driver.get(ALADIN_START_URL)
    time.sleep(2.5)
    close_safe_popups(driver, "aladin_sales_stats_direct")
    if "cmdGetStat" in driver.page_source:
        log("Aladin sales statistics page opened by direct URL.")
        return

    raise RuntimeError("Could not open Aladin sales statistics page.")

def aladin_auto_login(driver, login_row, label, output_dir):
    login_url = login_row.get("url") or ALADIN_START_URL
    if "aladin" not in login_url.lower():
        login_url = ALADIN_START_URL

    log(f"Opening Aladin page for {label}: {login_url}")
    driver.get(login_url)
    time.sleep(2.5)

    if "cmdGetStat" in driver.page_source:
        log("Aladin statistics page appears already accessible.")
        return

    # Important: Aladin has a global search input at the top.
    # Supplier ID field is exactly:
    # <input class="form1" type="text" name="SupplierId">
    supplier_id_el = None
    try:
        supplier_id_el = driver.find_element(By.NAME, "SupplierId")
        if not visible(supplier_id_el):
            supplier_id_el = None
    except Exception:
        supplier_id_el = None

    if supplier_id_el is None:
        try:
            supplier_id_el = driver.find_element(By.CSS_SELECTOR, "input[name='SupplierId'], input.form1[name='SupplierId']")
        except Exception:
            supplier_id_el = None

    pw_el = None
    try:
        pw_inputs = [x for x in driver.find_elements(By.CSS_SELECTOR, "input[type='password']") if visible(x)]
        if pw_inputs:
            pw_el = pw_inputs[0]
    except Exception:
        pw_el = None

    if supplier_id_el is None or pw_el is None:
        log("Exact Aladin supplier login fields were not found.")
        aladin_save_debug(driver, f"aladin_login_fields_not_found_{label}")
        raise RuntimeError("Could not find Aladin SupplierId/password fields.")

    safe_id = login_row["id"][:3] + "***" if len(login_row["id"]) >= 3 else "***"
    log(f"Typing Aladin SupplierId/password from scm_login.xlsx. id={safe_id}")
    aladin_set_text_input(supplier_id_el, login_row["id"])
    aladin_set_text_input(pw_el, login_row["password"])
    aladin_click_login_button(driver)
    time.sleep(3.0)
    close_safe_popups(driver, f"aladin_after_login_{label}")

    # Explicitly open statistics page after login.
    try:
        aladin_open_sales_statistics_page(driver)
    except Exception:
        aladin_save_debug(driver, f"aladin_stats_not_ready_{label}")
        raise


def aladin_option_texts(select_el):
    try:
        return [o.text.strip() for o in Select(select_el).options]
    except Exception:
        return []


def aladin_find_date_selects(driver):
    selects = [el for el in driver.find_elements(By.CSS_SELECTOR, "select") if visible(el)]
    candidates = []

    for el in selects:
        try:
            texts = aladin_option_texts(el)
            rect = driver.execute_script(
                "const r=arguments[0].getBoundingClientRect(); return {x:r.x,y:r.y,w:r.width,h:r.height};",
                el,
            )

            score = 0
            if any(t in texts for t in ["2026", "2025", "2024"]):
                score += 30
            if any(t in texts for t in ["1", "2", "3", "6", "12"]):
                score += 20
            if rect["w"] <= 120:
                score += 5

            if score >= 20:
                candidates.append((score, rect["y"], rect["x"], el, texts[:8], rect))
        except Exception:
            pass

    candidates.sort(key=lambda x: (x[1], x[2]))

    log("Aladin date select candidates:")
    for c in candidates[:10]:
        log(f"  score={c[0]}, x={c[2]:.1f}, y={c[1]:.1f}, options={c[4]}, rect={c[5]}")

    if len(candidates) < 6:
        raise RuntimeError(f"Could not find six Aladin date select fields. Found {len(candidates)}.")

    top_y = candidates[0][1]
    same_row = [c for c in candidates if abs(c[1] - top_y) < 80]
    if len(same_row) >= 6:
        selected = sorted(same_row[:6], key=lambda x: x[2])
    else:
        selected = sorted(candidates[:6], key=lambda x: (x[1], x[2]))

    return [c[3] for c in selected[:6]]


def aladin_select_value(select_el, value):
    sel = Select(select_el)
    value = str(value)

    for txt in (value, f"{value}년", f"{value}월", f"{value}일"):
        try:
            sel.select_by_visible_text(txt)
            return
        except Exception:
            pass

    for val in (value, value.zfill(2)):
        try:
            sel.select_by_value(val)
            return
        except Exception:
            pass

    for opt in sel.options:
        if opt.text.strip().startswith(value):
            opt.click()
            return

    raise RuntimeError(f"Could not select Aladin date value {value}.")


def aladin_set_date_fields(driver, date_text):
    d = datetime.strptime(date_text, "%Y-%m-%d").date()
    date_key = d.strftime("%Y%m%d")
    year = str(d.year)
    month = str(d.month)
    day = str(d.day)

    selects = aladin_find_date_selects(driver)
    log(f"Setting Aladin date to {date_key}")

    aladin_select_value(selects[0], year)
    aladin_select_value(selects[1], month)
    aladin_select_value(selects[2], day)
    aladin_select_value(selects[3], year)
    aladin_select_value(selects[4], month)
    aladin_select_value(selects[5], day)
    time.sleep(0.5)



def aladin_save_debug(driver, tag):
    try:
        debug_dir = SCRIPT_DIR / "aladin_debug"
        debug_dir.mkdir(parents=True, exist_ok=True)

        html_path = debug_dir / f"debug_aladin_{tag}.html"
        png_path = debug_dir / f"debug_aladin_{tag}.png"

        html_path.write_text(driver.page_source, encoding="utf-8", errors="ignore")
        driver.save_screenshot(str(png_path))

        log(f"Saved Aladin debug html: {html_path}")
        log(f"Saved Aladin debug png: {png_path}")
    except Exception as e:
        log(f"Could not save Aladin debug files: {e}")


def aladin_click_element(driver, el, name):
    driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", el)
    time.sleep(0.2)

    methods = [
        lambda: el.click(),
        lambda: ActionChains(driver).move_to_element(el).pause(0.2).click().perform(),
        lambda: driver.execute_script("arguments[0].click();", el),
        lambda: driver.execute_script(
            """
            const el = arguments[0];
            el.dispatchEvent(new MouseEvent('mouseover', {bubbles:true}));
            el.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
            el.dispatchEvent(new MouseEvent('mouseup', {bubbles:true}));
            el.dispatchEvent(new MouseEvent('click', {bubbles:true}));
            """,
            el,
        ),
    ]

    last_error = None
    for idx, method in enumerate(methods, start=1):
        try:
            method()
            log(f"{name}: click method {idx} ok")
            return True
        except Exception as e:
            last_error = e
            log(f"{name}: click method {idx} failed: {e}")

    log(f"{name}: all click methods failed. Last error: {last_error}")
    return False


def aladin_find_search_button(driver):
    for by, value in [
        (By.ID, "cmdGetStat"),
        (By.NAME, "cmdGetStat"),
        (By.CSS_SELECTOR, 'input[type="submit"][value="조회"]'),
    ]:
        try:
            btn = driver.find_element(by, value)
            log(f"Found Aladin search button: {by}={value}")
            return btn
        except Exception:
            pass

    raise RuntimeError("Could not find Aladin 조회 button: cmdGetStat.")


def aladin_click_search(driver):
    btn = aladin_find_search_button(driver)
    if aladin_click_element(driver, btn, "Aladin search button cmdGetStat"):
        return

    try:
        driver.execute_script("document.getElementById('cmdGetStat').click();")
        log("Aladin search button cmdGetStat: document.getElementById click ok")
        return
    except Exception as e:
        log(f"Aladin search button cmdGetStat: JS id click failed: {e}")

    raise RuntimeError("Could not click Aladin 조회 button cmdGetStat.")


def aladin_extract_table_rows(driver):
    expected_header = ["출판사", "도서명", "ISBN", "저자", "정가", "판매권수", "추이", "경향"]

    all_rows = driver.find_elements(By.CSS_SELECTOR, "tr")
    parsed = []

    for idx, tr in enumerate(all_rows):
        cells = tr.find_elements(By.CSS_SELECTOR, "th,td")
        if not cells:
            continue

        values = []
        for cell in cells:
            txt = cell.text.replace("\xa0", " ").strip()
            txt = "\n".join([part.strip() for part in txt.splitlines() if part.strip()])
            values.append(txt)

        if any(v for v in values):
            parsed.append((idx, tr, values))

    log(f"Aladin total visible/nonblank TR candidates: {len(parsed)}")

    header_pos = None
    for p_idx, (dom_idx, tr, values) in enumerate(parsed):
        compact = [v.strip() for v in values if v.strip()]
        if compact[:8] == expected_header:
            header_pos = p_idx
            log(f"Found Aladin exact header row at parsed index {p_idx}, dom tr index {dom_idx}")
            break

        if all(h in compact for h in expected_header) and len(compact) <= 12:
            header_pos = p_idx
            log(f"Found Aladin header row by loose matching at parsed index {p_idx}, dom tr index {dom_idx}")
            break

    if header_pos is None:
        log("Could not find Aladin exact header row. First 20 parsed rows:")
        for i, (_, _, vals) in enumerate(parsed[:20]):
            log(f"  row {i}: {vals[:12]}")
        raise RuntimeError("Could not find Aladin result header row.")

    rows = [expected_header]

    for p_idx in range(header_pos + 1, len(parsed)):
        dom_idx, tr, values = parsed[p_idx]
        compact = [v.strip() for v in values if v.strip()]
        if not compact:
            continue

        first = compact[0]
        joined = " ".join(compact)

        if first in ("총계", "총 계") or joined.startswith("총 계") or joined.startswith("총계"):
            log(f"Reached Aladin total row at parsed index {p_idx}: {compact[:8]}")
            break

        if first in ("회사소개", "채용안내", "이용약관") or "Aladin Communication" in joined:
            log(f"Reached Aladin footer row at parsed index {p_idx}: {compact[:8]}")
            break

        if len(compact) >= 6:
            row = compact[:8]
            while len(row) < 8:
                row.append("")
            if row[:8] != expected_header:
                rows.append(row)

    if len(rows) <= 1:
        raise RuntimeError("No Aladin data rows extracted after header.")

    log(f"Extracted Aladin clean rows including header: {len(rows)}")
    return rows


def aladin_save_rows_to_xlsx(rows, target_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Aladin"

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    header_fill = PatternFill("solid", fgColor="9DC3E6")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9EAF7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = {1: 16, 2: 32, 3: 16, 4: 32, 5: 12, 6: 12, 7: 10, 8: 10}
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(c, 14)

    ws.row_dimensions[1].height = 22
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 45

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)




def aladin_partner_config(kind):
    if kind == "child":
        return {
            "sheet_hint": "알라딘_아동",
            "sheet_keywords": ["알라딘", "아동"],
            "file_label": "아이세움_아동",
        }
    return {
        "sheet_hint": "알라딘_성인",
        "sheet_keywords": ["알라딘", "성인"],
        "file_label": "북폴리오_성인",
    }


def aladin_resolve_or_create_target_worksheet(dst_wb, aladin_partner):
    sheet_hint = aladin_partner["sheet_hint"]
    keywords = aladin_partner["sheet_keywords"]
    available = []

    # 1) Exact sheet name.
    for ws in dst_wb.Worksheets:
        try:
            name = str(ws.Name).strip()
            available.append(name)
            if name == sheet_hint:
                return ws, name, "exact-hint"
        except Exception:
            pass

    # 2) Normalized exact sheet name.
    norm_hint = normalize_korean_sheet_name(sheet_hint)
    for ws in dst_wb.Worksheets:
        try:
            name = str(ws.Name).strip()
            if normalize_korean_sheet_name(name) == norm_hint:
                return ws, name, "normalized-hint"
        except Exception:
            pass

    # 3) Keyword match: 알라딘 + 아동 / 알라딘 + 성인.
    norm_keywords = [normalize_korean_sheet_name(k) for k in keywords]
    matches = []
    for ws in dst_wb.Worksheets:
        try:
            name = str(ws.Name).strip()
            norm_name = normalize_korean_sheet_name(name)
            if all(k in norm_name for k in norm_keywords):
                matches.append((ws, name))
        except Exception:
            pass

    if matches:
        return matches[0][0], matches[0][1], "keyword-match"

    # 4) If template does not have the Aladin sheet, create it.
    log(f"Aladin target sheet not found. Creating new sheet: {sheet_hint}. Available={available}")
    new_ws = dst_wb.Worksheets.Add(After=dst_wb.Worksheets(dst_wb.Worksheets.Count))
    try:
        new_ws.Name = sheet_hint[:31]
    except Exception:
        new_ws.Name = ("알라딘_" + str(dst_wb.Worksheets.Count))[:31]
    return new_ws, str(new_ws.Name), "created-new-sheet"


def aladin_copy_xlsx_to_date_workbook(aladin_file: Path, target_workbook: Path, aladin_partner):
    aladin_file = Path(aladin_file).resolve()
    target_workbook = Path(target_workbook).resolve()

    if not aladin_file.exists():
        raise FileNotFoundError(f"Aladin source file not found: {aladin_file}")
    if not target_workbook.exists():
        raise FileNotFoundError(f"Target date workbook not found: {target_workbook}")

    sheet_hint = aladin_partner["sheet_hint"]
    log(f"Copying Aladin file into {target_workbook.name} / {sheet_hint}: {aladin_file.name}")

    excel = get_excel_app()
    src_wb = None
    dst_wb = None

    try:
        src_wb = excel.Workbooks.Open(str(aladin_file), UpdateLinks=0, ReadOnly=True)
        dst_wb = excel.Workbooks.Open(str(target_workbook), UpdateLinks=0, ReadOnly=False)

        src_ws = src_wb.Worksheets(1)
        dst_ws, actual_sheet_name, method = aladin_resolve_or_create_target_worksheet(dst_wb, aladin_partner)
        log(f"Resolved Aladin target sheet by {method}: {actual_sheet_name}")

        try:
            dst_ws.Activate()
        except Exception:
            pass

        dst_ws.Cells.Clear()
        src_ws.UsedRange.Copy()
        dst_ws.Range("A1").PasteSpecial(Paste=-4104)  # xlPasteAll
        try:
            dst_ws.Range("A1").PasteSpecial(Paste=8)  # xlPasteColumnWidths
        except Exception:
            pass

        excel.CutCopyMode = False

        try:
            rows_count = src_ws.UsedRange.Rows.Count
            cols_count = src_ws.UsedRange.Columns.Count
            for r in range(1, min(rows_count, 500) + 1):
                dst_ws.Rows(r).RowHeight = src_ws.Rows(r).RowHeight
            for c in range(1, min(cols_count, 120) + 1):
                dst_ws.Columns(c).ColumnWidth = src_ws.Columns(c).ColumnWidth
        except Exception:
            pass

        dst_wb.Save()
        log(f"Saved date workbook with Aladin sheet: {target_workbook} / sheet={actual_sheet_name}")

    finally:
        try:
            if src_wb is not None:
                src_wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if dst_wb is not None:
                dst_wb.Close(SaveChanges=True)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass

def aladin_process_one_date(driver, date_text, output_dir, file_label, target_workbook, aladin_partner):
    ymd = date_to_ymd(date_text)

    log("=" * 70)
    log(f"Processing Aladin {ymd}")
    close_safe_popups(driver, f"aladin_before_date_{ymd}")

    aladin_set_date_fields(driver, date_text)

    log("Clicking Aladin 조회 button.")
    aladin_click_search(driver)

    log("Waiting after Aladin search: 4 seconds")
    time.sleep(4)

    rows = aladin_extract_table_rows(driver)
    safe_label = re.sub(r"[^0-9A-Za-z가-힣_\-]+", "_", file_label)
    target = output_dir / f"{ymd}_알라딘_{safe_label}.xlsx"
    aladin_save_rows_to_xlsx(rows, target)
    log(f"Saved Aladin file: {target}")

    # V50: also copy the Aladin sheet into the date workbook, not only into download folder.
    aladin_copy_xlsx_to_date_workbook(target, target_workbook, aladin_partner)


def run_aladin_one_account(kind, label, output_folder, selected_dates, output_map):
    # V49: Use the existing common download folder instead of creating separate Aladin folders.
    output_dir = DOWNLOAD_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    aladin_partner = aladin_partner_config(kind)
    login_row = aladin_find_login_row(kind)

    log("")
    log("=" * 70)
    log(f"ALADIN {label} batch starts.")
    log(f"Output folder: {output_dir}")
    log("=" * 70)

    driver = build_driver()
    try:
        aladin_auto_login(driver, login_row, label, output_dir)

        for d in selected_dates:
            try:
                aladin_process_one_date(driver, d, output_dir, label, output_map[d], aladin_partner)
            except Exception as e:
                log(f"ERROR on Aladin {label} {date_to_ymd(d)}: {e}")
                log(traceback.format_exc())
                aladin_save_debug(driver, f"aladin_{label}_{date_to_ymd(d)}")
                raise

        log(f"ALADIN {label} batch finished.")
    finally:
        try:
            driver.quit()
        except Exception:
            pass


def run_aladin_batch(selected_dates, output_map):
    log("")
    log("=" * 70)
    log("STEP 1/3: Aladin SCM batch starts. Child first, adult next.")
    log("=" * 70)

    run_aladin_one_account("child", "아이세움_아동", "aladin_download_child", selected_dates, output_map)
    run_aladin_one_account("adult", "북폴리오_성인", "aladin_download_adlt", selected_dates, output_map)

    log("Aladin SCM batch finished.")



def show_completion_popup(title, message):
    """
    Show a Windows-style completion popup.
    If tkinter is unavailable, fall back to PowerShell MessageBox.
    If both fail, just log and continue.
    """
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        messagebox.showinfo(title, message, parent=root)
        root.destroy()
        return
    except Exception as e:
        log(f"tkinter completion popup failed: {e}")

    try:
        import subprocess
        ps = (
            "Add-Type -AssemblyName PresentationFramework; "
            f"[System.Windows.MessageBox]::Show('{message}', '{title}')"
        )
        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps],
            check=False,
        )
    except Exception as e:
        log(f"PowerShell completion popup failed: {e}")


def main():
    ensure_dirs()
    selected_dates = select_date_range()
    output_map = create_date_workbooks(selected_dates)

    log("SCM downloader only")
    log("VERSION: V72-KYOBO-STALE-DATE-GRID-GUARD")
    log(f"Current directory: {Path.cwd()}")
    log(f"Script path: {Path(__file__).resolve()}")
    log(f"SCRIPT_DIR/BAT_RESULT_DIR: {SCRIPT_DIR}")
    log(f"PHYSICAL_PY_DIR: {Path(__file__).resolve().parent}")
    log(f"DOWNLOAD_DIR: {DOWNLOAD_DIR}")
    log(f"MASTER_DATA_DIR: {MASTER_DATA_DIR}")
    log(f"DATE_COUNT: {len(selected_dates)}")
    log(f"DATE_RANGE: {selected_dates[0]} ~ {selected_dates[-1]}")
    log("")

    try:
        run_yes24_batch(selected_dates, output_map)
        run_aladin_batch(selected_dates, output_map)
        run_kyobo_batch(selected_dates, output_map)
        run_ypscm_batch(selected_dates, output_map)

        log("=" * 70)
        log("SCM 다운로드만 완료되었습니다.")
        log("생성/갱신된 날짜별 파일:")
        for d in selected_dates:
            log(f" - {output_map[d]}")
        log("=" * 70)

    except Exception as e:
        log("")
        log(f"ERROR: {e}")
        log(traceback.format_exc())
        log("SCM 다운로드 중 오류가 발생했습니다.")
        raise SystemExit(1)


if __name__ == "__main__":
    main()
