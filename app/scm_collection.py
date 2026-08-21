from __future__ import annotations

import hashlib
import shutil
import threading
import time
import uuid
from copy import deepcopy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl

from .config import BASE_DIR, YES24_DOWNLOAD_DIR
from .scm_import import ScmLedger, sync_scm_dataset
from .yes24_demographics import Yes24Demographics, parse_yes24_demographic_files
from .scm_collectors.workbook_parser import parse_date_workbooks
from .scm_credentials import get_account, get_recipient


CLIENTS = {
    "YES24": {"name": "YES24", "runner": "run_yes24_batch"},
    "KYOBO": {"name": "교보문고", "runner": "run_kyobo_batch"},
    "YPBOOKS": {"name": "영풍문고", "runner": "run_ypscm_batch"},
    "ALADIN": {"name": "알라딘", "runner": "run_aladin_batch"},
}
COLLECTION_ORDER = ("YES24", "KYOBO", "YPBOOKS", "ALADIN")
SCM_ROOT = Path(r"Y:\출판사업본부\05. 영업 실적")
SCM_INPUT_DIR = SCM_ROOT / "01. 실판매"
SCM_MASTER_DIR = SCM_ROOT / "80. Master Data"
SCM_TEMPLATE = SCM_INPUT_DIR / "SCM 복사 양식.xlsx"
SCM_LOGIN_FILE = SCM_MASTER_DIR / "scm_login.xlsx"
RUN_ROOT = BASE_DIR / "data" / "scm_runs"
MAX_RECOLLECT_DAYS = 31

_jobs: dict[str, dict[str, Any]] = {}
_lock = threading.RLock()
_input_waiters: dict[str, dict[str, Any]] = {}


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _job_update(job_id: str, **values: Any) -> None:
    with _lock:
        _jobs[job_id].update(values)


def _client_update(job_id: str, code: str, **values: Any) -> None:
    with _lock:
        _jobs[job_id]["clients"][code].update(values)


def _log(job_id: str, message: str) -> None:
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {message}"
    with _lock:
        logs = _jobs[job_id]["logs"]
        logs.append(line)
        if len(logs) > 500:
            del logs[:-500]


def get_scm_collection_status(job_id: str) -> dict[str, Any]:
    with _lock:
        job = _jobs.get(str(job_id))
        if not job:
            return {"ok": False, "message": "SCM 수집 작업을 찾을 수 없습니다."}
        return {"ok": True, "job": deepcopy(job)}

def submit_scm_collection_input(job_id: str, request_id: str, value: str) -> dict[str, Any]:
    with _lock:
        waiter = _input_waiters.get(str(job_id))
        if not waiter or waiter["request_id"] != str(request_id):
            return {"ok": False, "message": "이미 처리되었거나 유효하지 않은 인증 요청입니다."}
        code = str(value or "").strip()
        if not code: return {"ok": False, "message": "인증번호를 입력해주세요."}
        waiter["value"] = code; waiter["event"].set()
        return {"ok": True}

def _wait_for_input(job_id: str, title: str, prompt: str, timeout: int = 600) -> str:
    request_id = str(uuid.uuid4()); waiter = {"request_id":request_id,"event":threading.Event(),"value":""}
    with _lock:
        _input_waiters[job_id] = waiter
        _jobs[job_id]["input_request"] = {"request_id":request_id,"kind":"sms_code","title":title,"prompt":prompt}
        _jobs[job_id]["stage"] = "YES24 인증번호 입력 대기"
    if not waiter["event"].wait(timeout): raise TimeoutError("YES24 인증번호 입력 시간이 초과되었습니다.")
    with _lock:
        _input_waiters.pop(job_id, None); _jobs[job_id]["input_request"] = None
    return str(waiter["value"])


def _parse_date(value: Any) -> date:
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"날짜 형식은 YYYY-MM-DD여야 합니다: {value}") from exc


def _date_range(start: date, end: date) -> list[str]:
    if end < start:
        return []
    days = (end - start).days + 1
    return [(start + timedelta(days=offset)).isoformat() for offset in range(days)]


def _latest_dates(backend: Any, client_codes: list[str]) -> dict[str, str | None]:
    if not backend.db:
        raise RuntimeError("Supabase가 연결되지 않았습니다.")
    result: dict[str, str | None] = {}
    for code in client_codes:
        rows = (
            backend.db.client.table("SCM일별실판매")
            .select("판매일")
            .eq("거래처코드", code)
            .order("판매일", desc=True)
            .limit(1)
            .execute()
        ).data or []
        result[code] = str(rows[0]["판매일"]) if rows else None
    return result


def plan_scm_collection(backend: Any, options: dict[str, Any] | None = None) -> dict[str, Any]:
    options = options or {}
    requested = options.get("clients") or list(CLIENTS)
    requested_codes = {str(code).upper() for code in requested}
    client_codes = [code for code in COLLECTION_ORDER if code in requested_codes]
    if not client_codes:
        raise ValueError("수집할 거래처가 선택되지 않았습니다.")
    latest = _latest_dates(backend, client_codes)
    explicit_from = options.get("date_from")
    explicit_to = options.get("date_to")
    targets: dict[str, list[str]] = {}
    if explicit_from or explicit_to:
        if not explicit_from or not explicit_to:
            raise ValueError("재수집은 시작일과 종료일을 모두 입력해야 합니다.")
        start, end = _parse_date(explicit_from), _parse_date(explicit_to)
        dates = _date_range(start, end)
        if not dates:
            raise ValueError("재수집 종료일은 시작일보다 빠를 수 없습니다.")
        if len(dates) > MAX_RECOLLECT_DAYS:
            raise ValueError(f"한 번에 재수집할 수 있는 기간은 {MAX_RECOLLECT_DAYS}일입니다.")
        targets = {code: dates for code in client_codes}
        mode = "selected"
    else:
        yesterday = date.today() - timedelta(days=1)
        targets = {code: [yesterday.isoformat()] for code in client_codes}
        mode = "selected"
    return {"mode": mode, "latest": latest, "targets": targets, "clients": client_codes}


def start_scm_collection(backend: Any, options: dict[str, Any] | None = None) -> dict[str, Any]:
    options = options or {}
    with _lock:
        if any(job.get("state") in {"queued", "running"} for job in _jobs.values()):
            return {"ok": False, "message": "이미 SCM 데이터 수집이 진행 중입니다."}
    plan = plan_scm_collection(backend, options)
    plan["yes24_recipient_id"] = str(options.get("yes24_recipient_id") or "")
    account_map = {"KYOBO":["KYOBO"],"YPBOOKS":["YPBOOKS"],"YES24":["YES24_CHILD","YES24_ADULT"],"ALADIN":["ALADIN_CHILD","ALADIN_ADULT"]}
    for code in plan["clients"]:
        for key in account_map[code]: get_account(key)
    if "YES24" in plan["clients"]: get_recipient(plan["yes24_recipient_id"])
    if not any(plan["targets"].values()):
        return {"ok": True, "started": False, "message": "모든 거래처가 오늘 날짜까지 동기화되어 있습니다.", "plan": plan}
    job_id = str(uuid.uuid4())
    clients = {
        code: {
            "name": CLIENTS[code]["name"],
            "state": "waiting" if plan["targets"][code] else "skipped",
            "dates": plan["targets"][code],
            "rows": 0,
            "quantity": 0,
            "duration_seconds": 0.0,
            "error": "",
        }
        for code in plan["clients"]
    }
    with _lock:
        _jobs[job_id] = {
            "job_id": job_id,
            "state": "queued",
            "stage": "준비",
            "mode": plan["mode"],
            "started_at": _now(),
            "finished_at": None,
            "duration_seconds": 0.0,
            "clients": clients,
            "latest": plan["latest"],
            "logs": [],
            "error": "",
            "input_request": None,
        }
    threading.Thread(target=_run_collection, args=(job_id, backend, plan), daemon=True).start()
    return {"ok": True, "started": True, "job_id": job_id, "plan": plan}


def _configure_legacy(run_dir: Path, job_id: str, yes24_recipient_id: str = "") -> Any:
    from .scm_collectors import legacy_v72 as engine

    engine.SCRIPT_DIR = run_dir
    engine.BASE_DIR = SCM_ROOT
    engine.DOWNLOAD_DIR = YES24_DOWNLOAD_DIR
    engine.MASTER_DATA_DIR = run_dir
    original_log = engine.log

    def progress_log(message: Any = "") -> None:
        _log(job_id, str(message))
        original_log(message)

    engine.log = progress_log

    engine.read_kyobo_login_info = lambda: tuple(get_account("KYOBO")[x] for x in ("login_id", "password", "url"))
    engine.read_ypscm_login_info = lambda: tuple(get_account("YPBOOKS")[x] for x in ("login_id", "password", "url"))
    def yes_login(kind: str | None = None) -> dict[str, str]:
        row = get_account("YES24_CHILD" if kind == "child" else "YES24_ADULT")
        return {"id":row["login_id"],"password":row["password"],"url":row["url"],"row_text":kind or "YES24"}
    def aladin_login(kind: str) -> dict[str, str]:
        row = get_account("ALADIN_CHILD" if kind == "child" else "ALADIN_ADULT")
        return {"id":row["login_id"],"password":row["password"],"url":row["url"],"row_text":kind}
    engine.yes24_find_login_row = yes_login
    engine.aladin_find_login_row = aladin_login
    if yes24_recipient_id:
        recipient = get_recipient(yes24_recipient_id)
        popup_count = {"value": 0}
        def inline_yes24_input(title: str, prompt: str, default_value: str = "") -> str:
            popup_count["value"] += 1
            if popup_count["value"] % 2 == 1:
                progress_log(f"YES24 인증 수신자: {recipient['name']}")
                return recipient["phone"]
            progress_log(f"YES24 {recipient['name']}님에게 발송된 인증번호 입력 대기")
            masked = recipient['phone'][:3] + '-****-' + recipient['phone'][-4:]
            return _wait_for_input(job_id, title, f"{recipient['name']}님({masked})에게 발송된 인증번호를 입력해주세요.")
        engine.yes24_popup_input = inline_yes24_input
        engine.yes24_popup_info = lambda title, message: progress_log(f"{title}: {message}")

    def copy_sheet_without_excel(source_file: Path, target_file: Path, partner: dict[str, Any]) -> None:
        """COM/Excel 설치 없이 다운로드 원본 값을 임시 날짜 작업파일에 복사합니다."""
        try:
            with Path(source_file).open("rb") as stream:
                source_book = openpyxl.load_workbook(stream, read_only=True, data_only=False)
                try:
                    source_sheet = source_book.active
                    if source_sheet.calculate_dimension() in {"A1", "A1:A1"}:
                        source_sheet.reset_dimensions()
                    values = list(source_sheet.iter_rows(values_only=True))
                finally:
                    source_book.close()
        except Exception as exc:
            if "zip" not in str(exc).lower():
                raise
            import xlrd

            binary_book = xlrd.open_workbook(file_contents=Path(source_file).read_bytes())
            binary_sheet = binary_book.sheet_by_index(0)
            values = [tuple(binary_sheet.row_values(index)) for index in range(binary_sheet.nrows)]
        target_book = openpyxl.load_workbook(target_file)
        try:
            sheet_name = str(partner["sheet_hint"])
            if sheet_name in target_book.sheetnames:
                position = target_book.sheetnames.index(sheet_name)
                del target_book[sheet_name]
                target_sheet = target_book.create_sheet(sheet_name, position)
            else:
                target_sheet = target_book.create_sheet(sheet_name)
            for row in values:
                target_sheet.append(list(row))
            target_book.save(target_file)
        finally:
            target_book.close()
        progress_log(f"임시 작업파일 시트 반영 완료(openpyxl): {Path(target_file).name} / {sheet_name}")

    engine.copy_download_to_target_sheet = copy_sheet_without_excel
    engine.aladin_copy_xlsx_to_date_workbook = copy_sheet_without_excel

    def run_aladin_allow_empty_account(selected_dates: list[str], output_map: dict[str, Path]) -> None:
        completed = 0
        failures: list[str] = []
        for kind, label, folder in (
            ("child", "아이세움_아동", "aladin_download_child"),
            ("adult", "북폴리오_성인", "aladin_download_adlt"),
        ):
            try:
                engine.run_aladin_one_account(kind, label, folder, selected_dates, output_map)
                completed += 1
            except Exception as exc:
                if "No Aladin data rows extracted" in str(exc):
                    progress_log(f"알라딘 {label}: 선택 기간 판매 0건 (정상 빈 결과)")
                    continue
                failures.append(f"{label}: {exc}")
        if failures:
            raise RuntimeError("; ".join(failures))
        if completed == 0:
            raise RuntimeError("알라딘 선택 기간의 아동·성인 판매 데이터가 모두 0건입니다.")

    engine.run_aladin_batch = run_aladin_allow_empty_account
    return engine


def _filter_ledger(ledger: ScmLedger, client_code: str, requested_dates: list[str]) -> ScmLedger:
    rows = [row for row in ledger.rows if row["거래처코드"] == client_code and row["판매일"] in requested_dates]
    if not rows:
        raise RuntimeError(f"{CLIENTS[client_code]['name']} 수집 결과에 유효한 판매 데이터가 없습니다.")
    found_dates = {row["판매일"] for row in rows}
    missing_dates = sorted(set(requested_dates) - found_dates)
    if missing_dates:
        raise RuntimeError("판매 데이터가 없는 수집일: " + ", ".join(missing_dates))
    summary = {
        client_code: {
            "rows": len(rows),
            "quantity": sum(row["판매수량"] for row in rows),
            "unmatched": sum(1 for row in rows if not row.get("제품코드")),
        }
    }
    return ScmLedger(
        rows=rows,
        source_count=len(rows),
        collapsed_count=0,
        skipped_count=0,
        date_from=min(found_dates),
        date_to=max(found_dates),
        total_quantity=summary[client_code]["quantity"],
        client_summary=summary,
        product_codes={str(row["제품코드"]) for row in rows if row.get("제품코드")},
    )


def _filter_selected_scope(ledger: ScmLedger, client_code: str, requested_dates: list[str]) -> ScmLedger:
    """선택 범위를 보존하며 판매 0건 날짜도 정상적인 교체 범위로 반환합니다."""
    sale_date_key = "\ud310\ub9e4\uc77c"
    client_code_key = "\uac70\ub798\ucc98\ucf54\ub4dc"
    product_code_key = "\uc81c\ud488\ucf54\ub4dc"
    quantity_key = "\ud310\ub9e4\uc218\ub7c9"
    rows = [
        row for row in ledger.rows
        if row[client_code_key] == client_code and row[sale_date_key] in requested_dates
    ]
    summary = {
        client_code: {
            "rows": len(rows),
            "quantity": sum(row[quantity_key] for row in rows),
            "unmatched": sum(1 for row in rows if not row.get(product_code_key)),
        }
    }
    return ScmLedger(
        rows=rows,
        source_count=len(rows),
        collapsed_count=0,
        skipped_count=0,
        date_from=min(requested_dates),
        date_to=max(requested_dates),
        total_quantity=summary[client_code]["quantity"],
        client_summary=summary,
        product_codes={str(row[product_code_key]) for row in rows if row.get(product_code_key)},
    )


def _yes24_files(dates: list[str]) -> list[Path]:
    paths: list[Path] = []
    for value in dates:
        ymd = value.replace("-", "")
        for kind in ("아동", "성인"):
            for suffix in ("xls", "xlsx"):
                candidate = YES24_DOWNLOAD_DIR / f"{ymd}_예스24_{kind}.{suffix}"
                if candidate.exists():
                    paths.append(candidate)
                    break
    return paths


def _combined_hash(paths: list[Path]) -> str:
    digest = hashlib.sha256()
    for path in sorted(paths):
        digest.update(path.name.encode("utf-8"))
        with path.open("rb") as stream:
            for block in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(block)
    return digest.hexdigest()


def _verify_rows(backend: Any, ledger: ScmLedger, client_code: str) -> None:
    rows = (
        backend.db.client.table("SCM일별실판매")
        .select("판매일,거래처코드,ISBN13,판매수량")
        .eq("거래처코드", client_code)
        .gte("판매일", ledger.date_from)
        .lte("판매일", ledger.date_to)
        .execute()
    ).data or []
    expected = {(row["판매일"], row["ISBN13"]): row["판매수량"] for row in ledger.rows}
    actual = {(str(row["판매일"]), str(row["ISBN13"])): int(row["판매수량"]) for row in rows}
    if expected != actual:
        raise RuntimeError(
            f"DB 재검증 실패: 원천 {len(expected):,}건/{sum(expected.values()):,}부, "
            f"DB {len(actual):,}건/{sum(actual.values()):,}부"
        )


def _run_collection(job_id: str, backend: Any, plan: dict[str, Any]) -> None:
    started = time.perf_counter()
    run_dir = RUN_ROOT / job_id
    successful: list[str] = []
    try:
        _job_update(job_id, state="running", stage="작업파일 준비")
        run_dir.mkdir(parents=True, exist_ok=False)
        if not SCM_TEMPLATE.exists():
            raise FileNotFoundError(f"SCM 복사 양식을 찾을 수 없습니다: {SCM_TEMPLATE}")
        shutil.copy2(SCM_TEMPLATE, run_dir / SCM_TEMPLATE.name)
        engine = _configure_legacy(run_dir, job_id, plan.get("yes24_recipient_id", ""))
        all_dates = sorted({value for dates in plan["targets"].values() for value in dates})
        output_map = engine.create_date_workbooks(all_dates)

        for code in plan["clients"]:
            dates = plan["targets"][code]
            if not dates:
                continue
            client_started = time.perf_counter()
            _job_update(job_id, stage=f"{CLIENTS[code]['name']} 수집")
            _client_update(job_id, code, state="collecting")
            try:
                getattr(engine, CLIENTS[code]["runner"])(dates, output_map)
                _client_update(job_id, code, state="collected")
                successful.append(code)
            except Exception as exc:
                _log(job_id, f"{CLIENTS[code]['name']} 실패: {exc}")
                _client_update(job_id, code, state="failed", error=str(exc))
            finally:
                _client_update(job_id, code, duration_seconds=round(time.perf_counter() - client_started, 2))

        if not successful:
            raise RuntimeError("모든 거래처 수집이 실패했습니다.")

        _job_update(job_id, stage="정규화 및 Supabase 저장")
        parsed = parse_date_workbooks(output_map.values(), allow_empty=True)
        for code in successful:
            client_started = time.perf_counter()
            try:
                ledger = _filter_selected_scope(parsed, code, plan["targets"][code])
                demographics = Yes24Demographics([], 0, "", "", 0, 0)
                sources = [Path(output_map[value]) for value in plan["targets"][code]]
                if code == "YES24":
                    yes_files = _yes24_files(plan["targets"][code])
                    expected_file_count = len(plan["targets"][code]) * 2
                    if len(yes_files) != expected_file_count:
                        raise RuntimeError(f"YES24 구매자 원본이 부족합니다: {len(yes_files)}/{expected_file_count}개")
                    demographics = parse_yes24_demographic_files(yes_files)
                    sources = yes_files
                result = sync_scm_dataset(
                    backend,
                    ledger,
                    demographics,
                    ",".join(path.name for path in sources)[:500],
                    _combined_hash(sources),
                    [
                        {"\ud310\ub9e4\uc77c": sale_date, "\uac70\ub798\ucc98\ucf54\ub4dc": code}
                        for sale_date in plan["targets"][code]
                    ],
                )
                _verify_rows(backend, ledger, code)
                _client_update(
                    job_id,
                    code,
                    state="completed",
                    rows=result["total"],
                    quantity=result["quantity"],
                    sync_id=result["sync_id"],
                    duration_seconds=round(
                        _jobs[job_id]["clients"][code]["duration_seconds"] + time.perf_counter() - client_started,
                        2,
                    ),
                )
            except Exception as exc:
                _log(job_id, f"{CLIENTS[code]['name']} 저장/검증 실패: {exc}")
                _client_update(job_id, code, state="failed", error=str(exc))

        states = [values["state"] for values in _jobs[job_id]["clients"].values()]
        final_state = "completed" if all(state in {"completed", "skipped"} for state in states) else "partial"
        final_stage = "Supabase에 DB 등록을 완료했습니다." if final_state == "completed" else "일부 거래처 저장에 실패했습니다."
        _job_update(job_id, state=final_state, stage=final_stage, db_registered=final_state == "completed")
    except Exception as exc:
        _log(job_id, f"SCM 수집 작업 실패: {exc}")
        _job_update(job_id, state="failed", stage="실패", error=str(exc))
    finally:
        _job_update(
            job_id,
            finished_at=_now(),
            duration_seconds=round(time.perf_counter() - started, 2),
        )
