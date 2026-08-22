from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import openpyxl


TABLE_NAME = "ERP제품마스터"
PRODUCT_CODE_COLUMN = "제품코드"
EXPECTED_ROW_COUNT = 11_054
EXPECTED_COLUMN_COUNT = 34
DEFAULT_BATCH_SIZE = 500
DEFAULT_PAGE_SIZE = 1_000

SOURCE_COLUMNS = (
    "브랜드명", "계열명", "시리즈명", "품목", "품목명", "위탁구분", "수주유형",
    "초판발행일", "첫출고일", "첫출고일(제품)", "제품코드", "제품명", "약어", "정가",
    "구정가", "ISBN", "UCI코드", "학년", "학기", "과목", "손익항목코드", "손익항목명",
    "입력자사번", "입력자명", "자료입력일", "자료수정일", "제품상태", "과세구분",
    "아선방", "페이지수", "제본방식", "규격", "박스부수", "묶음부수",
)
DATE_COLUMNS = {"초판발행일", "첫출고일", "첫출고일(제품)", "자료입력일", "자료수정일"}
INTEGER_COLUMNS = {"정가", "구정가", "페이지수", "박스부수", "묶음부수"}


@dataclass(frozen=True)
class ExcelValidation:
    row_count: int
    column_count: int
    null_product_codes: int
    duplicate_product_codes: int
    unique_product_codes: int


@dataclass(frozen=True)
class ParsedProductMaster:
    path: Path
    rows: list[dict[str, Any]]
    validation: ExcelValidation

    def aggregate(self) -> dict[str, Any]:
        return aggregate_rows(self.rows)


@dataclass(frozen=True)
class ImportResult:
    source_rows: int
    inserted: int
    updated: int
    batches: int
    batch_size: int


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _text(value: Any) -> str | None:
    if _is_blank(value):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _date_text(value: Any, column: str, row_number: int) -> str | None:
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"{row_number}행 {column} 날짜값을 해석할 수 없습니다: {text!r}")


def _integer(value: Any, column: str, row_number: int) -> int | None:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        raise ValueError(f"{row_number}행 {column} 값이 숫자가 아닙니다: {value!r}")
    try:
        number = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{row_number}행 {column} 값이 숫자가 아닙니다: {value!r}") from exc
    if not number.is_integer():
        raise ValueError(f"{row_number}행 {column} 값이 정수가 아닙니다: {value!r}")
    return int(number)


def _normalise_row(values: tuple[Any, ...], row_number: int) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for column, value in zip(SOURCE_COLUMNS, values):
        if column in DATE_COLUMNS:
            row[column] = _date_text(value, column, row_number)
        elif column in INTEGER_COLUMNS:
            row[column] = _integer(value, column, row_number)
        else:
            row[column] = _text(value)
    return row


def parse_product_master_excel(
    path: str | Path,
    *,
    expected_rows: int = EXPECTED_ROW_COUNT,
    expected_columns: int = EXPECTED_COLUMN_COUNT,
) -> ParsedProductMaster:
    source_path = Path(path)
    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = tuple(_text(value) or "" for value in next(iterator))
        except StopIteration as exc:
            raise ValueError("Excel 파일에 헤더가 없습니다.") from exc

        if headers != SOURCE_COLUMNS:
            raise ValueError(
                "Excel 컬럼이 ERP제품마스터 34개 컬럼과 일치하지 않습니다. "
                f"확인된 컬럼 수: {len(headers)}"
            )

        rows: list[dict[str, Any]] = []
        raw_codes: list[str | None] = []
        for row_number, values in enumerate(iterator, start=2):
            if all(_is_blank(value) for value in values):
                continue
            if len(values) != expected_columns:
                raise ValueError(f"{row_number}행의 컬럼 수가 {expected_columns}개가 아닙니다.")
            row = _normalise_row(values, row_number)
            rows.append(row)
            raw_codes.append(row[PRODUCT_CODE_COLUMN])
    finally:
        workbook.close()

    valid_codes = [code for code in raw_codes if code]
    counts = Counter(valid_codes)
    validation = ExcelValidation(
        row_count=len(rows),
        column_count=len(headers),
        null_product_codes=len(raw_codes) - len(valid_codes),
        duplicate_product_codes=sum(count - 1 for count in counts.values() if count > 1),
        unique_product_codes=len(counts),
    )
    problems: list[str] = []
    if validation.row_count != expected_rows:
        problems.append(f"데이터 행 {validation.row_count}건(예상 {expected_rows}건)")
    if validation.column_count != expected_columns:
        problems.append(f"컬럼 {validation.column_count}개(예상 {expected_columns}개)")
    if validation.null_product_codes:
        problems.append(f"제품코드 NULL {validation.null_product_codes}건")
    if validation.duplicate_product_codes:
        problems.append(f"제품코드 중복 {validation.duplicate_product_codes}건")
    if validation.unique_product_codes != expected_rows:
        problems.append(f"제품코드 unique {validation.unique_product_codes}건(예상 {expected_rows}건)")
    if problems:
        raise ValueError("적재 전 검증 실패: " + ", ".join(problems))

    return ParsedProductMaster(source_path, rows, validation)


def _chunks(values: list[Any], size: int) -> Iterable[list[Any]]:
    for start in range(0, len(values), size):
        yield values[start:start + size]


def _existing_product_codes(client: Any, codes: list[str]) -> set[str]:
    existing: set[str] = set()
    for code_batch in _chunks(codes, 200):
        response = (
            client.table(TABLE_NAME)
            .select(PRODUCT_CODE_COLUMN)
            .in_(PRODUCT_CODE_COLUMN, code_batch)
            .execute()
        )
        existing.update(str(row[PRODUCT_CODE_COLUMN]) for row in (response.data or []))
    return existing


def upsert_product_master(
    client: Any,
    parsed: ParsedProductMaster,
    *,
    batch_size: int = DEFAULT_BATCH_SIZE,
) -> ImportResult:
    if batch_size < 1:
        raise ValueError("batch_size는 1 이상이어야 합니다.")
    codes = [str(row[PRODUCT_CODE_COLUMN]) for row in parsed.rows]
    existing = _existing_product_codes(client, codes)
    imported_at = datetime.now(timezone.utc).isoformat()
    payload = [
        {**row, "원본파일명": parsed.path.name, "적재일시": imported_at}
        for row in parsed.rows
    ]
    batches = 0
    for batch in _chunks(payload, batch_size):
        client.table(TABLE_NAME).upsert(batch, on_conflict=PRODUCT_CODE_COLUMN).execute()
        batches += 1
    return ImportResult(
        source_rows=len(payload),
        inserted=sum(1 for code in codes if code not in existing),
        updated=sum(1 for code in codes if code in existing),
        batches=batches,
        batch_size=batch_size,
    )


def fetch_all_product_master_rows(client: Any, *, page_size: int = DEFAULT_PAGE_SIZE) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for start in range(0, 10_000_000, page_size):
        batch = (
            client.table(TABLE_NAME)
            # `첫출고일(제품)`의 괄호를 PostgREST가 관계식으로 해석하지 않도록
            # 전체 행을 조회한 뒤 비교 단계에서 SOURCE_COLUMNS만 사용합니다.
            .select("*")
            .order(PRODUCT_CODE_COLUMN)
            .range(start, start + page_size - 1)
            .execute()
        ).data or []
        rows.extend(batch)
        if len(batch) < page_size:
            break
    return rows


def aggregate_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    codes = [row.get(PRODUCT_CODE_COLUMN) for row in rows]
    valid_codes = [str(code) for code in codes if code]
    code_counts = Counter(valid_codes)
    dates = {
        column: [str(row[column]) for row in rows if row.get(column)]
        for column in ("초판발행일", "첫출고일")
    }
    return {
        "total_rows": len(rows),
        "distinct_product_codes": len(code_counts),
        "null_product_codes": len(codes) - len(valid_codes),
        "duplicate_product_codes": sum(count - 1 for count in code_counts.values() if count > 1),
        "null_product_names": sum(1 for row in rows if not row.get("제품명")),
        "status_counts": dict(sorted(Counter(row.get("제품상태") for row in rows).items(), key=lambda item: str(item[0]))),
        "brand_counts": dict(sorted(Counter(row.get("브랜드명") for row in rows).items(), key=lambda item: str(item[0]))),
        "first_edition_date_min": min(dates["초판발행일"], default=None),
        "first_edition_date_max": max(dates["초판발행일"], default=None),
        "first_shipment_date_min": min(dates["첫출고일"], default=None),
        "first_shipment_date_max": max(dates["첫출고일"], default=None),
        "isbn_count": sum(1 for row in rows if row.get("ISBN")),
    }


def compare_source_and_database(parsed: ParsedProductMaster, database_rows: list[dict[str, Any]]) -> list[str]:
    source_by_code = {str(row[PRODUCT_CODE_COLUMN]): row for row in parsed.rows}
    database_by_code = {str(row[PRODUCT_CODE_COLUMN]): row for row in database_rows}
    errors: list[str] = []
    if source_by_code.keys() != database_by_code.keys():
        missing = sorted(source_by_code.keys() - database_by_code.keys())
        extra = sorted(database_by_code.keys() - source_by_code.keys())
        errors.append(f"제품코드 불일치: 누락 {len(missing)}건, 추가 {len(extra)}건")
    mismatch_count = sum(
        1 for code in source_by_code.keys() & database_by_code.keys()
        if any(source_by_code[code].get(column) != database_by_code[code].get(column) for column in SOURCE_COLUMNS)
    )
    if mismatch_count:
        errors.append(f"34개 원본 필드 값 불일치: {mismatch_count}건")
    return errors
