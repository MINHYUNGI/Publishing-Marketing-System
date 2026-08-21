from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl

from .config import SCM_LEDGER_FILE
from .yes24_demographics import Yes24Demographics, parse_yes24_demographics, preview_yes24_demographics


SYNC_VERSION = "SCM-SUPABASE-V1"
REQUIRED_COLUMNS = ("판매기준일", "거래처", "ISBN", "판매합계")
CLIENT_CODES = {
    "교보문고": "KYOBO",
    "영풍문고": "YPBOOKS",
    "예스24": "YES24",
    "YES24": "YES24",
    "알라딘": "ALADIN",
}
COMPARE_COLUMNS = ("제품코드", "판매수량", "원본파일명", "원본시트")


@dataclass(frozen=True)
class ScmLedger:
    rows: list[dict[str, Any]]
    source_count: int
    collapsed_count: int
    skipped_count: int
    date_from: str
    date_to: str
    total_quantity: int
    client_summary: dict[str, dict[str, int]]
    product_codes: set[str]


def _text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _code(value: Any) -> str:
    text = _text(value)
    return text[:-2] if text.endswith(".0") else text


def _isbn(value: Any) -> str:
    text = _code(value)
    return "".join(ch for ch in text if ch.isdigit())


def _date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value).replace("/", "-").replace(".", "-")
    parts = [part for part in text.split("-") if part]
    if len(parts) >= 3:
        try:
            return date(int(parts[0]), int(parts[1]), int(parts[2])).isoformat()
        except ValueError:
            return ""
    if len(text) == 8 and text.isdigit():
        try:
            return date(int(text[:4]), int(text[4:6]), int(text[6:])).isoformat()
        except ValueError:
            return ""
    return ""


def _integer(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(round(float(str(value).replace(",", ""))))
    except (TypeError, ValueError):
        return 0


def _chunks(values: list[dict[str, Any]], size: int = 500) -> Iterable[list[dict[str, Any]]]:
    for start in range(0, len(values), size):
        yield values[start : start + size]


def parse_scm_ledger(path: Path = SCM_LEDGER_FILE) -> ScmLedger:
    """기존 SCM 통합원장을 수정하지 않고 읽어 DB Grain으로 정규화합니다."""
    if not path.exists():
        raise FileNotFoundError(f"SCM 실판매 통합원장을 찾을 수 없습니다: {path}")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "통합원장" not in workbook.sheetnames:
            raise RuntimeError("SCM 통합원장에 '통합원장' 시트가 없습니다.")
        worksheet = workbook["통합원장"]
        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [_text(value) for value in (header_values or [])]
        missing = [column for column in REQUIRED_COLUMNS if column not in headers]
        if missing:
            raise RuntimeError("SCM 통합원장 필수 컬럼이 없습니다: " + ", ".join(missing))

        positions = {name: headers.index(name) for name in headers if name}
        normalized: dict[tuple[str, str, str], dict[str, Any]] = {}
        source_count = 0
        skipped_count = 0
        product_codes: set[str] = set()

        for values in worksheet.iter_rows(min_row=2, values_only=True):
            source_count += 1
            sale_date = _date_text(values[positions["판매기준일"]])
            client_name = _text(values[positions["거래처"]])
            client_code = CLIENT_CODES.get(client_name)
            isbn13 = _isbn(values[positions["ISBN"]])
            quantity = _integer(values[positions["판매합계"]])
            if not sale_date or not client_code or not isbn13 or quantity == 0:
                skipped_count += 1
                continue

            product_code = _code(values[positions["제품코드"]]) if "제품코드" in positions else ""
            product_name = ""
            for column in ("분석상품명", "공식상품명", "SCM상품명"):
                if column in positions:
                    product_name = _text(values[positions[column]])
                    if product_name:
                        break
            key = (sale_date, client_code, isbn13)
            row = {
                "판매일": sale_date,
                "거래처코드": client_code,
                "ISBN13": isbn13,
                "제품코드": product_code or None,
                "SCM상품명": product_name or None,
                "출판일자": (_date_text(values[positions["출판일자"]]) or None) if "출판일자" in positions else None,
                "판매수량": quantity,
                "원본파일명": _text(values[positions["원본파일"]]) if "원본파일" in positions else path.name,
                "원본시트": _text(values[positions["원본시트"]]) if "원본시트" in positions else None,
            }
            if key in normalized:
                normalized[key]["판매수량"] += quantity
            else:
                normalized[key] = row
            if product_code:
                product_codes.add(product_code)
    finally:
        workbook.close()

    rows = sorted(normalized.values(), key=lambda row: (row["판매일"], row["거래처코드"], row["ISBN13"]))
    if not rows:
        raise RuntimeError("SCM 통합원장에서 동기화할 판매 데이터가 없습니다.")
    client_summary: dict[str, dict[str, int]] = defaultdict(lambda: {"rows": 0, "quantity": 0, "unmatched": 0})
    for row in rows:
        summary = client_summary[row["거래처코드"]]
        summary["rows"] += 1
        summary["quantity"] += row["판매수량"]
        if not row["제품코드"]:
            summary["unmatched"] += 1

    return ScmLedger(
        rows=rows,
        source_count=source_count,
        collapsed_count=source_count - skipped_count - len(rows),
        skipped_count=skipped_count,
        date_from=rows[0]["판매일"],
        date_to=rows[-1]["판매일"],
        total_quantity=sum(row["판매수량"] for row in rows),
        client_summary=dict(client_summary),
        product_codes=product_codes,
    )


def _fetch_existing(client: Any, date_from: str, date_to: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    page_size = 1000
    offset = 0
    while True:
        page = (
            client.table("SCM일별실판매")
            .select("판매일,거래처코드,ISBN13,제품코드,판매수량,원본파일명,원본시트")
            .gte("판매일", date_from)
            .lte("판매일", date_to)
            .range(offset, offset + page_size - 1)
            .execute()
        ).data or []
        rows.extend(page)
        if len(page) < page_size:
            return rows
        offset += page_size


def _same(old: dict[str, Any], new: dict[str, Any]) -> bool:
    return all((old.get(column) or None) == (new.get(column) or None) for column in COMPARE_COLUMNS)


def preview_scm_sync(path: Path = SCM_LEDGER_FILE) -> dict[str, Any]:
    ledger = parse_scm_ledger(path)
    yes24 = preview_yes24_demographics()
    return {
        "ok": True,
        "source_file": str(path),
        "source_rows": ledger.source_count,
        "rows": len(ledger.rows),
        "collapsed": ledger.collapsed_count,
        "skipped": ledger.skipped_count,
        "date_from": ledger.date_from,
        "date_to": ledger.date_to,
        "total_quantity": ledger.total_quantity,
        "client_summary": ledger.client_summary,
        "missing_product_code_rows": sum(1 for row in ledger.rows if not row["제품코드"]),
        "yes24_demographics": yes24,
    }


def sync_scm_dataset(
    backend: Any,
    ledger: ScmLedger,
    yes24: Yes24Demographics,
    source_file: str,
    source_hash: str,
) -> dict[str, Any]:
    """정규화된 증분 범위를 스테이징한 뒤 DB 함수로 원자적으로 확정합니다."""
    if not backend.db:
        raise RuntimeError("Supabase가 연결되지 않았습니다.")
    client = backend.db.client
    isbn_values = sorted({row["ISBN13"] for row in ledger.rows})
    mapping_by_isbn: dict[str, str] = {}
    for isbn_batch in (isbn_values[start : start + 200] for start in range(0, len(isbn_values), 200)):
        mapped = (
            client.table("SCM제품매핑")
            .select("ISBN13,제품코드")
            .in_("ISBN13", isbn_batch)
            .execute()
        ).data or []
        mapping_by_isbn.update(
            {str(row["ISBN13"]): str(row["제품코드"]) for row in mapped if row.get("제품코드")}
        )
    for row in ledger.rows:
        if not row.get("제품코드"):
            row["제품코드"] = mapping_by_isbn.get(row["ISBN13"])
    valid_product_codes = {
        str(row.get("제품코드"))
        for row in backend.product_index
        if row.get("제품코드")
    }
    for row in ledger.rows:
        if row["제품코드"] and row["제품코드"] not in valid_product_codes:
            row["제품코드"] = None
    for summary in ledger.client_summary.values():
        summary["unmatched"] = 0
    for row in ledger.rows:
        if not row["제품코드"]:
            ledger.client_summary[row["거래처코드"]]["unmatched"] += 1

    existing = _fetch_existing(client, ledger.date_from, ledger.date_to)
    existing_map = {
        (str(row.get("판매일")), str(row.get("거래처코드")), str(row.get("ISBN13"))): row
        for row in existing
    }
    inserted = updated = unchanged = 0
    by_client = defaultdict(lambda: {"inserted": 0, "updated": 0, "unchanged": 0})
    for row in ledger.rows:
        key = (row["판매일"], row["거래처코드"], row["ISBN13"])
        old = existing_map.get(key)
        if old is None:
            inserted += 1
            by_client[row["거래처코드"]]["inserted"] += 1
        elif _same(old, row):
            unchanged += 1
            by_client[row["거래처코드"]]["unchanged"] += 1
        else:
            updated += 1
            by_client[row["거래처코드"]]["updated"] += 1

    history = {
        "대상시작일": ledger.date_from,
        "대상종료일": ledger.date_to,
        "원천파일명": source_file,
        "원천파일해시": source_hash,
        "상태": "진행",
        "전체건수": len(ledger.rows),
        "신규건수": inserted,
        "수정건수": updated,
        "동일건수": unchanged,
        "미매칭건수": sum(1 for row in ledger.rows if not row["제품코드"]),
        "YES24원본파일수": yes24.file_count,
        "YES24구매자스냅샷건수": len(yes24.rows),
        "YES24구매자분포건수": yes24.distribution_count,
        "원천판매수량합계": ledger.total_quantity,
        "프로그램버전": SYNC_VERSION,
    }
    created = client.table("SCM동기화이력").insert(history).execute().data or []
    if not created or not created[0].get("동기화ID"):
        raise RuntimeError("SCM 동기화 이력을 생성하지 못했습니다.")
    sync_id = created[0]["동기화ID"]

    try:
        staging_rows = [{"동기화ID": sync_id, **row} for row in ledger.rows]
        for batch in _chunks(staging_rows):
            client.table("SCM동기화스테이징").upsert(
                batch,
                on_conflict="동기화ID,판매일,거래처코드,ISBN13",
            ).execute()

        yes24_staging_rows = [{"동기화ID": sync_id, **row} for row in yes24.rows]
        for batch in _chunks(yes24_staging_rows):
            client.table("YES24구매자스테이징").upsert(
                batch,
                on_conflict="동기화ID,기준일,계정구분,ISBN13",
            ).execute()

        result = client.rpc("SCM동기화확정", {"대상동기화ID": sync_id}).execute().data or []
        applied = int((result[0] if result else {}).get("반영건수") or 0)
        db_quantity = int((result[0] if result else {}).get("판매수량합계") or 0)
        buyer_count = int((result[0] if result else {}).get("YES24스냅샷건수") or 0)
        distribution_count = int((result[0] if result else {}).get("YES24분포건수") or 0)
        if applied != len(ledger.rows) or db_quantity != ledger.total_quantity:
            raise RuntimeError(
                f"SCM 확정 검증 실패: 원천 {len(ledger.rows):,}건/{ledger.total_quantity:,}부, "
                f"DB {applied:,}건/{db_quantity:,}부"
            )
        if buyer_count != len(yes24.rows) or distribution_count != yes24.distribution_count:
            raise RuntimeError(
                f"YES24 구매자 검증 실패: 원천 {len(yes24.rows):,}건/{yes24.distribution_count:,}분포, "
                f"DB {buyer_count:,}건/{distribution_count:,}분포"
            )

        for client_code, summary in ledger.client_summary.items():
            counts = by_client[client_code]
            client.table("SCM동기화거래처결과").upsert(
                {
                    "동기화ID": sync_id,
                    "거래처코드": client_code,
                    "처리건수": summary["rows"],
                    "판매수량합계": summary["quantity"],
                    "신규건수": counts["inserted"],
                    "수정건수": counts["updated"],
                    "동일건수": counts["unchanged"],
                },
                on_conflict="동기화ID,거래처코드",
            ).execute()

        client.table("SCM동기화이력").update(
            {"상태": "성공", "종료일시": datetime.now().isoformat(), "DB판매수량합계": db_quantity}
        ).eq("동기화ID", sync_id).execute()
        return {
            "ok": True,
            "sync_id": sync_id,
            "file_name": source_file,
            "date_from": ledger.date_from,
            "date_to": ledger.date_to,
            "total": len(ledger.rows),
            "quantity": ledger.total_quantity,
            "inserted": inserted,
            "updated": updated,
            "unchanged": unchanged,
            "unmatched": history["미매칭건수"],
            "client_summary": ledger.client_summary,
            "yes24_demographics": {
                "files": yes24.file_count,
                "rows": len(yes24.rows),
                "distribution_rows": yes24.distribution_count,
                "quantity": yes24.total_quantity,
                "date_from": yes24.date_from,
                "date_to": yes24.date_to,
            },
        }
    except Exception as exc:
        client.table("SCM동기화이력").update(
            {"상태": "실패", "종료일시": datetime.now().isoformat(), "오류건수": 1, "오류내용": str(exc)}
        ).eq("동기화ID", sync_id).execute()
        raise


def sync_scm_ledger(backend: Any, path: Path = SCM_LEDGER_FILE) -> dict[str, Any]:
    """기존 전체 원장 수동 동기화 호환 진입점입니다."""
    from .security import sha256_file

    return sync_scm_dataset(
        backend,
        parse_scm_ledger(path),
        parse_yes24_demographics(),
        path.name,
        sha256_file(path),
    )
