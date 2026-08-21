from datetime import date, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook

from app.scm_collection import plan_scm_collection
from app.scm_collectors.workbook_parser import parse_date_workbooks
from tests.fakes import database_with


class ScmCollectionTests(unittest.TestCase):
    def test_date_workbook_parser_matches_four_client_rules(self):
        temporary = TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        path = Path(temporary.name) / "20260820.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)

        kyobo = workbook.create_sheet("\uad50\ubcf4_\uc544\ub3d9")
        kyobo.append(["\uad50\ubcf4\ubb38\uace0 \ud310\ub9e4\ucd94\uc774"])
        kyobo.append([])
        kyobo.append([])
        kyobo.append(["ISBN", "\uc0c1\ud488\uba85", "\ucd9c\ud310\uc77c\uc790", "\ud310\ub9e4\n(\uc601\uc5c5\uc810)", "\ud310\ub9e4\n(\uc628\ub77c\uc778)", "\ud310\ub9e4\n(\uc778\ud130\ud30c\ud06c)", "\ud310\ub9e4\n(\ubc95\uc778)"])
        kyobo.append(["9790000000001", "\uad50\ubcf4 \ub3c4\uc11c", "2026-01-01", 1, 2, 3, 99])

        ypbooks = workbook.create_sheet("\uc601\ud48d\ubb38\uace0")
        ypbooks.append(["\ubc14\ucf54\ub4dc", "\ub3c4\uc11c\uba85", "\ud310\ub9e4\uc218\ub7c9", "\ubc1c\ud589\uc77c"])
        ypbooks.append(["9790000000002", "\uc601\ud48d \ub3c4\uc11c", 4, "2026-01-02"])

        yes24 = workbook.create_sheet("\uc608\uc2a424_\uc544\ub3d9")
        yes24.append(["ISBN13", "\uc0c1\ud488\uba85", "\ucd1d\uacc4", "\ub0a8", "\ub140"])
        yes24.append(["", "\ud569\uacc4", 5, 2, 3])
        yes24.append(["9790000000003", "\uc608\uc2a4 \ub3c4\uc11c", 5, 2, 3])

        aladin = workbook.create_sheet("\uc54c\ub77c\ub518_\uc131\uc778")
        aladin.append(["\ub3c4\uc11c\uba85", "ISBN", "\ud310\ub9e4\uad8c\uc218"])
        aladin.append(["\uc54c\ub77c\ub518 \ub3c4\uc11c", "9790000000004", "6 \uad8c"])

        workbook.create_sheet("\uc608\uc2a424_\uc131\uc778")
        empty_aladin = workbook.create_sheet("\uc54c\ub77c\ub518_\uc544\ub3d9")
        empty_aladin.append(["\ucd9c\ud310\uc0ac"])
        workbook.save(path)

        parsed = parse_date_workbooks([path])
        quantities = {(row["\uac70\ub798\ucc98\ucf54\ub4dc"], row["ISBN13"]): row["\ud310\ub9e4\uc218\ub7c9"] for row in parsed.rows}
        self.assertEqual(quantities, {
            ("KYOBO", "9790000000001"): 6,
            ("YPBOOKS", "9790000000002"): 4,
            ("YES24", "9790000000003"): 5,
            ("ALADIN", "9790000000004"): 6,
        })
        self.assertEqual(parsed.total_quantity, 21)

    def test_incremental_collection_plans_each_client_after_its_latest_date(self):
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        today = date.today().isoformat()
        database = database_with({
            "SCM\uc77c\ubcc4\uc2e4\ud310\ub9e4": [
                {"\uac70\ub798\ucc98\ucf54\ub4dc": "KYOBO", "\ud310\ub9e4\uc77c": yesterday},
                {"\uac70\ub798\ucc98\ucf54\ub4dc": "YES24", "\ud310\ub9e4\uc77c": today},
            ]
        })
        backend = type("BackendStub", (), {"db": database})()
        plan = plan_scm_collection(backend, {"clients": ["KYOBO", "YES24"]})
        self.assertEqual(plan["targets"]["KYOBO"], [today])
        self.assertEqual(plan["targets"]["YES24"], [])

    def test_recollection_is_bounded_and_client_selective(self):
        database = database_with({"SCM\uc77c\ubcc4\uc2e4\ud310\ub9e4": []})
        backend = type("BackendStub", (), {"db": database})()
        plan = plan_scm_collection(
            backend,
            {"clients": ["ALADIN"], "date_from": "2026-08-18", "date_to": "2026-08-20"},
        )
        self.assertEqual(plan["mode"], "recollect")
        self.assertEqual(plan["targets"], {"ALADIN": ["2026-08-18", "2026-08-19", "2026-08-20"]})
